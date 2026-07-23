from __future__ import annotations

import hashlib
import json
import zipfile
from collections import Counter, defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from .certified_metrics import canonical_hash
from .models import ImportContract, RawImportBatch, ReferenceMappingSet, SourceSnapshot


class SalesLedgerRejected(ValueError):
    pass


def _validate_xlsx(uploaded) -> None:
    size = int(getattr(uploaded, "size", 0) or 0)
    if size <= 0 or size > 512 * 1024 * 1024:
        raise SalesLedgerRejected("XLSX 文件为空或超过 512MB 上限")
    uploaded.seek(0)
    if not zipfile.is_zipfile(uploaded):
        raise SalesLedgerRejected("销售明细账必须是有效 XLSX")
    uploaded.seek(0)
    with zipfile.ZipFile(uploaded) as archive:
        members = archive.infolist()
        total_uncompressed = sum(item.file_size for item in members)
        total_compressed = sum(max(item.compress_size, 1) for item in members)
        if len(members) > 20_000 or total_uncompressed > 2 * 1024 * 1024 * 1024:
            raise SalesLedgerRejected("XLSX 解压规模超过安全上限")
        if total_uncompressed / max(total_compressed, 1) > 200:
            raise SalesLedgerRejected("XLSX 压缩比异常")
    uploaded.seek(0)


def _normalize(value: Any) -> str:
    return " ".join(str(value or "").strip().split()).casefold()


def _decimal(value: Any, field: str) -> Decimal:
    try:
        result = Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, AttributeError) as exc:
        raise SalesLedgerRejected(f"{field} 不是有效十进制数") from exc
    if not result.is_finite() or abs(result) > Decimal("1000000000000000"):
        raise SalesLedgerRejected(f"{field} 超出允许范围")
    return result


def _date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _uploaded_hash(uploaded) -> str:
    digest = hashlib.sha256()
    uploaded.seek(0)
    for chunk in iter(lambda: uploaded.read(1024 * 1024), b""):
        digest.update(chunk)
    uploaded.seek(0)
    return f"sha256:{digest.hexdigest()}"


def _mapping_value(mapping_set: ReferenceMappingSet, raw: Any) -> Any:
    normalized = _normalize(raw)
    normalized_map = {_normalize(key): value for key, value in (mapping_set.mappings or {}).items()}
    return normalized_map.get(normalized)


def _require_confirmed(row, organization, label: str) -> None:
    if row is None or row.organization_id != organization.id:
        raise SalesLedgerRejected(f"{label} 不属于当前企业")
    if row.status != ImportContract.SignoffStatus.CONFIRMED:
        raise SalesLedgerRejected(f"{label} 尚未人工确认")


def _read_rows(uploaded, contract: ImportContract):
    schema = contract.schema or {}
    fields = schema.get("fields") or {}
    required = tuple(fields.values())
    if not required:
        raise SalesLedgerRejected("导入契约缺少字段白名单")
    uploaded.seek(0)
    workbook = load_workbook(uploaded, read_only=True, data_only=True)
    sheet_name = str(schema.get("sheet") or workbook.sheetnames[0])
    if sheet_name not in workbook.sheetnames:
        raise SalesLedgerRejected("工作表与导入契约不一致")
    try:
        sheet = workbook[sheet_name]
        iterator = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(iterator, ())]
        if not headers or len(set(headers)) != len(headers):
            raise SalesLedgerRejected("表头为空或存在重复列")
        missing = sorted(set(required) - set(headers))
        if missing:
            raise SalesLedgerRejected(f"缺少契约字段：{', '.join(missing)}")
        indexes = {canonical: headers.index(source) for canonical, source in fields.items()}
        for values in iterator:
            yield {canonical: values[index] if index < len(values) else None for canonical, index in indexes.items()}
    finally:
        workbook.close()


@transaction.atomic
def import_sales_ledger(*, organization, user, uploaded, manifest: dict, contract, channel_mapping, product_mapping):
    if contract.organization_id not in (None, organization.id):
        raise SalesLedgerRejected("导入契约不属于当前企业")
    if contract.signoff_status != ImportContract.SignoffStatus.CONFIRMED:
        raise SalesLedgerRejected("导入契约尚未人工确认")
    _require_confirmed(channel_mapping, organization, "渠道映射")
    _require_confirmed(product_mapping, organization, "商品映射")
    if manifest.get("schema") != "yiran_governed_raw_manifest_v1":
        raise SalesLedgerRejected("Raw Manifest schema 不受信任")
    _validate_xlsx(uploaded)
    manifest_hash = canonical_hash(manifest)
    content_hash = _uploaded_hash(uploaded)
    if str(manifest.get("content_sha256") or "") != content_hash:
        raise SalesLedgerRejected("上传内容与 Manifest hash 不一致")
    if str(manifest.get("contract_hash") or "") != contract.contract_hash:
        raise SalesLedgerRejected("Manifest 未绑定当前导入契约")

    schema = contract.schema or {}
    accepted_statuses = {_normalize(value) for value in schema.get("accepted_statuses", [])}
    accepted_currencies = {_normalize(value) for value in schema.get("accepted_currencies", [])}
    accepted_order_types = {_normalize(value) for value in schema.get("accepted_order_types", [])}
    brand_aliases = {_normalize(value) for value in schema.get("brand_aliases", [])}
    return_keywords = tuple(_normalize(value) for value in schema.get("return_type_keywords", []))
    window = manifest.get("window") or schema.get("window") or {}
    window_start = _date(window.get("start"))
    window_end = _date(window.get("end"))
    if not window_start or not window_end or window_start > window_end:
        raise SalesLedgerRejected("Manifest 缺少有效统计窗口")

    quarantines: Counter[str] = Counter()
    quality: Counter[str] = Counter()
    aggregates: dict[tuple[str, str], dict[str, Decimal]] = defaultdict(
        lambda: {"sales_qty": Decimal("0"), "sales_amount": Decimal("0")}
    )
    fingerprints: set[str] = set()
    actual_dates: list[date] = []
    row_count = 0
    accepted_count = 0

    for row in _read_rows(uploaded, contract):
        row_count += 1
        raw_date = _date(row.get("order_created_at"))
        if raw_date:
            actual_dates.append(raw_date)
        channel = _mapping_value(channel_mapping, row.get("channel"))
        product = _mapping_value(product_mapping, row.get("sku"))
        order_class = channel.get("order_class") if isinstance(channel, dict) else channel
        sku_id = product.get("sku_id") if isinstance(product, dict) else product
        common_reasons: list[str] = []
        if not order_class:
            common_reasons.append("unmapped_channel")
        if not sku_id:
            common_reasons.append("unmapped_sku")
        if accepted_statuses and _normalize(row.get("order_status")) not in accepted_statuses:
            common_reasons.append("status_not_allowlisted")
        order_type = _normalize(row.get("order_type"))
        if accepted_order_types and order_type not in accepted_order_types:
            common_reasons.append("order_type_not_allowlisted")
        if any(keyword and keyword in order_type for keyword in return_keywords):
            common_reasons.append("return_or_after_sales")
        if brand_aliases and _normalize(row.get("brand")) not in brand_aliases:
            common_reasons.append("brand_not_allowlisted")
        if raw_date is None or not window_start <= raw_date <= window_end:
            common_reasons.append("outside_or_invalid_date")
        m2_reasons = list(common_reasons)
        m3_reasons = list(common_reasons)
        if accepted_currencies and _normalize(row.get("currency")) not in accepted_currencies:
            m3_reasons.append("currency_not_allowlisted")
        try:
            qty = _decimal(row.get("qty"), "qty")
        except SalesLedgerRejected:
            m2_reasons.append("invalid_quantity")
            qty = Decimal("0")
        try:
            amount = _decimal(row.get("amount"), "amount")
        except SalesLedgerRejected:
            m3_reasons.append("invalid_amount")
            amount = Decimal("0")
        if qty < 0:
            m2_reasons.append("negative_quantity")
        if amount < 0:
            m3_reasons.append("negative_amount")

        fingerprint_payload = {
            "order": hashlib.sha256(str(row.get("order_id") or "").encode("utf-8")).hexdigest(),
            "sku": _normalize(row.get("sku")), "qty": str(qty), "amount": str(amount), "date": str(raw_date),
        }
        fingerprint = canonical_hash(fingerprint_payload)
        if fingerprint in fingerprints:
            quality["duplicate_fingerprint_suspect"] += 1
        fingerprints.add(fingerprint)
        key = (str(sku_id), str(order_class))
        if m2_reasons:
            quarantines.update(f"m2:{reason}" for reason in set(m2_reasons))
        else:
            aggregates[key]["sales_qty"] += qty
        if m3_reasons:
            quarantines.update(f"m3:{reason}" for reason in set(m3_reasons))
        else:
            aggregates[key]["sales_amount"] += amount
        if not m2_reasons and not m3_reasons:
            accepted_count += 1

    expected_rows = manifest.get("row_count")
    if expected_rows is not None and int(expected_rows) != row_count:
        raise SalesLedgerRejected("实际行数与 Manifest 不一致")
    actual_start = min(actual_dates) if actual_dates else None
    actual_end = max(actual_dates) if actual_dates else None
    boundary_covered = bool(actual_start and actual_end and actual_start <= window_start and actual_end >= window_end)
    safe_rows = [
        {"sku": sku, "order_class": order_class, "sales_qty": str(values["sales_qty"]),
         "sales_amount": str(values["sales_amount"])}
        for (sku, order_class), values in sorted(aggregates.items())
    ]
    receipt = {
        "schema": "yiran_sales_ledger_aggregate_receipt_v1",
        "classification": "aggregate_only_not_certified",
        "rows": safe_rows,
        "window": {"start": str(window_start), "end": str(window_end)},
        "external_write_performed": False,
    }
    status = RawImportBatch.Status.QUARANTINED
    return RawImportBatch.objects.create(
        organization=organization, dataset_type="sales_ledger", source_system="jackyun_sales_ledger_export",
        contract=contract, channel_mapping=channel_mapping, product_mapping=product_mapping, status=status,
        manifest_hash=manifest_hash, content_hash=content_hash,
        schema_version=str(manifest.get("schema_version") or "v1")[:32], window_start=window_start,
        window_end=window_end, actual_start=actual_start, actual_end=actual_end, boundary_covered=boundary_covered,
        source_complete=False, row_count=row_count, accepted_row_count=accepted_count,
        quarantine_summary=dict(quarantines), quality_signals=dict(quality), aggregate_receipt=receipt, created_by=user,
    )


@transaction.atomic
def reconcile_sales_ledger(*, batch: RawImportBatch, organization, user, reconciliation_hash: str):
    if batch.organization_id != organization.id:
        raise PermissionError("导入批次不属于当前企业")
    if batch.reconciliation_status == RawImportBatch.ReconciliationStatus.PASSED and batch.snapshot_id:
        return batch
    if not batch.boundary_covered:
        raise SalesLedgerRejected("数据未覆盖完整统计窗口")
    if any(int(value or 0) > 0 for value in (batch.quarantine_summary or {}).values()):
        raise SalesLedgerRejected("仍有隔离行未完成业务对账")
    if int((batch.quality_signals or {}).get("duplicate_fingerprint_suspect") or 0) > 0:
        raise SalesLedgerRejected("重复嫌疑尚未完成业务键对账")
    expected = canonical_hash({
        "manifest_hash": batch.manifest_hash,
        "content_hash": batch.content_hash,
        "receipt": batch.aggregate_receipt,
    })
    if reconciliation_hash != expected:
        raise SalesLedgerRejected("对账 hash 不匹配")
    payload = {"rows": (batch.aggregate_receipt or {}).get("rows", [])}
    snapshot = SourceSnapshot.objects.create(
        organization=organization, source_system="jackyun_sales_ledger_export", source_mode=SourceSnapshot.SourceMode.LIVE,
        scope={"dataset_type": "sales_ledger", "window_start": str(batch.window_start), "window_end": str(batch.window_end)},
        as_of=timezone.now(), complete=True,
        completeness={"pagination_complete": True, "external_reconciliation": True}, schema_version=batch.schema_version,
        row_count=len(payload["rows"]), content_hash=canonical_hash(payload), payload=payload,
        governance_status="governed", reconciliation_status="passed", manifest_hash=batch.manifest_hash,
        boundary_covered=True, source_complete=True, created_by=user,
    )
    batch.snapshot = snapshot
    batch.status = RawImportBatch.Status.READY
    batch.reconciliation_status = RawImportBatch.ReconciliationStatus.PASSED
    batch.source_complete = True
    batch.reconciled_by = user
    batch.reconciled_at = timezone.now()
    batch.save(update_fields=[
        "snapshot", "status", "reconciliation_status", "source_complete", "reconciled_by", "reconciled_at",
    ])
    return batch
