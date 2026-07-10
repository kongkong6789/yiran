"""
会议多格式产物生成:Markdown 方案 / HTML 图表分析报告 / Excel 指标表。
"""
from __future__ import annotations

import base64
import html
import io
from datetime import datetime

from . import knowledge, graph_knowledge
from .models import Meeting, Message


def _safe_filename(title: str, ext: str) -> str:
    base = "".join(c if c.isalnum() or c in "-_" else "_" for c in title[:40])
    return f"{base or 'meeting'}_{datetime.now().strftime('%Y%m%d')}.{ext}"


def _svg_bar_chart(metrics: list[dict], width: int = 560, height: int = 220) -> str:
    """简易 SVG 柱状图(指标值对比)。"""
    if not metrics:
        return '<p class="muted">暂无指标数据</p>'
    items = metrics[:8]
    vals = []
    for m in items:
        try:
            vals.append(float(m.get("value") or 0))
        except (TypeError, ValueError):
            vals.append(0.0)
    max_v = max(vals) or 1.0
    bar_w = max(24, (width - 40) // len(items) - 12)
    gap = 12
    bars = []
    labels = []
    for i, (m, v) in enumerate(zip(items, vals)):
        x = 30 + i * (bar_w + gap)
        h = int((v / max_v) * (height - 50))
        y = height - 30 - h
        name = str(m.get("metric_name") or m.get("metric") or f"M{i+1}")[:10]
        bars.append(
            f'<rect x="{x}" y="{y}" width="{bar_w}" height="{h}" rx="4" fill="#7d7aff"/>'
            f'<text x="{x + bar_w/2}" y="{y - 6}" text-anchor="middle" font-size="10" fill="#aab0d6">{v:.2g}</text>'
        )
        labels.append(
            f'<text x="{x + bar_w/2}" y="{height - 8}" text-anchor="middle" font-size="9" fill="#8b90ad">{html.escape(name)}</text>'
        )
    return (
        f'<svg width="{width}" height="{height}" viewBox="0 0 {width} {height}">'
        + "".join(bars) + "".join(labels) + "</svg>"
    )


def build_html_report(m: Meeting, md_plan: str, agent_points: list[str]) -> str:
    """自包含 HTML 分析报告(含 SVG 图表),可直接浏览器打开。"""
    metrics = knowledge.fetch_metrics_rows()
    anomalies = knowledge.fetch_anomaly_rows()
    refs = graph_knowledge.search_graph(m.question).get("refs", [])
    participants = ", ".join(f"{a.emoji}{a.name}" for a in m.participants.all())

    anomaly_rows = ""
    for a in anomalies[:8]:
        anomaly_rows += (
            f"<tr><td>{html.escape(str(a.get('dt', '')))}</td>"
            f"<td><span class='tag {a.get('level', '')}'>{html.escape(str(a.get('level', '')))}</span></td>"
            f"<td>{html.escape(str(a.get('metric_name', '')))}</td>"
            f"<td>{html.escape(str(a.get('dim_value', '')))}</td>"
            f"<td>{html.escape(str(a.get('detail', '')))}</td></tr>"
        )

    ref_tags = "".join(
        f"<span class='chip'>{html.escape(r.get('otype', ''))} · {html.escape(r.get('name', ''))}</span>"
        for r in refs
    ) or "<span class='muted'>未匹配到图谱实体</span>"

    highlights = "".join(
        f"<li><b>{html.escape(p.split(':', 1)[0] if ':' in p else 'Agent')}</b>: "
        f"{html.escape(p.split(':', 1)[-1][:120])}</li>"
        for p in agent_points[-8:]
    )

    metric_table = ""
    for r in metrics[:12]:
        mom = r.get("mom")
        mom_s = f"{mom:+.1%}" if isinstance(mom, (int, float)) else str(mom or "-")
        metric_table += (
            f"<tr><td>{html.escape(str(r.get('metric_name', '')))}</td>"
            f"<td>{html.escape(str(r.get('dim_value', '')))}</td>"
            f"<td>{r.get('value', '')}</td><td>{mom_s}</td>"
            f"<td>{html.escape(str(r.get('unit') or ''))}</td></tr>"
        )

    plan_html = html.escape(md_plan).replace("\n", "<br/>")

    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="utf-8"/>
<title>{html.escape(m.title)} · 分析报告</title>
<style>
  body {{ font-family: -apple-system, "Segoe UI", sans-serif; background: #0a0c18; color: #e7e9f3; margin: 0; padding: 24px; }}
  .wrap {{ max-width: 920px; margin: 0 auto; }}
  h1 {{ font-size: 22px; margin: 0 0 8px; }}
  h2 {{ font-size: 15px; color: #c4b5fd; border-bottom: 1px solid #2a2e42; padding-bottom: 6px; margin-top: 28px; }}
  .meta {{ color: #8b90ad; font-size: 13px; margin-bottom: 20px; }}
  .card {{ background: #12151f; border: 1px solid #2a2e42; border-radius: 10px; padding: 16px; margin: 12px 0; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  th, td {{ border-bottom: 1px solid #232636; padding: 8px 10px; text-align: left; }}
  th {{ color: #8b90ad; font-weight: 500; }}
  .chip {{ display: inline-block; background: #1e2240; border: 1px solid #3d4270; border-radius: 6px; padding: 2px 8px; margin: 2px; font-size: 12px; }}
  .tag {{ padding: 2px 6px; border-radius: 4px; font-size: 11px; }}
  .tag.high, .tag.critical {{ background: #3d1f2a; color: #ff8fa3; }}
  .tag.medium, .tag.warn {{ background: #3d321f; color: #ffd77a; }}
  .muted {{ color: #8b90ad; }}
  ul {{ margin: 8px 0; padding-left: 20px; line-height: 1.7; }}
  .plan {{ background: #0e1018; padding: 12px; border-radius: 8px; font-size: 14px; line-height: 1.65; }}
</style>
</head>
<body>
<div class="wrap">
  <h1>圆桌会议分析报告</h1>
  <div class="meta">
    问题:{html.escape(m.question)}<br/>
    参会:{html.escape(participants)} · 轮次:{m.round} · 生成:{datetime.now().strftime('%Y-%m-%d %H:%M')}
  </div>

  <h2>核心结论摘要</h2>
  <div class="card"><div class="plan">{html.escape(m.context_summary or '(暂无纪要)')}</div></div>

  <h2>关键指标图表</h2>
  <div class="card">{_svg_bar_chart(metrics)}</div>

  <h2>指标明细</h2>
  <div class="card"><table>
    <tr><th>指标</th><th>维度</th><th>值</th><th>环比</th><th>单位</th></tr>
    {metric_table or '<tr><td colspan="5" class="muted">暂无数据</td></tr>'}
  </table></div>

  <h2>异常预警</h2>
  <div class="card"><table>
    <tr><th>日期</th><th>级别</th><th>指标</th><th>对象</th><th>详情</th></tr>
    {anomaly_rows or '<tr><td colspan="5" class="muted">暂无异常</td></tr>'}
  </table></div>

  <h2>图谱引用实体</h2>
  <div class="card">{ref_tags}</div>

  <h2>讨论要点</h2>
  <div class="card"><ul>{highlights or '<li class="muted">暂无发言</li>'}</ul></div>

  <h2>执行方案(Markdown)</h2>
  <div class="card plan">{plan_html}</div>
</div>
</body>
</html>"""


def build_xlsx_bytes(m: Meeting, agent_messages: list[Message]) -> bytes:
    """Excel 指标与会议记录(多 Sheet)。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4E4AFF")

    # Sheet1: 指标快照
    ws1 = wb.active
    ws1.title = "指标快照"
    cols1 = ["日期", "指标", "维度类型", "维度值", "值", "环比", "同比", "单位", "口径"]
    ws1.append(cols1)
    for c in range(1, len(cols1) + 1):
        cell = ws1.cell(1, c)
        cell.font = header_font
        cell.fill = header_fill
    for r in knowledge.fetch_metrics_rows():
        mom = r.get("mom")
        yoy = r.get("yoy")
        ws1.append([
            str(r.get("dt", "")),
            r.get("metric_name", r.get("metric", "")),
            r.get("dim_type", ""),
            r.get("dim_value", r.get("dim", "")),
            r.get("value"),
            f"{mom:+.1%}" if isinstance(mom, (int, float)) else mom,
            f"{yoy:+.1%}" if isinstance(yoy, (int, float)) else yoy,
            r.get("unit", ""),
            r.get("formula", ""),
        ])

    # Sheet2: 异常预警
    ws2 = wb.create_sheet("异常预警")
    cols2 = ["日期", "级别", "指标", "对象", "规则", "详情"]
    ws2.append(cols2)
    for c in range(1, len(cols2) + 1):
        ws2.cell(1, c).font = header_font
        ws2.cell(1, c).fill = header_fill
    for a in knowledge.fetch_anomaly_rows():
        ws2.append([
            str(a.get("dt", "")),
            a.get("level", ""),
            a.get("metric_name", a.get("metric", "")),
            a.get("dim_value", a.get("scope", "")),
            a.get("rule", ""),
            a.get("detail", ""),
        ])

    # Sheet3: 会议摘要
    ws3 = wb.create_sheet("会议摘要")
    ws3["A1"] = "字段"
    ws3["B1"] = "内容"
    ws3["A1"].font = header_font
    ws3["B1"].font = header_font
    summary = [
        ("会议标题", m.title),
        ("核心问题", m.question),
        ("轮次", m.round),
        ("状态", m.status),
        ("纪要摘要", m.context_summary or ""),
        ("参会人", ", ".join(a.name for a in m.participants.all())),
    ]
    for i, (k, v) in enumerate(summary, start=2):
        ws3.cell(i, 1, k)
        ws3.cell(i, 2, v)
        ws3.cell(i, 2).alignment = Alignment(wrap_text=True)

    # Sheet4: 发言记录
    ws4 = wb.create_sheet("发言记录")
    ws4.append(["轮次", "发言方", "角色", "内容", "时间"])
    for c in range(1, 6):
        ws4.cell(1, c).font = header_font
        ws4.cell(1, c).fill = header_fill
    for msg in agent_messages:
        ws4.append([
            msg.round,
            msg.speaker_name,
            msg.speaker_type,
            msg.content,
            msg.created_at.strftime("%Y-%m-%d %H:%M"),
        ])

    for ws in (ws1, ws2, ws4):
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
