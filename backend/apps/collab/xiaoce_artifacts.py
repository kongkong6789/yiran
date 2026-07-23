"""Small, safe file artifacts produced directly by Xiaoce requests.

This module deliberately does not depend on the chat provider.  It gives the
Xiaoce worker a deterministic way to honour a narrowly scoped request such as
"create an Excel with hello in it", even when the model does not have a file
tool available.
"""
from __future__ import annotations

import os
import posixpath
import re
import tempfile
import uuid
import zipfile
from pathlib import Path
from urllib.parse import urlsplit
from xml.etree import ElementTree

from apps.core.attachments import attachments_root, resolve_attachment_path

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
DEFAULT_XLSX_NAME = "xiaoce-export.xlsx"
MAX_CELL_TEXT_CHARS = 8_000
MAX_XLSX_SOURCE_BYTES = 20 * 1024 * 1024
MAX_XLSX_ZIP_ENTRIES = 512
MAX_XLSX_EXPANDED_BYTES = 96 * 1024 * 1024
MAX_XLSX_SINGLE_ENTRY_BYTES = 32 * 1024 * 1024
MAX_XLSX_WORKSHEETS = 64
MAX_XLSX_WORKSHEET_XML_BYTES = 32 * 1024 * 1024
MAX_XLSX_WORKSHEET_ROWS = 250_000
MAX_XLSX_WORKSHEET_CELLS = 250_000
MAX_XLSX_TOTAL_CELLS = 500_000
MAX_XLSX_DIMENSION_CELLS = 5_000_000
MAX_XLSX_WORKSHEET_XML_ELEMENTS = 1_000_000
MAX_XLSX_PACKAGE_XML_BYTES = 4 * 1024 * 1024
MAX_XLSX_PACKAGE_XML_ELEMENTS = 100_000

_WORKBOOK_PATH = "xl/workbook.xml"
_WORKBOOK_RELS_PATH = "xl/_rels/workbook.xml.rels"
_WORKBOOK_NAMESPACES = {
    "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "http://purl.oclc.org/ooxml/spreadsheetml/main",
}
_OFFICE_RELATIONSHIP_NAMESPACES = {
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "http://purl.oclc.org/ooxml/officeDocument/relationships",
}
_PACKAGE_RELATIONSHIP_NAMESPACES = {
    "http://schemas.openxmlformats.org/package/2006/relationships",
    "http://purl.oclc.org/ooxml/package/relationships",
}
_WORKSHEET_RELATIONSHIP_TYPES = {
    f"{namespace}/worksheet"
    for namespace in _OFFICE_RELATIONSHIP_NAMESPACES
}


class XlsxResourceLimitError(ValueError):
    """Raised before openpyxl when an uploaded workbook exceeds safe limits."""

# Require both a file-production verb and an explicit Excel-type target.  This
# avoids producing files for ordinary questions *about* an uploaded workbook.
_CLAUSE_GAP = r"[^。！？!?；;,，\n]"
_ZH_PRODUCTION_VERB = (
    r"(?:生成|产出|创建|制作|导出|输出|做|写)"
    r"(?!\s*(?:的|过的|出来的|好的))"
)
_ZH_POST_TARGET_VERB = (
    r"(?:生成|产出|创建|制作|导出|输出|写)"
    r"(?!\s*(?:的|过的|出来的|好的))"
)
_EXCEL_GENERATION_REQUEST = re.compile(
    r"(?:"
    rf"{_ZH_PRODUCTION_VERB}{_CLAUSE_GAP}{{0,32}}?(?:excel|xlsx|xls|电子表格)"
    rf"|(?:excel|xlsx|xls|电子表格){_CLAUSE_GAP}{{0,32}}?{_ZH_POST_TARGET_VERB}"
    rf"|(?:generate|produce|create|make|export|write){_CLAUSE_GAP}{{0,40}}?(?:excel|xlsx|xls|spreadsheet)"
    rf"|(?:excel|xlsx|xls|spreadsheet){_CLAUSE_GAP}{{0,40}}?(?:generate|produce|create|make|export|write)"
    r")",
    re.IGNORECASE | re.DOTALL,
)
_NEGATED_GENERATION_REQUEST = re.compile(
    r"(?:"
    r"(?:不要|别|无需|不用|不需要|禁止|停止|取消)(?:\s|再|继续|帮我|替我|给我){0,10}"
    r"(?:生成|产出|创建|制作|导出|输出|做|写)"
    r"|(?:为什么|为何|怎么会).{0,16}(?:(?<!能)不能|无法|没法|(?<!可)不可以|不支持|做不到)"
    r"|(?:(?<!能)不能|无法|没法|(?<!可)不可以|不允许|不支持|做不到).{0,16}"
    r"(?:生成|产出|创建|制作|导出|输出|做|写)"
    r"|\b(?:do\s+not|don't|dont|never|no\s+need\s+to)\b.{0,16}"
    r"\b(?:generate|produce|create|make|export|write)\b"
    r"|\bwhy\b.{0,24}\b(?:cannot|can't|cant|unable)\b"
    r"|\b(?:cannot|can't|cant|unable)\b.{0,24}"
    r"\b(?:generat|produc|creat|mak|export|writ)"
    r"|\b(?:should|must)\s+not\s+(?:be\s+)?"
    r"(?:generat|produc|creat|mak|export|writ)"
    r")",
    re.IGNORECASE | re.DOTALL,
)
_CLAUSE_BOUNDARY = re.compile(r"[。！？!?；;,，\n]")
_SOURCE_WORKBOOK_REQUEST = re.compile(
    r"(?:"
    r"(?:已有|现有|原(?:有)?|这个|该|当前|上传(?:的)?|附件(?:中|里)?的?)"
    r"[^。！？!?；;,，\n]{0,16}(?:excel|xlsx|工作簿|电子表格)"
    r"|(?:excel|xlsx|工作簿|电子表格)"
    r"[^。！？!?；;,，\n]{0,16}(?:已有|现有|原(?:有)?|这个|该|当前|上传(?:的)?|附件)"
    r"|\b(?:uploaded|existing|attached|original|this)\b"
    r"[^.!?;,\n]{0,24}\b(?:excel|xlsx|workbook|spreadsheet)\b"
    r"|\b(?:excel|xlsx|workbook|spreadsheet)\b"
    r"[^.!?;,\n]{0,24}\b(?:uploaded|existing|attached|original)\b"
    r")",
    re.IGNORECASE,
)

_CHINESE_CONTENT = re.compile(
    r"(?:内容|数据)(?:里面|里)?\s*(?:只)?\s*"
    r"(?:写入?|填写|输入)\s*(?:一个|一条|一行)?\s*(?P<value>.+)",
    re.IGNORECASE | re.DOTALL,
)
_A1_CONTENT = re.compile(
    r"(?:A\s*1|第一(?:个)?单元格)\s*(?:里|中)?\s*"
    r"(?:写(?:入|上)?(?:值)?|填写|输入|放入?|改(?:成|为)|设置(?:成|为))\s*(?P<value>.+)",
    re.IGNORECASE | re.DOTALL,
)
_ENGLISH_CONTENT = re.compile(
    r"(?:content|(?:cell\s*)?A\s*1)\s*(?:should\s*)?"
    r"(?:be|say|contain|read|write)\s*(?P<value>.+)",
    re.IGNORECASE | re.DOTALL,
)
_CHINESE_CELL_CONTENT = re.compile(
    r"(?:在\s*)?(?P<cell>[A-Z]{1,3}\s*[1-9]\d{0,6})(?:\s*单元格)?"
    r"\s*(?:里|中)?\s*(?:写(?:入|上)?(?:值)?|填写|输入|放入?|改(?:成|为)|设置(?:成|为))"
    r"\s*(?:一个|一条|一行)?\s*(?P<value>.+)",
    re.IGNORECASE | re.DOTALL,
)
_ENGLISH_CELL_CONTENT = re.compile(
    r"(?:cell\s*)?(?P<cell>[A-Z]{1,3}\s*[1-9]\d{0,6})\s*"
    r"(?:should\s*)?(?:be|say|contain|read|write|set\s+to)\s*(?P<value>.+)",
    re.IGNORECASE | re.DOTALL,
)
_FOLLOW_UP_INSTRUCTION = re.compile(
    r"(?:"
    r"\s*(?:[，,；;。]\s*)?(?:然后|随后|接着)"
    r"|\s*(?:[，,；;。]\s*)?并(?:且)?\s*(?:把|将)?\s*(?:生成的|这个|该)?\s*"
    r"文件\s*(?:发|发送|给|返回|回传|保存|下载|命名)"
    r"|\s*(?:[，,；;。]\s*)?并(?:且)?\s*(?:导出|输出|保存|下载)"
    r"|\s*(?:[，,；;。]\s*)?(?:文件名|档案名)\s*(?:为|叫|是|：|:)"
    r"|\s*(?:,\s*)?\bthen\b"
    r"|\s*(?:,\s*)?\band\s+(?:send|return|give|name|save|download|export)\b"
    r"(?:\s+the)?\s+(?:file|workbook|spreadsheet)"
    r"|\s*(?:,\s*)?\bfile\s*name\s*(?:is|should\s+be|:)"
    r")",
    re.IGNORECASE | re.DOTALL,
)
_ENGLISH_PUT_CONTENT = re.compile(
    r"(?:write|put|enter)\s+(?P<value>.+?)\s+"
    r"(?:in|into)\s+(?:the\s+)?(?:excel|xlsx|spreadsheet)",
    re.IGNORECASE | re.DOTALL,
)


def is_excel_generation_request(message: str | None) -> bool:
    """Whether *message* explicitly asks Xiaoce to create an Excel file."""
    text = str(message or "")
    for match in _EXCEL_GENERATION_REQUEST.finditer(text):
        preceding = list(_CLAUSE_BOUNDARY.finditer(text, 0, match.start()))
        clause_start = preceding[-1].end() if preceding else 0
        following = _CLAUSE_BOUNDARY.search(text, match.end())
        clause_end = following.start() if following else len(text)
        clause = text[clause_start:clause_end]
        if not _NEGATED_GENERATION_REQUEST.search(clause):
            return True
    return False


def _clean_requested_value(value: str) -> str:
    """Keep the requested payload compact and remove request-closing words."""
    value = str(value or "").strip()
    quote_pairs = {"\"": "\"", "'": "'", "“": "”", "‘": "’", "`": "`"}
    if value[:1] in quote_pairs:
        closing = quote_pairs[value[0]]
        closing_index = value.find(closing, 1)
        if closing_index >= 1:
            return value[1:closing_index][:MAX_CELL_TEXT_CHARS]
    value = _FOLLOW_UP_INSTRUCTION.split(value, maxsplit=1)[0]
    value = re.split(r"\s*(?:就好|即可|就可以|就行|就够了)(?:\s|[。！!？?，,、]|$).*", value, maxsplit=1)[0]
    value = value.strip().strip("\"'“”‘’` ")
    return value[:MAX_CELL_TEXT_CHARS]


def _column_index(column: str) -> int:
    value = 0
    for char in column.upper():
        if not "A" <= char <= "Z":
            raise ValueError("Excel 列地址无效")
        value = value * 26 + ord(char) - ord("A") + 1
    return value


def _coordinate_parts(value: str) -> tuple[int, int] | None:
    address = re.sub(r"[\s$]+", "", str(value or "")).upper()
    match = re.fullmatch(r"([A-Z]{1,3})([1-9]\d{0,6})", address)
    if not match:
        return None
    column = _column_index(match.group(1))
    row = int(match.group(2))
    if column > 16_384 or row > 1_048_576:
        return None
    return row, column


def _normalize_cell_address(value: str) -> str | None:
    address = re.sub(r"\s+", "", str(value or "")).upper()
    return address if _coordinate_parts(address) is not None else None


def extract_excel_cell_update(message: str | None) -> tuple[str, str] | None:
    """Extract a safe target cell and its explicit requested text value."""
    text = str(message or "")
    for pattern in (_CHINESE_CELL_CONTENT, _ENGLISH_CELL_CONTENT):
        match = pattern.search(text)
        if not match:
            continue
        cell = _normalize_cell_address(match.group("cell"))
        value = _clean_requested_value(match.group("value"))
        if cell and value:
            return cell, value
    return None


def extract_excel_a1_content(message: str | None) -> str | None:
    """Extract an explicitly requested A1 value, if the wording supplies one."""
    text = str(message or "")
    cell_update = extract_excel_cell_update(text)
    if cell_update is not None:
        return cell_update[1]
    for pattern in (_CHINESE_CONTENT, _A1_CONTENT, _ENGLISH_CONTENT, _ENGLISH_PUT_CONTENT):
        match = pattern.search(text)
        if match:
            value = _clean_requested_value(match.group("value"))
            if value:
                return value
    return None


def _safe_excel_text(value: str) -> str:
    """Prevent a user-provided string from becoming an Excel formula on open."""
    text = "".join(char for char in str(value or "") if char >= " " or char in "\n\t")
    text = text[:MAX_CELL_TEXT_CHARS]
    return f"'{text}" if text.startswith(("=", "+", "-", "@")) else text


def _safe_xlsx_name(filename: str | None) -> str:
    requested_name = Path(str(filename or DEFAULT_XLSX_NAME).replace("\\", "/")).name
    if not requested_name.lower().endswith(".xlsx"):
        requested_name = DEFAULT_XLSX_NAME
    safe_name = re.sub(r"[^A-Za-z0-9._-]+", "-", requested_name).strip(".-")
    if not safe_name or safe_name in {".", ".."}:
        safe_name = DEFAULT_XLSX_NAME
    if not safe_name.lower().endswith(".xlsx"):
        safe_name = f"{safe_name}.xlsx"
    return safe_name


def _save_workbook_artifact(*, user_id: int, workbook, filename: str) -> dict:
    safe_name = _safe_xlsx_name(filename)
    root = attachments_root(user_id)
    root.mkdir(parents=True, exist_ok=True)
    stored_id = f"{uuid.uuid4().hex}_{safe_name}"
    target = root / stored_id
    temporary_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="wb",
            suffix=".xlsx",
            prefix=".xiaoce-",
            dir=root,
            delete=False,
        ) as temporary:
            temporary_path = Path(temporary.name)
        workbook.save(temporary_path)
        os.replace(temporary_path, target)
        temporary_path = None
    finally:
        if temporary_path is not None:
            temporary_path.unlink(missing_ok=True)

    return {
        "id": stored_id,
        "name": safe_name,
        "size": target.stat().st_size,
        "mime": XLSX_MIME,
        "has_text": False,
        "is_image": False,
        "is_file": True,
    }


def create_excel_artifact(
    *,
    user_id: int,
    content: str,
    filename: str = DEFAULT_XLSX_NAME,
    cell: str = "A1",
) -> dict:
    """Write a one-cell workbook inside the current user's attachment root.

    The returned item intentionally uses the same local attachment metadata
    shape as ``run_chat(...)[\"generated_files\"]`` so that
    ``_xiaoce_output_attachments`` can expose it without special handling.
    """
    from openpyxl import Workbook

    safe_cell = _normalize_cell_address(cell)
    if safe_cell is None:
        raise ValueError("Excel 单元格地址无效")
    workbook = Workbook()
    try:
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet[safe_cell] = _safe_excel_text(content)
        return _save_workbook_artifact(
            user_id=user_id,
            workbook=workbook,
            filename=filename,
        )
    finally:
        workbook.close()


def _worksheet_dimension_area(reference: str) -> int:
    endpoints = str(reference or "").split(":")
    if len(endpoints) not in {1, 2}:
        raise XlsxResourceLimitError("xlsx 工作表维度格式无效")
    start = _coordinate_parts(endpoints[0])
    end = _coordinate_parts(endpoints[-1])
    if start is None or end is None:
        raise XlsxResourceLimitError("xlsx 工作表维度超出 Excel 范围")
    row_span = abs(end[0] - start[0]) + 1
    column_span = abs(end[1] - start[1]) + 1
    return row_span * column_span


class _LimitedZipReader:
    """Cap actual decompressed worksheet bytes even if ZIP metadata is forged."""

    def __init__(self, source, limit: int, error_message: str):
        self.source = source
        self.limit = limit
        self.error_message = error_message
        self.bytes_read = 0

    def read(self, size: int = -1) -> bytes:
        remaining_with_probe = self.limit - self.bytes_read + 1
        if remaining_with_probe <= 0:
            raise XlsxResourceLimitError(self.error_message)
        if size < 0 or size > remaining_with_probe:
            size = remaining_with_probe
        chunk = self.source.read(size)
        self.bytes_read += len(chunk)
        if self.bytes_read > self.limit:
            raise XlsxResourceLimitError(self.error_message)
        return chunk


def _split_xml_name(value: str) -> tuple[str, str]:
    text = str(value or "")
    if text.startswith("{") and "}" in text:
        namespace, local_name = text[1:].split("}", 1)
        return namespace, local_name
    return "", text


def _package_xml_events(
    archive: zipfile.ZipFile,
    info: zipfile.ZipInfo,
):
    with archive.open(info, "r") as source:
        reader = _LimitedZipReader(
            source,
            MAX_XLSX_PACKAGE_XML_BYTES,
            "xlsx 核心 XML 解压后过大",
        )
        element_count = 0
        for event, element in ElementTree.iterparse(
            reader,
            events=("start", "end"),
        ):
            if event == "start":
                element_count += 1
                if element_count > MAX_XLSX_PACKAGE_XML_ELEMENTS:
                    raise XlsxResourceLimitError("xlsx 核心 XML 元素数量过多")
            yield event, element


def _parse_workbook_sheet_ids(
    archive: zipfile.ZipFile,
    info: zipfile.ZipInfo,
) -> list[str]:
    sheet_ids: list[str] = []
    seen_ids: set[str] = set()
    stack: list[str] = []
    root_checked = False
    for event, element in _package_xml_events(archive, info):
        namespace, local_name = _split_xml_name(element.tag)
        if event == "start":
            if not root_checked:
                root_checked = True
                if local_name != "workbook" or namespace not in _WORKBOOK_NAMESPACES:
                    raise XlsxResourceLimitError("xlsx workbook.xml 格式无法识别")
            stack.append(local_name)
            if (
                local_name == "sheet"
                and len(stack) >= 2
                and stack[-2] == "sheets"
            ):
                relationship_ids = [
                    element.attrib.get(f"{{{relationship_namespace}}}id")
                    for relationship_namespace in _OFFICE_RELATIONSHIP_NAMESPACES
                    if element.attrib.get(f"{{{relationship_namespace}}}id")
                ]
                if len(relationship_ids) != 1:
                    raise XlsxResourceLimitError("xlsx 工作表关系 ID 无法识别")
                relationship_id = relationship_ids[0]
                if relationship_id in seen_ids:
                    raise XlsxResourceLimitError("xlsx 工作表关系 ID 重复")
                seen_ids.add(relationship_id)
                sheet_ids.append(relationship_id)
                if len(sheet_ids) > MAX_XLSX_WORKSHEETS:
                    raise XlsxResourceLimitError("xlsx 工作表数量过多")
        else:
            element.clear()
            if stack:
                stack.pop()
    if not root_checked or not sheet_ids:
        raise XlsxResourceLimitError("xlsx workbook.xml 未包含可识别工作表")
    return sheet_ids


def _parse_workbook_relationships(
    archive: zipfile.ZipFile,
    info: zipfile.ZipInfo,
) -> dict[str, tuple[str, str, str]]:
    relationships: dict[str, tuple[str, str, str]] = {}
    root_checked = False
    for event, element in _package_xml_events(archive, info):
        namespace, local_name = _split_xml_name(element.tag)
        if event == "start":
            if not root_checked:
                root_checked = True
                if (
                    local_name != "Relationships"
                    or namespace not in _PACKAGE_RELATIONSHIP_NAMESPACES
                ):
                    raise XlsxResourceLimitError(
                        "xlsx workbook 关系文件格式无法识别",
                    )
            if local_name != "Relationship":
                continue
            relationship_id = str(element.attrib.get("Id") or "")
            relationship_type = str(element.attrib.get("Type") or "")
            target = str(element.attrib.get("Target") or "")
            target_mode = str(element.attrib.get("TargetMode") or "")
            if not relationship_id or not relationship_type or not target:
                raise XlsxResourceLimitError("xlsx workbook 关系条目不完整")
            if relationship_id in relationships:
                raise XlsxResourceLimitError("xlsx workbook 关系 ID 重复")
            relationships[relationship_id] = (
                relationship_type,
                target,
                target_mode,
            )
        else:
            element.clear()
    if not root_checked:
        raise XlsxResourceLimitError("xlsx workbook 关系文件无法识别")
    return relationships


def _normalize_relationship_target(target: str) -> str:
    if not target or "\x00" in target or "\\" in target:
        raise XlsxResourceLimitError("xlsx 工作表关系目标路径无效")
    parsed = urlsplit(target)
    if parsed.scheme or parsed.netloc or parsed.query or parsed.fragment:
        raise XlsxResourceLimitError("xlsx 工作表关系目标必须位于文件包内")
    raw_path = parsed.path
    raw_parts = raw_path.lstrip("/").split("/")
    if not raw_path or any(part == ".." for part in raw_parts):
        raise XlsxResourceLimitError("xlsx 工作表关系目标存在路径穿越")
    if raw_path.startswith("/"):
        normalized = posixpath.normpath(raw_path.lstrip("/"))
    else:
        normalized = posixpath.normpath(
            posixpath.join(posixpath.dirname(_WORKBOOK_PATH), raw_path),
        )
    if (
        not normalized
        or normalized in {".", ".."}
        or normalized.startswith("../")
        or normalized.startswith("/")
    ):
        raise XlsxResourceLimitError("xlsx 工作表关系目标存在路径穿越")
    return normalized


def _resolve_worksheet_entries(
    archive: zipfile.ZipFile,
    entries_by_name: dict[str, zipfile.ZipInfo],
) -> list[zipfile.ZipInfo]:
    workbook_info = entries_by_name.get(_WORKBOOK_PATH)
    relationships_info = entries_by_name.get(_WORKBOOK_RELS_PATH)
    if workbook_info is None or relationships_info is None:
        raise XlsxResourceLimitError("xlsx 缺少 workbook 核心关系文件")

    sheet_ids = _parse_workbook_sheet_ids(archive, workbook_info)
    relationships = _parse_workbook_relationships(
        archive,
        relationships_info,
    )
    worksheet_relationship_ids = {
        relationship_id
        for relationship_id, (relationship_type, _, _) in relationships.items()
        if relationship_type in _WORKSHEET_RELATIONSHIP_TYPES
    }
    if worksheet_relationship_ids != set(sheet_ids):
        raise XlsxResourceLimitError("xlsx 包含未识别或游离的工作表关系")

    worksheet_entries: list[zipfile.ZipInfo] = []
    seen_targets: set[str] = set()
    for relationship_id in sheet_ids:
        relationship_type, target, target_mode = relationships[relationship_id]
        if relationship_type not in _WORKSHEET_RELATIONSHIP_TYPES:
            raise XlsxResourceLimitError("xlsx 工作表关系类型无法识别")
        if target_mode and target_mode.lower() != "internal":
            raise XlsxResourceLimitError("xlsx 工作表不允许使用外部关系")
        normalized_target = _normalize_relationship_target(target)
        dedupe_key = normalized_target.casefold()
        if dedupe_key in seen_targets:
            raise XlsxResourceLimitError("xlsx 多个工作表指向重复目标")
        seen_targets.add(dedupe_key)
        target_info = entries_by_name.get(normalized_target)
        if (
            target_info is None
            or target_info.is_dir()
            or not normalized_target.lower().endswith(".xml")
        ):
            raise XlsxResourceLimitError("xlsx 工作表关系目标无法识别")
        worksheet_entries.append(target_info)
    return worksheet_entries


def _inspect_worksheet_xml(archive: zipfile.ZipFile, info: zipfile.ZipInfo) -> int:
    cell_count = 0
    row_count = 0
    element_count = 0
    min_row = min_column = max_row = max_column = None
    root_checked = False

    with archive.open(info, "r") as source:
        reader = _LimitedZipReader(
            source,
            MAX_XLSX_WORKSHEET_XML_BYTES,
            "xlsx 工作表 XML 解压后过大",
        )
        for event, element in ElementTree.iterparse(reader, events=("start", "end")):
            if event == "start":
                element_count += 1
                if element_count > MAX_XLSX_WORKSHEET_XML_ELEMENTS:
                    raise XlsxResourceLimitError("xlsx 工作表 XML 元素数量过多")

                namespace, tag = _split_xml_name(element.tag)
                if not root_checked:
                    root_checked = True
                    if tag != "worksheet" or namespace not in _WORKBOOK_NAMESPACES:
                        raise XlsxResourceLimitError("xlsx 工作表关系目标无法识别")
                if tag == "dimension":
                    area = _worksheet_dimension_area(element.attrib.get("ref", ""))
                    if area > MAX_XLSX_DIMENSION_CELLS:
                        raise XlsxResourceLimitError("xlsx 工作表声明维度过大")
                elif tag == "row":
                    row_count += 1
                    if row_count > MAX_XLSX_WORKSHEET_ROWS:
                        raise XlsxResourceLimitError("xlsx 工作表行数过多")
                elif tag == "c":
                    cell_count += 1
                    if cell_count > MAX_XLSX_WORKSHEET_CELLS:
                        raise XlsxResourceLimitError("xlsx 单个工作表单元格数量过多")
                    coordinate = _coordinate_parts(element.attrib.get("r", ""))
                    if coordinate is not None:
                        row, column = coordinate
                        min_row = row if min_row is None else min(min_row, row)
                        min_column = column if min_column is None else min(min_column, column)
                        max_row = row if max_row is None else max(max_row, row)
                        max_column = column if max_column is None else max(max_column, column)
                        actual_area = (
                            (max_row - min_row + 1)
                            * (max_column - min_column + 1)
                        )
                        if actual_area > MAX_XLSX_DIMENSION_CELLS:
                            raise XlsxResourceLimitError("xlsx 工作表实际维度过大")
            else:
                element.clear()

    if not root_checked:
        raise XlsxResourceLimitError("xlsx 工作表 XML 为空")
    return cell_count


def _validate_xlsx_resource_limits(path: Path) -> None:
    """Reject resource-heavy XLSX packages before openpyxl parses them."""
    if path.stat().st_size > MAX_XLSX_SOURCE_BYTES:
        raise XlsxResourceLimitError("xlsx 原文件体积超过上限")

    with zipfile.ZipFile(path, "r") as archive:
        entries = archive.infolist()
        if len(entries) > MAX_XLSX_ZIP_ENTRIES:
            raise XlsxResourceLimitError("xlsx ZIP 条目数量过多")

        expanded_bytes = 0
        seen_names: set[str] = set()
        entries_by_name: dict[str, zipfile.ZipInfo] = {}
        for info in entries:
            raw_name = info.filename
            package_name = raw_name[:-1] if info.is_dir() else raw_name
            if (
                not package_name
                or "\x00" in package_name
                or "\\" in package_name
                or package_name.startswith("/")
                or any(
                    part in {"", ".", ".."}
                    for part in package_name.split("/")
                )
            ):
                raise XlsxResourceLimitError("xlsx ZIP 条目路径无效")
            normalized_name = posixpath.normpath(package_name)
            if (
                normalized_name in {".", ".."}
                or normalized_name.startswith("../")
                or normalized_name.startswith("/")
            ):
                raise XlsxResourceLimitError("xlsx ZIP 条目存在路径穿越")
            dedupe_name = normalized_name.casefold()
            if dedupe_name in seen_names:
                raise XlsxResourceLimitError("xlsx ZIP 包含重复条目")
            seen_names.add(dedupe_name)
            if info.flag_bits & 0x1:
                raise XlsxResourceLimitError("xlsx ZIP 不支持加密条目")
            if info.file_size > MAX_XLSX_SINGLE_ENTRY_BYTES:
                raise XlsxResourceLimitError("xlsx ZIP 单个条目解压后过大")
            expanded_bytes += info.file_size
            if expanded_bytes > MAX_XLSX_EXPANDED_BYTES:
                raise XlsxResourceLimitError("xlsx ZIP 解压后总体积过大")
            if not info.is_dir():
                entries_by_name[normalized_name] = info

        worksheet_entries = _resolve_worksheet_entries(
            archive,
            entries_by_name,
        )
        if len(worksheet_entries) > MAX_XLSX_WORKSHEETS:
            raise XlsxResourceLimitError("xlsx 工作表数量过多")

        total_cells = 0
        for info in worksheet_entries:
            if info.file_size > MAX_XLSX_WORKSHEET_XML_BYTES:
                raise XlsxResourceLimitError("xlsx 工作表 XML 解压后过大")
            total_cells += _inspect_worksheet_xml(archive, info)
            if total_cells > MAX_XLSX_TOTAL_CELLS:
                raise XlsxResourceLimitError("xlsx 全部工作表单元格数量过多")


def _source_xlsx(
    source_attachments: list[dict] | None,
    user_id: int,
) -> tuple[Path, str] | None:
    """Resolve the first current-user-owned .xlsx source attachment."""
    owned_root = attachments_root(user_id).resolve()
    for item in source_attachments or []:
        if not isinstance(item, dict):
            continue
        stored_id = str(item.get("id") or item.get("stored_id") or "").strip()
        path = resolve_attachment_path(user_id, stored_id)
        if path is None:
            continue
        try:
            resolved_path = path.resolve(strict=True)
        except (OSError, RuntimeError):
            continue
        if not resolved_path.is_relative_to(owned_root):
            continue
        display_name = Path(str(item.get("name") or path.name).replace("\\", "/")).name
        if path.suffix.lower() != ".xlsx" and not display_name.lower().endswith(".xlsx"):
            continue
        return resolved_path, display_name
    return None


def modify_source_excel_artifact(
    *,
    user_id: int,
    source_attachments: list[dict] | None,
    cell: str,
    content: str,
) -> dict:
    """Copy a user-owned source workbook, modify one cell, and export the copy."""
    source = _source_xlsx(source_attachments, user_id)
    if source is None:
        raise ValueError("未找到当前用户可用的 xlsx 原文件")
    source_path, source_name = source
    safe_cell = _normalize_cell_address(cell)
    if safe_cell is None:
        raise ValueError("Excel 单元格地址无效")

    _validate_xlsx_resource_limits(source_path)

    from openpyxl import load_workbook

    workbook = load_workbook(source_path, data_only=False, keep_links=True)
    try:
        workbook.active[safe_cell] = _safe_excel_text(content)
        output_name = f"{Path(source_name).stem}-edited.xlsx"
        return _save_workbook_artifact(
            user_id=user_id,
            workbook=workbook,
            filename=output_name,
        )
    finally:
        workbook.close()


def maybe_generate_excel_artifact(
    *,
    user_id: int,
    request_text: str | None,
    model_reply: str | None,
    source_attachments: list[dict] | None = None,
) -> list[dict]:
    """Generate an Excel attachment only for an explicit Excel request.

    An explicit requested value wins; otherwise the model response is written
    to A1 so normal Xiaoce answers can still be exported as requested.
    """
    if not is_excel_generation_request(request_text):
        return []
    cell_update = extract_excel_cell_update(request_text)
    explicit_content = cell_update[1] if cell_update else extract_excel_a1_content(request_text)
    content = explicit_content or str(model_reply or "").strip()
    if not content:
        return []
    cell = cell_update[0] if cell_update else "A1"
    if _SOURCE_WORKBOOK_REQUEST.search(str(request_text or "")) and cell_update is not None:
        return [
            modify_source_excel_artifact(
                user_id=user_id,
                source_attachments=source_attachments,
                cell=cell,
                content=content,
            ),
        ]
    return [create_excel_artifact(user_id=user_id, content=content, cell=cell)]


__all__ = [
    "DEFAULT_XLSX_NAME",
    "XLSX_MIME",
    "XlsxResourceLimitError",
    "create_excel_artifact",
    "extract_excel_a1_content",
    "extract_excel_cell_update",
    "is_excel_generation_request",
    "maybe_generate_excel_artifact",
    "modify_source_excel_artifact",
]
