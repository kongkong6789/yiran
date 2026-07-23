import io
import tempfile
import uuid
from pathlib import Path
from unittest.mock import patch

from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import override_settings
from openpyxl import Workbook, load_workbook
from rest_framework.test import APITestCase

from apps.collab import views
from apps.collab.mentions import get_xiaoce_bot_user
from apps.collab.models import CollabMessage, CollabParticipant, CollabRoom, XiaoceRun
from apps.collab.xiaoce_runs import cancel_xiaoce_run, complete_xiaoce_run
from apps.collab.tests.test_xiaoce_runs import prepared_skill
from apps.core.attachments import resolve_attachment_path
from apps.skills.models import SkillAsset, UserSkill


def xlsx_upload(name="补货计划.xlsx"):
    stream = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "补货计划"
    sheet.append(["SKU", "建议补货量"])
    sheet.append(["SKU-001", 120])
    workbook.save(stream)
    workbook.close()
    return SimpleUploadedFile(
        name,
        stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


class XiaoceApiTests(APITestCase):
    def setUp(self):
        realtime_patcher = patch("apps.collab.views.ws_push.publish_sync")
        realtime_patcher.start()
        self.addCleanup(realtime_patcher.stop)
        self.user = User.objects.create_user("api-owner", password="pw")
        self.other = User.objects.create_user("api-other", password="pw")
        self.colleague = User.objects.create_user("api-colleague", password="pw")
        self.staff = User.objects.create_user("api-staff", password="pw", is_staff=True)
        bot = get_xiaoce_bot_user()
        self.bot = bot
        self.room = CollabRoom.objects.create(created_by=self.user, room_kind="dm")
        CollabParticipant.objects.create(room=self.room, user=self.user)
        CollabParticipant.objects.create(room=self.room, user=bot)
        self.client.force_authenticate(self.user)

    def create_context_room(self, *, owner=None, title="旧任务：夏季上新"):
        owner = owner or self.user
        room = CollabRoom.objects.create(
            created_by=owner,
            room_kind="dm",
            title=title,
        )
        CollabParticipant.objects.create(room=room, user=owner)
        CollabParticipant.objects.create(room=room, user=self.bot)
        return room

    @property
    def messages_url(self):
        return f"/api/collab/rooms/{self.room.id}/messages/"

    @property
    def tasks_url(self):
        return "/api/collab/xiaoce-tasks/"

    def test_create_xiaoce_task_always_creates_an_independent_room(self):
        first = self.client.post(self.tasks_url, {}, format="json")
        second = self.client.post(self.tasks_url, {}, format="json")

        self.assertEqual(first.status_code, 201)
        self.assertEqual(second.status_code, 201)
        self.assertNotEqual(first.data["id"], second.data["id"])
        self.assertEqual(first.data["title"], "小策bot（新任务）")
        self.assertEqual(second.data["title"], "小策bot（新任务）")
        for payload in (first.data, second.data):
            self.assertEqual(payload["room_kind"], "dm")
            self.assertEqual(payload["display_title"], "小策bot（新任务）")
            self.assertEqual(len(payload["messages"]), 1)
            self.assertEqual(payload["messages"][0]["ai_kind"], "xiaoce")

    def test_create_xiaoce_task_trims_and_limits_custom_title(self):
        response = self.client.post(
            self.tasks_url,
            {"title": f"  {'GMV' * 60}  "},
            format="json",
        )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(len(response.data["title"]), 120)
        self.assertTrue(response.data["title"].startswith("GMV"))

    def test_room_list_uses_task_title_only_for_xiaoce_direct_messages(self):
        task = self.client.post(
            self.tasks_url,
            {"title": "小策bot（GMV运算处理任务）"},
            format="json",
        ).data
        normal = self.client.post(
            "/api/collab/rooms/",
            {"peer_username": self.colleague.username, "room_kind": "dm", "title": "内部标题"},
            format="json",
        ).data

        listed = self.client.get("/api/collab/rooms/").data["results"]
        by_id = {row["id"]: row for row in listed}
        self.assertEqual(by_id[task["id"]]["display_title"], "小策bot（GMV运算处理任务）")
        self.assertEqual(by_id[normal["id"]]["display_title"], self.colleague.username)

    def test_xiaoce_task_is_visible_only_to_its_owner(self):
        trigger = CollabMessage.objects.create(
            room=self.room,
            sender=self.user,
            content="long running task",
            msg_type="user",
        )
        run = XiaoceRun.objects.create(
            id=uuid.uuid4(),
            room=self.room,
            user=self.user,
            trigger_message=trigger,
        )

        owner_detail = self.client.get(f"/api/collab/rooms/{self.room.id}/")
        owner_list = self.client.get("/api/collab/rooms/")
        owner_room = next(
            row for row in owner_list.data["results"]
            if row["id"] == str(self.room.id)
        )
        self.assertEqual(owner_detail.status_code, 200)
        self.assertEqual(owner_detail.data["active_xiaoce_run"]["id"], str(run.id))
        self.assertEqual(owner_room["active_xiaoce_run"]["id"], str(run.id))

        # 管理员、普通外部用户，以及共享 bot 账号本身均不能读取。
        for outsider in (self.staff, self.other, self.bot):
            self.client.force_authenticate(outsider)
            self.assertEqual(
                self.client.get(f"/api/collab/rooms/{self.room.id}/").status_code,
                403,
            )
            self.assertEqual(
                self.client.get(f"/api/collab/rooms/{self.room.id}/messages/").status_code,
                403,
            )
            self.assertEqual(
                self.client.post(f"/api/collab/rooms/{self.room.id}/read/").status_code,
                403,
            )
            listed_ids = {
                row["id"]
                for row in self.client.get("/api/collab/rooms/").data["results"]
            }
            self.assertNotIn(str(self.room.id), listed_ids)
            searched_ids = {
                row["room"]["id"]
                for row in self.client.get(
                    "/api/collab/search/",
                    {"q": "long running task"},
                ).data["results"]
            }
            self.assertNotIn(str(self.room.id), searched_ids)
            unread_ids = {
                row["room_id"]
                for row in self.client.get("/api/collab/unread/").data["results"]
            }
            self.assertNotIn(str(self.room.id), unread_ids)

    def test_existing_room_endpoint_reuses_the_latest_xiaoce_task(self):
        first = self.client.post(self.tasks_url, {"title": "任务一"}, format="json").data
        second = self.client.post(self.tasks_url, {"title": "任务二"}, format="json").data

        opened = self.client.post(
            "/api/collab/rooms/",
            {"peer_username": "小策bot", "room_kind": "dm"},
            format="json",
        )

        self.assertEqual(opened.status_code, 200)
        self.assertNotEqual(first["id"], second["id"])
        self.assertEqual(opened.data["id"], second["id"])

    @patch("apps.collab.views.threading.Thread")
    def test_send_returns_progress_run_and_room_detail_recovers_it(self, thread_cls):
        run_id = uuid.uuid4()

        response = self.client.post(
            self.messages_url,
            {"content": "分析", "run_id": str(run_id)},
        )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["xiaoce_run"]["id"], str(run_id))
        self.assertEqual(
            response.data["xiaoce_run"]["progress_steps"][0]["code"],
            "understanding",
        )
        self.assertEqual(response.data["message"]["meta"]["run_id"], str(run_id))
        # Viewer-relative presence must not ride on a room-wide message event.
        self.assertNotIn("peer_online", response.data["room"])
        self.assertNotIn("online_count", response.data["room"])
        thread_cls.return_value.start.assert_called_once()
        thread_args = thread_cls.call_args.kwargs["args"]
        self.assertEqual(thread_args, (run_id,))

        detail = self.client.get(f"/api/collab/rooms/{self.room.id}/")
        self.assertEqual(detail.data["active_xiaoce_run"]["id"], str(run_id))
        self.assertTrue(detail.data["peer_online"])

        presence = self.client.get(
            "/api/collab/presence/",
            {"user_ids": str(self.bot.id)},
        )
        self.assertTrue(presence.data["users"][str(self.bot.id)]["online"])

    @patch("apps.collab.views.threading.Thread")
    def test_second_send_is_rejected_without_creating_an_orphan_message(self, _thread_cls):
        self.client.post(
            self.messages_url,
            {"content": "第一次", "run_id": str(uuid.uuid4())},
        )
        before = CollabMessage.objects.count()

        response = self.client.post(
            self.messages_url,
            {"content": "第二次", "run_id": str(uuid.uuid4())},
        )

        self.assertEqual(response.status_code, 409)
        self.assertEqual(CollabMessage.objects.count(), before)

    @patch("apps.collab.views.threading.Thread")
    def test_send_can_reference_one_owned_xiaoce_task(self, _thread_cls):
        context_room = self.create_context_room()
        CollabMessage.objects.create(
            room=context_room,
            sender=self.user,
            content="目标是八月上新 3 个 SKU",
            msg_type="user",
        )
        CollabMessage.objects.create(
            room=context_room,
            sender=self.bot,
            content="已确定先完成定价表",
            msg_type="ai",
            ai_kind="xiaoce",
        )

        response = self.client.post(
            self.messages_url,
            {
                "content": "@「旧任务：夏季上新」 继续排期",
                "run_id": str(uuid.uuid4()),
                "context_room_ids": [str(context_room.id)],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201)
        refs = response.data["message"]["meta"]["context_rooms"]
        self.assertEqual(refs[0]["id"], str(context_room.id))
        self.assertEqual(refs[0]["title"], "旧任务：夏季上新")
        self.assertEqual(refs[0]["message_count"], 2)
        self.assertEqual(
            refs[0]["last_message_id"],
            context_room.messages.order_by("-id").values_list("id", flat=True).first(),
        )
        saved_message = CollabMessage.objects.get(id=response.data["message"]["id"])
        self.assertEqual(views._xiaoce_trigger_prompt(saved_message), "继续排期")

    @patch("apps.collab.views.threading.Thread")
    def test_formdata_reference_payload_is_supported_for_file_sends(self, _thread_cls):
        context_room = self.create_context_room()

        response = self.client.post(
            self.messages_url,
            {
                "content": "继续",
                "run_id": str(uuid.uuid4()),
                "context_room_ids": f'["{context_room.id}"]',
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(
            response.data["message"]["meta"]["context_rooms"][0]["id"],
            str(context_room.id),
        )

    @patch("apps.collab.views.threading.Thread")
    def test_file_only_message_starts_a_xiaoce_run(self, thread_cls):
        run_id = uuid.uuid4()
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(CHAT_ATTACHMENTS_ROOT=Path(tmp)):
                response = self.client.post(
                    self.messages_url,
                    {"content": "", "run_id": str(run_id), "files": [xlsx_upload()]},
                    format="multipart",
                )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["xiaoce_run"]["id"], str(run_id))
        self.assertEqual(len(response.data["message"]["attachments"]), 1)
        self.assertTrue(response.data["message"]["attachments"][0]["has_text"])
        thread_cls.return_value.start.assert_called_once()

    @patch("apps.collab.views.threading.Thread")
    @patch("apps.core.agent_chat.run_chat")
    def test_worker_reads_excel_and_can_return_the_downloadable_file(
        self,
        run_chat,
        _thread_cls,
    ):
        run_id = uuid.uuid4()

        def answer(**kwargs):
            return {
                "ok": True,
                "reply": "已读取补货计划。",
                "attachments": kwargs["attachments"],
            }

        run_chat.side_effect = answer
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(CHAT_ATTACHMENTS_ROOT=Path(tmp)):
                upload = xlsx_upload()
                expected_bytes = upload.read()
                upload.seek(0)
                response = self.client.post(
                    self.messages_url,
                    {
                        "content": "请读取补货计划，并把文件返回给我",
                        "run_id": str(run_id),
                        "files": [upload],
                    },
                    format="multipart",
                )
                self.assertEqual(response.status_code, 201)

                views._run_xiaoce_reply_async(run_id)

                model_attachment = run_chat.call_args.kwargs["attachments"][0]
                self.assertIn("SKU-001", model_attachment["text"])
                self.assertIn("建议补货量", model_attachment["text"])

                run = XiaoceRun.objects.select_related("result_message").get(id=run_id)
                returned = run.result_message.attachments[0]
                self.assertEqual(returned["name"], "补货计划.xlsx")
                self.assertTrue(returned["url"].startswith("/api/collab/attachments/"))

                download = self.client.get(f'{returned["url"]}?download=1')
                self.assertEqual(download.status_code, 200)
                self.assertEqual(
                    b"".join(download.streaming_content),
                    expected_bytes,
                )

    @patch("apps.collab.views.threading.Thread")
    @patch("apps.core.agent_chat.run_chat")
    def test_regular_file_analysis_does_not_duplicate_the_input_attachment(
        self,
        run_chat,
        _thread_cls,
    ):
        run_id = uuid.uuid4()

        def answer(**kwargs):
            return {"ok": True, "reply": "已分析。", "attachments": kwargs["attachments"]}

        run_chat.side_effect = answer
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(CHAT_ATTACHMENTS_ROOT=Path(tmp)):
                response = self.client.post(
                    self.messages_url,
                    {
                        "content": "这里面什么内容",
                        "run_id": str(run_id),
                        "files": [xlsx_upload()],
                    },
                    format="multipart",
                )
                self.assertEqual(response.status_code, 201)
                views._run_xiaoce_reply_async(run_id)

        run = XiaoceRun.objects.select_related("result_message").get(id=run_id)
        self.assertEqual(run.result_message.attachments, [])

    def test_rejected_message_does_not_write_uploaded_attachment_bytes(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            with override_settings(CHAT_ATTACHMENTS_ROOT=root):
                response = self.client.post(
                    self.messages_url,
                    {
                        "content": "x" * 4001,
                        "run_id": str(uuid.uuid4()),
                        "files": [xlsx_upload("must-not-be-written.xlsx")],
                    },
                    format="multipart",
                )

                self.assertEqual(response.status_code, 400)
                self.assertFalse((root / str(self.user.id)).exists())

    def test_transaction_failure_cleans_uploaded_attachment_bytes(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            with (
                override_settings(CHAT_ATTACHMENTS_ROOT=root),
                patch(
                    "apps.collab.views.create_xiaoce_run",
                    side_effect=RuntimeError("database write failed"),
                ),
            ):
                with self.assertRaisesRegex(RuntimeError, "database write failed"):
                    self.client.post(
                        self.messages_url,
                        {
                            "content": "分析这个文件",
                            "run_id": str(uuid.uuid4()),
                            "files": [xlsx_upload("must-be-cleaned.xlsx")],
                        },
                        format="multipart",
                    )
                user_root = root / str(self.user.id)
                remaining = list(user_root.iterdir()) if user_root.exists() else []

        self.assertEqual(remaining, [])

    @patch("apps.core.agent_chat.run_chat")
    def test_attachment_followup_reuses_the_latest_uploaded_excel(self, run_chat):
        run_chat.return_value = {"ok": True, "reply": "已完成分析。"}
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(CHAT_ATTACHMENTS_ROOT=Path(tmp)):
                processed = views.process_uploaded_files([xlsx_upload()], self.user.id)
                CollabMessage.objects.create(
                    room=self.room,
                    sender=self.user,
                    content="这里面都有什么",
                    attachments=views.attachment_public_meta(processed),
                    msg_type="user",
                )
                CollabMessage.objects.create(
                    room=self.room,
                    sender=self.bot,
                    content="文件包含补货计划。",
                    msg_type="ai",
                    ai_kind="xiaoce",
                )
                trigger = CollabMessage.objects.create(
                    room=self.room,
                    sender=self.user,
                    content="读取这里面的数据给我做一个分析",
                    msg_type="user",
                )
                run = XiaoceRun.objects.create(
                    id=uuid.uuid4(),
                    room=self.room,
                    user=self.user,
                    trigger_message=trigger,
                )

                views._run_xiaoce_reply_async(run.id)

        inherited = run_chat.call_args.kwargs["attachments"]
        self.assertEqual(len(inherited), 1)
        self.assertIn("SKU-001", inherited[0]["text"])
        self.assertIn("建议补货量", inherited[0]["text"])

    @patch("apps.core.agent_chat.run_chat")
    def test_non_attachment_followup_does_not_inherit_an_uploaded_file(self, run_chat):
        run_chat.return_value = {"ok": True, "reply": "已完成。"}
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(CHAT_ATTACHMENTS_ROOT=Path(tmp)):
                uploaded = views.process_uploaded_files([xlsx_upload()], self.user.id)
                CollabMessage.objects.create(
                    room=self.room,
                    sender=self.user,
                    content="这是上周的补货表",
                    attachments=views.attachment_public_meta(uploaded),
                    msg_type="user",
                )
                trigger = CollabMessage.objects.create(
                    room=self.room,
                    sender=self.user,
                    content="帮我写一份今天的会议纪要",
                    msg_type="user",
                )
                run = XiaoceRun.objects.create(
                    id=uuid.uuid4(),
                    room=self.room,
                    user=self.user,
                    trigger_message=trigger,
                )

                views._run_xiaoce_reply_async(run.id)

        self.assertEqual(run_chat.call_args.kwargs["attachments"], [])

    @patch("apps.core.agent_chat.run_chat")
    def test_current_attachment_takes_priority_over_inherited_file(self, run_chat):
        run_chat.return_value = {"ok": True, "reply": "已完成。"}
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(CHAT_ATTACHMENTS_ROOT=Path(tmp)):
                previous = views.process_uploaded_files(
                    [xlsx_upload("previous.xlsx")], self.user.id,
                )
                current = views.process_uploaded_files(
                    [xlsx_upload("current.xlsx")], self.user.id,
                )
                CollabMessage.objects.create(
                    room=self.room,
                    sender=self.user,
                    content="先看这份旧表",
                    attachments=views.attachment_public_meta(previous),
                    msg_type="user",
                )
                trigger = CollabMessage.objects.create(
                    room=self.room,
                    sender=self.user,
                    content="读取这个表格并分析",
                    attachments=views.attachment_public_meta(current),
                    msg_type="user",
                )
                run = XiaoceRun.objects.create(
                    id=uuid.uuid4(),
                    room=self.room,
                    user=self.user,
                    trigger_message=trigger,
                )

                views._run_xiaoce_reply_async(run.id)

        attachments = run_chat.call_args.kwargs["attachments"]
        self.assertEqual([item["name"] for item in attachments], ["current.xlsx"])

    @patch("apps.core.agent_chat.run_chat")
    def test_attachment_followup_uses_requesters_latest_file_and_ignores_other_sources(
        self,
        run_chat,
    ):
        run_chat.return_value = {"ok": True, "reply": "已完成。"}
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(CHAT_ATTACHMENTS_ROOT=Path(tmp)):
                own_old = views.process_uploaded_files(
                    [xlsx_upload("own-old.xlsx")], self.user.id,
                )
                own_latest = views.process_uploaded_files(
                    [xlsx_upload("own-latest.xlsx")], self.user.id,
                )
                other_attachment = views.process_uploaded_files(
                    [xlsx_upload("other.xlsx")], self.user.id,
                )
                bot_attachment = views.process_uploaded_files(
                    [xlsx_upload("bot.xlsx")], self.user.id,
                )
                for sender, content, attachment in (
                    (self.user, "我的旧附件", own_old),
                    (self.user, "我的最新附件", own_latest),
                    (self.other, "其他用户的附件", other_attachment),
                    (self.bot, "bot 的附件", bot_attachment),
                ):
                    CollabMessage.objects.create(
                        room=self.room,
                        sender=sender,
                        content=content,
                        attachments=views.attachment_public_meta(attachment),
                        msg_type="user" if sender != self.bot else "ai",
                        ai_kind="reply" if sender == self.bot else "",
                    )
                trigger = CollabMessage.objects.create(
                    room=self.room,
                    sender=self.user,
                    content="读取这里面的数据给我做一个分析",
                    msg_type="user",
                )
                run = XiaoceRun.objects.create(
                    id=uuid.uuid4(),
                    room=self.room,
                    user=self.user,
                    trigger_message=trigger,
                )

                views._run_xiaoce_reply_async(run.id)

        attachments = run_chat.call_args.kwargs["attachments"]
        self.assertEqual([item["name"] for item in attachments], ["own-latest.xlsx"])

    @patch("apps.core.agent_chat.run_chat")
    def test_attachment_followup_can_read_a_file_generated_by_xiaoce(self, run_chat):
        run_chat.return_value = {"ok": True, "reply": "文件内容是 hello。"}
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(CHAT_ATTACHMENTS_ROOT=Path(tmp)):
                generated = views.process_uploaded_files(
                    [xlsx_upload("xiaoce-generated.xlsx")],
                    self.user.id,
                )
                CollabMessage.objects.create(
                    room=self.room,
                    sender=self.bot,
                    content="已生成文件。",
                    attachments=views.attachment_public_meta(generated),
                    msg_type="ai",
                    ai_kind="xiaoce",
                )
                trigger = CollabMessage.objects.create(
                    room=self.room,
                    sender=self.user,
                    content="读取这个文件并告诉我里面的数据",
                    msg_type="user",
                )
                run = XiaoceRun.objects.create(
                    id=uuid.uuid4(),
                    room=self.room,
                    user=self.user,
                    trigger_message=trigger,
                )

                views._run_xiaoce_reply_async(run.id)

        attachments = run_chat.call_args.kwargs["attachments"]
        self.assertEqual([item["name"] for item in attachments], ["xiaoce-generated.xlsx"])
        self.assertIn("SKU-001", attachments[0]["text"])
        run.refresh_from_db()
        self.assertEqual(run.result_message.attachments, [])

    @patch("apps.core.agent_chat.run_chat")
    def test_followup_can_modify_a_just_generated_html_file(self, run_chat):
        run_chat.return_value = {"ok": True, "reply": "已读取并准备修改。"}
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(CHAT_ATTACHMENTS_ROOT=Path(tmp)):
                generated = views.process_uploaded_files(
                    [
                        SimpleUploadedFile(
                            "report.html",
                            b"<h1>hello</h1>",
                            content_type="text/html",
                        ),
                    ],
                    self.user.id,
                )
                CollabMessage.objects.create(
                    room=self.room,
                    sender=self.bot,
                    content="已生成文件。",
                    attachments=views.attachment_public_meta(generated),
                    msg_type="ai",
                    ai_kind="xiaoce",
                )
                trigger = CollabMessage.objects.create(
                    room=self.room,
                    sender=self.user,
                    content="修改刚生成的 HTML 文件，把标题改成周报",
                    msg_type="user",
                )
                run = XiaoceRun.objects.create(
                    id=uuid.uuid4(),
                    room=self.room,
                    user=self.user,
                    trigger_message=trigger,
                )

                views._run_xiaoce_reply_async(run.id)

        attachments = run_chat.call_args.kwargs["attachments"]
        self.assertEqual([item["name"] for item in attachments], ["report.html"])
        self.assertIn("hello", attachments[0]["text"])

    @patch("apps.core.agent_chat.run_chat")
    def test_replying_to_an_attachment_beats_a_newer_file(self, run_chat):
        run_chat.return_value = {"ok": True, "reply": "已分析引用文件。"}
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(CHAT_ATTACHMENTS_ROOT=Path(tmp)):
                quoted_file = views.process_uploaded_files(
                    [xlsx_upload("quoted.xlsx")],
                    self.user.id,
                )
                newer_file = views.process_uploaded_files(
                    [xlsx_upload("newer.xlsx")],
                    self.user.id,
                )
                quoted = CollabMessage.objects.create(
                    room=self.room,
                    sender=self.user,
                    content="需要稍后分析的表",
                    attachments=views.attachment_public_meta(quoted_file),
                    msg_type="user",
                )
                CollabMessage.objects.create(
                    room=self.room,
                    sender=self.user,
                    content="另一份表",
                    attachments=views.attachment_public_meta(newer_file),
                    msg_type="user",
                )
                trigger = CollabMessage.objects.create(
                    room=self.room,
                    sender=self.user,
                    reply_to=quoted,
                    content="分析一下",
                    msg_type="user",
                )
                run = XiaoceRun.objects.create(
                    id=uuid.uuid4(),
                    room=self.room,
                    user=self.user,
                    trigger_message=trigger,
                )

                views._run_xiaoce_reply_async(run.id)

        attachments = run_chat.call_args.kwargs["attachments"]
        self.assertEqual([item["name"] for item in attachments], ["quoted.xlsx"])

    def test_generated_file_is_exposed_without_requiring_a_return_phrase(self):
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(CHAT_ATTACHMENTS_ROOT=Path(tmp)):
                root = Path(tmp) / str(self.user.id)
                root.mkdir(parents=True)
                (root / "generated_report.csv").write_bytes(b"sku,qty\nSKU-001,120\n")

                output = views._xiaoce_output_attachments(
                    {
                        "generated_files": [{
                            "id": "generated_report.csv",
                            "name": "补货分析.csv",
                            "mime": "text/csv",
                        }],
                    },
                    "生成补货分析报告",
                    self.user.id,
                )

        self.assertEqual(output[0]["name"], "补货分析.csv")
        self.assertEqual(output[0]["url"], "/api/collab/attachments/generated_report.csv/")

    def test_requested_files_take_priority_over_optional_images_at_attachment_limit(self):
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(CHAT_ATTACHMENTS_ROOT=Path(tmp)):
                root = Path(tmp) / str(self.user.id)
                root.mkdir(parents=True)
                requested = {
                    "id": "requested_report.pdf",
                    "name": "经营报告.pdf",
                    "mime": "application/pdf",
                }
                (root / requested["id"]).write_bytes(b"%PDF-safe")
                images = []
                for index in range(5):
                    item = {
                        "id": f"optional_{index}.png",
                        "name": f"optional-{index}.png",
                        "mime": "image/png",
                        "is_image": True,
                    }
                    (root / item["id"]).write_bytes(b"\x89PNG\r\n\x1a\n")
                    images.append(item)

                output = views._xiaoce_output_attachments(
                    {
                        "generated_images": images,
                        "generated_files": [requested],
                    },
                    "生成 PDF 报告",
                    self.user.id,
                )

        self.assertEqual(len(output), views.MAX_ATTACH_FILES)
        self.assertEqual(output[0]["name"], "经营报告.pdf")
        self.assertIn("optional-0.png", [item["name"] for item in output])
        self.assertNotIn("optional-4.png", [item["name"] for item in output])

    def test_generated_html_download_uses_explicit_storage_owner_and_safe_headers(self):
        stored_id = f"{uuid.uuid4().hex}_report.html"
        attachment = {
            "id": stored_id,
            "name": "report.html",
            "mime": "text/html",
            "storage_owner_id": self.user.id,
            "is_file": True,
            "is_image": False,
        }
        CollabMessage.objects.create(
            room=self.room,
            sender=self.bot,
            content="文件已生成",
            attachments=[attachment],
            msg_type="ai",
            ai_kind="xiaoce",
        )

        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(CHAT_ATTACHMENTS_ROOT=Path(tmp)):
                user_root = Path(tmp) / str(self.user.id)
                user_root.mkdir(parents=True)
                (user_root / stored_id).write_text("<h1>safe</h1>", encoding="utf-8")

                response = self.client.get(f"/api/collab/attachments/{stored_id}/")
                payload = b"".join(response.streaming_content)

                self.client.force_authenticate(self.other)
                forbidden = self.client.get(f"/api/collab/attachments/{stored_id}/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(payload, b"<h1>safe</h1>")
        self.assertEqual(response["X-Content-Type-Options"], "nosniff")
        self.assertEqual(response["Cache-Control"], "private, no-store")
        self.assertEqual(response["Content-Security-Policy"], "sandbox")
        self.assertIn("attachment;", response["Content-Disposition"])
        self.assertEqual(forbidden.status_code, 403)

    @patch("apps.core.agent_chat.run_chat")
    def test_worker_generates_and_returns_downloadable_excel_even_if_model_refuses(self, run_chat):
        run_chat.return_value = {
            "ok": True,
            "reply": "我无法直接生成并发送 Excel 文件。",
        }
        trigger = CollabMessage.objects.create(
            room=self.room,
            sender=self.user,
            content="可以帮我产出一个excel吗 内容里面写一个hello就好",
            msg_type="user",
        )
        run = XiaoceRun.objects.create(
            id=uuid.uuid4(),
            room=self.room,
            user=self.user,
            trigger_message=trigger,
        )

        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(CHAT_ATTACHMENTS_ROOT=Path(tmp)):
                views._run_xiaoce_reply_async(run.id)

                run.refresh_from_db()
                attachment = run.result_message.attachments[0]
                stored_path = Path(tmp) / str(self.user.id) / attachment["id"]
                workbook = load_workbook(stored_path, data_only=False)
                try:
                    self.assertEqual(workbook.active["A1"].value, "hello")
                finally:
                    workbook.close()

                download = self.client.get(f'{attachment["url"]}?download=1')
                downloaded = b"".join(download.streaming_content)

        self.assertEqual(
            run.result_message.content,
            "已生成文件：xiaoce-export.xlsx，可点击附件下载。",
        )
        self.assertEqual(len(run.result_message.attachments), 1)
        self.assertEqual(attachment["name"], "xiaoce-export.xlsx")
        self.assertEqual(download.status_code, 200)
        self.assertTrue(downloaded.startswith(b"PK"))
        run_chat.assert_not_called()

    @patch("apps.core.agent_chat.run_chat")
    def test_worker_generates_literal_markdown_without_calling_the_model(self, run_chat):
        trigger = CollabMessage.objects.create(
            room=self.room,
            sender=self.user,
            content="请生成一个 Markdown 文件，内容里面写一个 hello",
            msg_type="user",
        )
        run = XiaoceRun.objects.create(
            id=uuid.uuid4(),
            room=self.room,
            user=self.user,
            trigger_message=trigger,
        )

        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(CHAT_ATTACHMENTS_ROOT=Path(tmp)):
                views._run_xiaoce_reply_async(run.id)
                run.refresh_from_db()
                attachment = run.result_message.attachments[0]
                stored_path = resolve_attachment_path(self.user.id, attachment["id"])
                content = stored_path.read_text(encoding="utf-8")
                download = self.client.get(f'{attachment["url"]}?download=1')

        self.assertEqual(content, "hello\n")
        self.assertEqual(attachment["name"], "xiaoce-export.md")
        self.assertEqual(attachment["storage_owner_id"], self.user.id)
        self.assertEqual(download.status_code, 200)
        self.assertTrue(download["Content-Type"].startswith("text/markdown"))
        self.assertEqual(
            run.result_message.content,
            "已生成文件：xiaoce-export.md，可点击附件下载。",
        )
        run_chat.assert_not_called()

    @patch("apps.core.agent_chat.run_chat")
    def test_model_backed_file_generation_uses_the_trusted_renderer_contract(self, run_chat):
        run_chat.return_value = {
            "ok": True,
            "reply": "# 经营分析\n\n- GMV 稳定",
            "llm": True,
        }
        trigger = CollabMessage.objects.create(
            room=self.room,
            sender=self.user,
            content="请根据当前对话生成一个 Markdown 分析报告",
            msg_type="user",
        )
        run = XiaoceRun.objects.create(
            id=uuid.uuid4(),
            room=self.room,
            user=self.user,
            trigger_message=trigger,
        )

        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(CHAT_ATTACHMENTS_ROOT=Path(tmp)):
                views._run_xiaoce_reply_async(run.id)
                run.refresh_from_db()
                attachment = run.result_message.attachments[0]
                stored_path = resolve_attachment_path(self.user.id, attachment["id"])
                content = stored_path.read_text(encoding="utf-8")

        call = run_chat.call_args.kwargs
        self.assertEqual(call["max_tokens_floor"], 4200)
        self.assertEqual(
            call["internal_system_append"],
            views.XIAOCE_FILE_ARTIFACT_SYSTEM_APPEND,
        )
        self.assertIn("# 经营分析", content)
        self.assertEqual(
            run.result_message.content,
            "已生成文件：xiaoce-export.md，可点击附件下载。",
        )

    @patch("apps.core.agent_chat.run_chat")
    def test_worker_copies_and_updates_an_uploaded_excel(self, run_chat):
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(CHAT_ATTACHMENTS_ROOT=Path(tmp)):
                processed = views.process_uploaded_files([xlsx_upload("source.xlsx")], self.user.id)
                original_path = Path(processed[0]["stored_path"])
                trigger = CollabMessage.objects.create(
                    room=self.room,
                    sender=self.user,
                    content=(
                        "请在上传的 Excel 的 B2 单元格写入 hello，"
                        "然后把文件发给我，并导出 Excel"
                    ),
                    attachments=views.attachment_public_meta(processed),
                    msg_type="user",
                )
                run = XiaoceRun.objects.create(
                    id=uuid.uuid4(),
                    room=self.room,
                    user=self.user,
                    trigger_message=trigger,
                )

                views._run_xiaoce_reply_async(run.id)

                run.refresh_from_db()
                attachment = run.result_message.attachments[0]
                output_path = resolve_attachment_path(self.user.id, attachment["id"])
                output = load_workbook(output_path, data_only=False)
                original = load_workbook(original_path, data_only=False)
                try:
                    self.assertEqual(output.active["A1"].value, "SKU")
                    self.assertEqual(output.active["B2"].value, "hello")
                    self.assertEqual(original.active["B2"].value, 120)
                finally:
                    output.close()
                    original.close()

        self.assertEqual(run.status, XiaoceRun.Status.COMPLETED)
        self.assertEqual(
            run.result_message.content,
            "已生成文件：xiaoce-export.xlsx，可点击附件下载。",
        )
        run_chat.assert_not_called()

    @patch("apps.collab.views.maybe_generate_file_artifacts")
    @patch("apps.core.agent_chat.run_chat")
    def test_excel_generation_failure_does_not_fail_the_successful_chat(
        self,
        run_chat,
        generate_excel,
    ):
        generate_excel.side_effect = OSError("disk full")
        run_chat.return_value = {"ok": True, "reply": "这是模型生成的正文。"}
        trigger = CollabMessage.objects.create(
            room=self.room,
            sender=self.user,
            content="把这段内容导出 Excel",
            msg_type="user",
        )
        run = XiaoceRun.objects.create(
            id=uuid.uuid4(),
            room=self.room,
            user=self.user,
            trigger_message=trigger,
        )

        views._run_xiaoce_reply_async(run.id)

        run.refresh_from_db()
        self.assertEqual(run.status, XiaoceRun.Status.COMPLETED)
        self.assertEqual(run.result_message.attachments, [])
        self.assertIn("这是模型生成的正文。", run.result_message.content)
        self.assertIn("文件生成失败", run.result_message.content)

    @patch("apps.collab.views.maybe_generate_file_artifacts")
    @patch("apps.core.agent_chat.run_chat")
    def test_generic_export_does_not_wrap_an_llm_error_in_a_workbook(
        self,
        run_chat,
        generate_excel,
    ):
        run_chat.return_value = {
            "ok": True,
            "reply": "模型调用未成功。",
            "llm": True,
            "llm_error": "certificate failed",
        }
        trigger = CollabMessage.objects.create(
            room=self.room,
            sender=self.user,
            content="把分析结果导出 Excel",
            msg_type="user",
        )
        run = XiaoceRun.objects.create(
            id=uuid.uuid4(),
            room=self.room,
            user=self.user,
            trigger_message=trigger,
        )

        views._run_xiaoce_reply_async(run.id)

        run.refresh_from_db()
        self.assertEqual(run.status, XiaoceRun.Status.COMPLETED)
        self.assertEqual(run.result_message.attachments, [])
        generate_excel.assert_not_called()

    @patch("apps.core.agent_chat.run_chat")
    def test_cancelled_excel_generation_cleans_the_unpublished_file(self, run_chat):
        trigger = CollabMessage.objects.create(
            room=self.room,
            sender=self.user,
            content="生成 Excel，内容里面写一个 hello",
            msg_type="user",
        )
        run = XiaoceRun.objects.create(
            id=uuid.uuid4(),
            room=self.room,
            user=self.user,
            trigger_message=trigger,
        )
        created: list[dict] = []

        def generate_then_cancel(**kwargs):
            from apps.collab.xiaoce_artifacts import create_excel_artifact

            item = create_excel_artifact(user_id=self.user.id, content="hello")
            created.append(item)
            cancel_xiaoce_run(XiaoceRun.objects.get(id=run.id))
            return [item]

        with tempfile.TemporaryDirectory() as tmp:
            with (
                override_settings(CHAT_ATTACHMENTS_ROOT=Path(tmp)),
                patch(
                    "apps.collab.views.maybe_generate_file_artifacts",
                    side_effect=generate_then_cancel,
                ),
            ):
                views._run_xiaoce_reply_async(run.id)
                self.assertIsNone(resolve_attachment_path(self.user.id, created[0]["id"]))

        run.refresh_from_db()
        self.assertEqual(run.status, XiaoceRun.Status.CANCELLED)
        run_chat.assert_not_called()

    @patch("apps.collab.views.threading.Thread")
    def test_reference_rejects_unowned_or_current_task(self, _thread_cls):
        foreign_room = self.create_context_room(owner=self.other, title="他人任务")

        foreign = self.client.post(
            self.messages_url,
            {
                "content": "继续",
                "run_id": str(uuid.uuid4()),
                "context_room_ids": [str(foreign_room.id)],
            },
            format="json",
        )
        current = self.client.post(
            self.messages_url,
            {
                "content": "继续",
                "run_id": str(uuid.uuid4()),
                "context_room_ids": [str(self.room.id)],
            },
            format="json",
        )

        self.assertEqual(foreign.status_code, 400)
        self.assertEqual(current.status_code, 400)
        self.assertEqual(self.room.messages.count(), 0)

    @patch("apps.collab.views.threading.Thread")
    def test_cancel_is_idempotent_and_returns_persisted_snapshot(self, _thread_cls):
        run_id = uuid.uuid4()
        self.client.post(
            self.messages_url,
            {"content": "分析", "run_id": str(run_id)},
        )
        url = f"/api/collab/rooms/{self.room.id}/xiaoce-runs/{run_id}/cancel/"

        first = self.client.post(url)
        second = self.client.post(url)

        self.assertEqual(first.status_code, 200)
        self.assertEqual(second.status_code, 200)
        self.assertEqual(second.data["message"]["id"], first.data["message"]["id"])
        self.assertTrue(first.data["message"]["meta"]["cancelled"])
        self.assertEqual(first.data["xiaoce_run"]["status"], "cancelled")
        self.assertIsNone(first.data["active_xiaoce_run"])

    @patch("apps.collab.views.threading.Thread")
    def test_finished_or_unowned_run_cannot_be_cancelled(self, _thread_cls):
        run_id = uuid.uuid4()
        self.client.post(
            self.messages_url,
            {"content": "分析", "run_id": str(run_id)},
        )
        complete_xiaoce_run(run_id, "已完成")
        url = f"/api/collab/rooms/{self.room.id}/xiaoce-runs/{run_id}/cancel/"
        self.assertEqual(self.client.post(url).status_code, 409)

        self.client.force_authenticate(self.other)
        self.assertEqual(self.client.post(url).status_code, 404)

    @patch("apps.collab.views.threading.Thread")
    @patch("apps.collab.views.ws_push.publish_sync")
    def test_delete_running_xiaoce_task_prevents_late_worker_output(
        self,
        publish,
        _thread_cls,
    ):
        run_id = uuid.uuid4()
        self.client.post(
            self.messages_url,
            {"content": "长任务", "run_id": str(run_id)},
            format="json",
        )
        publish.reset_mock()

        response = self.client.delete(f"/api/collab/rooms/{self.room.id}/")
        views._run_xiaoce_reply_async(run_id)

        self.assertEqual(response.status_code, 200)
        self.assertFalse(CollabRoom.objects.filter(id=self.room.id).exists())
        self.assertFalse(XiaoceRun.objects.filter(id=run_id).exists())
        publish.assert_not_called()

    def test_xiaoce_task_rename_rejects_blank_and_has_no_group_announcement(self):
        before = CollabMessage.objects.filter(room=self.room, msg_type="system").count()

        renamed = self.client.patch(
            f"/api/collab/rooms/{self.room.id}/",
            {"title": "  小策bot（GMV运算处理任务）  "},
            format="json",
        )
        blank = self.client.patch(
            f"/api/collab/rooms/{self.room.id}/",
            {"title": "   "},
            format="json",
        )

        self.assertEqual(renamed.status_code, 200)
        self.assertEqual(renamed.data["title"], "小策bot（GMV运算处理任务）")
        self.assertEqual(blank.status_code, 400)
        self.assertEqual(blank.data["error"], "会话名称不能为空")
        self.assertEqual(
            CollabMessage.objects.filter(room=self.room, msg_type="system").count(),
            before,
        )

    @patch("apps.core.agent_chat.run_chat")
    def test_worker_persists_final_process_snapshot(self, run_chat):
        trigger = CollabMessage.objects.create(
            room=self.room,
            sender=self.user,
            content="分析",
            msg_type="user",
            meta={},
        )
        run = XiaoceRun.objects.create(
            id=uuid.uuid4(),
            room=self.room,
            user=self.user,
            trigger_message=trigger,
        )

        def answer(**kwargs):
            callback = kwargs["progress_callback"]
            callback("understanding", "running", {})
            callback("understanding", "completed", {})
            callback("composing", "running", {})
            callback("composing", "completed", {})
            return {"ok": True, "reply": "最终答案"}

        run_chat.side_effect = answer

        views._run_xiaoce_reply_async(run.id)

        run.refresh_from_db()
        self.assertEqual(run.status, XiaoceRun.Status.COMPLETED)
        self.assertEqual(run.result_message.content, "最终答案")
        self.assertEqual(run.result_message.meta["process_steps"], run.progress_steps)

    @patch("apps.core.agent_chat.run_chat")
    def test_worker_injects_referenced_task_transcript(self, run_chat):
        context_room = self.create_context_room()
        CollabMessage.objects.create(
            room=context_room,
            sender=self.user,
            content="方案代号是 Aurora，预算 20 万",
            msg_type="user",
        )
        snapshot_tail = CollabMessage.objects.create(
            room=context_room,
            sender=self.bot,
            content="下一步是确认供应商档期",
            msg_type="ai",
            ai_kind="xiaoce",
        )
        trigger = CollabMessage.objects.create(
            room=self.room,
            sender=self.user,
            content="继续这个任务",
            msg_type="user",
            meta={
                "context_rooms": [{
                    "id": str(context_room.id),
                    "title": context_room.title,
                    "message_count": 2,
                    "last_message_id": snapshot_tail.id,
                }],
            },
        )
        CollabMessage.objects.create(
            room=context_room,
            sender=self.user,
            content="引用后新增的内容不应读取",
            msg_type="user",
        )
        run = XiaoceRun.objects.create(
            id=uuid.uuid4(),
            room=self.room,
            user=self.user,
            trigger_message=trigger,
        )
        run_chat.return_value = {"ok": True, "reply": "已承接"}

        views._run_xiaoce_reply_async(run.id)

        self.assertEqual(run_chat.call_args.kwargs["usage_source"], "agent")
        block = run_chat.call_args.kwargs["extra_reference_blocks"][0]
        self.assertIn("方案代号是 Aurora", block)
        self.assertIn("下一步是确认供应商档期", block)
        self.assertIn(context_room.title, block)
        self.assertNotIn("引用后新增的内容不应读取", block)

        run_chat.reset_mock()
        followup = CollabMessage.objects.create(
            room=self.room,
            sender=self.user,
            content="那么现在先做什么？",
            msg_type="user",
            meta={},
        )
        followup_run = XiaoceRun.objects.create(
            id=uuid.uuid4(),
            room=self.room,
            user=self.user,
            trigger_message=followup,
        )

        views._run_xiaoce_reply_async(followup_run.id)

        inherited_block = run_chat.call_args.kwargs["extra_reference_blocks"][0]
        self.assertIn("方案代号是 Aurora", inherited_block)

    @patch("apps.core.agent_chat.run_chat")
    def test_worker_maps_internal_failure_to_safe_message(self, run_chat):
        trigger = CollabMessage.objects.create(
            room=self.room,
            sender=self.user,
            content="分析",
            msg_type="user",
        )
        run = XiaoceRun.objects.create(
            id=uuid.uuid4(),
            room=self.room,
            user=self.user,
            trigger_message=trigger,
        )
        run_chat.side_effect = RuntimeError("Traceback /srv/app.py sk-secret-value")

        views._run_xiaoce_reply_async(run.id)

        run.refresh_from_db()
        self.assertEqual(run.status, XiaoceRun.Status.FAILED)
        self.assertNotIn("Traceback", run.result_message.content)
        self.assertNotIn("sk-secret-value", run.result_message.content)

    @patch("apps.skills.repository.cos_enabled", return_value=False)
    @patch("apps.core.conversation_skill.prepare_conversation_skill")
    def test_worker_packages_explicit_request_and_enables_private_skill(
        self,
        prepare,
        _cos_enabled,
    ):
        trigger = CollabMessage.objects.create(
            room=self.room,
            sender=self.user,
            content="把当前聊天记录打包成 Skill 并自动上传平台",
            msg_type="user",
        )
        run = XiaoceRun.objects.create(
            id=uuid.uuid4(),
            room=self.room,
            user=self.user,
            trigger_message=trigger,
        )
        prepare.return_value = prepared_skill(self.room)

        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(SKILLS_WORKSPACE_ROOT=Path(tmp)):
                views._run_xiaoce_reply_async(run.id)

        run.refresh_from_db()
        asset = SkillAsset.objects.get()
        personal = UserSkill.objects.get()
        self.assertEqual(run.status, XiaoceRun.Status.COMPLETED)
        self.assertEqual(asset.visibility, SkillAsset.Visibility.PRIVATE)
        self.assertTrue(personal.enabled)
        self.assertEqual(run.result_message.meta["created_skill"]["skill_id"], asset.skill_id)
