import tempfile
import zipfile
from pathlib import Path
from unittest.mock import patch
from xml.etree import ElementTree

from django.test import SimpleTestCase, override_settings
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from apps.collab import xiaoce_artifacts
from apps.collab.xiaoce_artifacts import (
    XLSX_MIME,
    XlsxResourceLimitError,
    create_excel_artifact,
    extract_excel_a1_content,
    extract_excel_cell_update,
    is_excel_generation_request,
    maybe_generate_excel_artifact,
)
from apps.core.attachments import resolve_attachment_path


class XiaoceExcelArtifactTests(SimpleTestCase):
    @staticmethod
    def generate_from_existing(user_id: int, source_path: Path):
        return maybe_generate_excel_artifact(
            user_id=user_id,
            request_text="在上传的 xlsx 的 A1 写入 hello 并导出 Excel",
            model_reply=None,
            source_attachments=[{
                "id": source_path.name,
                "name": "source.xlsx",
            }],
        )

    @staticmethod
    def relocate_first_worksheet(
        source_path: Path,
        *,
        target: str = "evil.xml",
        package_path: str = "xl/evil.xml",
        dimension: str | None = None,
        target_mode: str | None = None,
    ) -> None:
        rewritten_path = source_path.with_name("rewritten.xlsx")
        with (
            zipfile.ZipFile(source_path, "r") as original,
            zipfile.ZipFile(
                rewritten_path,
                "w",
                compression=zipfile.ZIP_DEFLATED,
            ) as rewritten,
        ):
            for info in original.infolist():
                data = original.read(info.filename)
                output_name = info.filename
                if info.filename == "xl/_rels/workbook.xml.rels":
                    root = ElementTree.fromstring(data)
                    for relationship in root:
                        if str(relationship.attrib.get("Type") or "").endswith(
                            "/worksheet",
                        ):
                            relationship.set("Target", target)
                            if target_mode is not None:
                                relationship.set("TargetMode", target_mode)
                            break
                    data = ElementTree.tostring(
                        root,
                        encoding="utf-8",
                        xml_declaration=True,
                    )
                elif info.filename == "xl/worksheets/sheet1.xml":
                    output_name = package_path
                    if dimension is not None:
                        root = ElementTree.fromstring(data)
                        for element in root.iter():
                            if str(element.tag).rsplit("}", 1)[-1] == "dimension":
                                element.set("ref", dimension)
                                break
                        data = ElementTree.tostring(
                            root,
                            encoding="utf-8",
                            xml_declaration=True,
                        )
                rewritten.writestr(output_name, data)
        rewritten_path.replace(source_path)

    @staticmethod
    def duplicate_worksheet_relationship_target(source_path: Path) -> None:
        rewritten_path = source_path.with_name("duplicate-target.xlsx")
        with (
            zipfile.ZipFile(source_path, "r") as original,
            zipfile.ZipFile(
                rewritten_path,
                "w",
                compression=zipfile.ZIP_DEFLATED,
            ) as rewritten,
        ):
            for info in original.infolist():
                data = original.read(info.filename)
                if info.filename == "xl/_rels/workbook.xml.rels":
                    root = ElementTree.fromstring(data)
                    worksheet_relationships = [
                        relationship
                        for relationship in root
                        if str(relationship.attrib.get("Type") or "").endswith(
                            "/worksheet",
                        )
                    ]
                    duplicate_target = worksheet_relationships[0].attrib["Target"]
                    for relationship in worksheet_relationships[1:]:
                        relationship.set("Target", duplicate_target)
                    data = ElementTree.tostring(
                        root,
                        encoding="utf-8",
                        xml_declaration=True,
                    )
                rewritten.writestr(info.filename, data)
        rewritten_path.replace(source_path)

    def test_only_explicit_excel_production_requests_match(self):
        self.assertTrue(is_excel_generation_request("可以帮我产出一个excel吗，内容里面写一个hello就好"))
        self.assertTrue(is_excel_generation_request("Please create an xlsx spreadsheet"))
        self.assertTrue(is_excel_generation_request("把分析结果导出 Excel"))
        self.assertFalse(is_excel_generation_request("读取这个 Excel 并做分析"))
        self.assertFalse(is_excel_generation_request("这个 xlsx 里面都有什么？"))
        self.assertFalse(is_excel_generation_request("读取你刚才生成的这个 Excel 文件"))
        self.assertFalse(is_excel_generation_request("把之前创建的 xlsx 发给我"))
        self.assertTrue(is_excel_generation_request("重新生成这个 Excel"))

    def test_negated_or_inability_questions_do_not_generate(self):
        for request in (
            "不要帮我生成 Excel",
            "为什么不能生成 Excel？",
            "小策无法创建 xlsx 文件吗？",
            "Please do not create an Excel file.",
            "Why can't you generate an xlsx?",
            "This Excel must not be exported.",
            "Excel cannot be generated here.",
        ):
            with self.subTest(request=request):
                self.assertFalse(is_excel_generation_request(request))

        self.assertTrue(is_excel_generation_request("之前无法生成 Excel，现在请生成 Excel"))
        self.assertTrue(is_excel_generation_request("能不能帮我生成 Excel？"))

    def test_extracts_explicit_a1_content_in_chinese_and_english(self):
        self.assertEqual(
            extract_excel_a1_content("可以帮我产出一个excel吗 内容里面写一个hello就好、"),
            "hello",
        )
        self.assertEqual(extract_excel_a1_content("Create an xlsx; A1 should say hello"), "hello")
        self.assertIsNone(extract_excel_a1_content("请生成一个 Excel"))

    def test_content_extraction_stops_before_follow_up_instructions(self):
        self.assertEqual(
            extract_excel_a1_content("生成 Excel，内容里面写一个 hello，然后把文件发给我"),
            "hello",
        )
        self.assertEqual(
            extract_excel_a1_content("在 A1 写入 hello 并把文件发给我，文件名为 greeting.xlsx"),
            "hello",
        )
        self.assertEqual(
            extract_excel_a1_content("Create an xlsx; content should say hello and send the file back"),
            "hello",
        )
        self.assertEqual(
            extract_excel_cell_update("在上传的 Excel 的 C3 单元格写入 hello，然后导出"),
            ("C3", "hello"),
        )
        self.assertEqual(
            extract_excel_cell_update("在已有 xlsx 的 D5 写值 hello 并导出 Excel"),
            ("D5", "hello"),
        )

    def test_writes_a_safe_workbook_in_the_current_users_attachment_directory(self):
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(CHAT_ATTACHMENTS_ROOT=Path(tmp)):
                item = create_excel_artifact(
                    user_id=42,
                    content="=HYPERLINK(\"https://example.test\")",
                    filename="..\\..\\unsafe name.xlsx",
                )
                path = resolve_attachment_path(42, item["id"])
                self.assertIsNotNone(path)
                self.assertEqual(path.parent, (Path(tmp) / "42").resolve())
                self.assertEqual(item["name"], "unsafe-name.xlsx")
                self.assertEqual(item["mime"], XLSX_MIME)
                self.assertEqual(item["size"], path.stat().st_size)

                workbook = load_workbook(path, data_only=False)
                try:
                    self.assertEqual(workbook.active["A1"].value, "'=HYPERLINK(\"https://example.test\")")
                    self.assertEqual(workbook.active["A1"].data_type, "s")
                finally:
                    workbook.close()

    def test_generates_requested_text_or_falls_back_to_the_model_reply(self):
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(CHAT_ATTACHMENTS_ROOT=Path(tmp)):
                requested = maybe_generate_excel_artifact(
                    user_id=7,
                    request_text="帮我生成 Excel，内容里面写一个 hello 就好",
                    model_reply="这段回复不应该写进去",
                )
                fallback = maybe_generate_excel_artifact(
                    user_id=7,
                    request_text="Export an Excel spreadsheet",
                    model_reply="模型生成的分析结论",
                )
                absent = maybe_generate_excel_artifact(
                    user_id=7,
                    request_text="分析这个 Excel",
                    model_reply="不会生成文件",
                )

                self.assertEqual(len(requested), 1)
                self.assertEqual(len(fallback), 1)
                self.assertEqual(absent, [])
                for item, expected in ((requested[0], "hello"), (fallback[0], "模型生成的分析结论")):
                    workbook = load_workbook(resolve_attachment_path(7, item["id"]), data_only=False)
                    try:
                        self.assertEqual(workbook.active["A1"].value, expected)
                    finally:
                        workbook.close()

    def test_general_export_requires_a_nonempty_model_reply(self):
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(CHAT_ATTACHMENTS_ROOT=Path(tmp)):
                output = maybe_generate_excel_artifact(
                    user_id=7,
                    request_text="导出 Excel",
                    model_reply=None,
                )

                self.assertEqual(output, [])
                self.assertFalse((Path(tmp) / "7").exists())

    def test_existing_workbook_is_copied_and_only_the_requested_cell_changes(self):
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(CHAT_ATTACHMENTS_ROOT=Path(tmp)):
                root = Path(tmp) / "7"
                root.mkdir(parents=True)
                source_path = root / "source_workbook.xlsx"
                source = Workbook()
                source.active.title = "原数据"
                source.active["A1"] = "保留标题"
                source.active["B2"] = "旧值"
                source.active["C3"] = "=1+1"
                source.active["A1"].font = Font(bold=True)
                source.create_sheet("其他表")["D4"] = "仍需保留"
                source.save(source_path)
                source.close()

                generated = maybe_generate_excel_artifact(
                    user_id=7,
                    request_text=(
                        "请在上传的 Excel 的 B2 单元格写入 hello，"
                        "然后把文件发给我，并导出 Excel"
                    ),
                    model_reply=None,
                    source_attachments=[{
                        "id": source_path.name,
                        "name": "业务模板.xlsx",
                        "mime": XLSX_MIME,
                    }],
                )

                self.assertEqual(len(generated), 1)
                self.assertNotEqual(generated[0]["id"], source_path.name)
                result_path = resolve_attachment_path(7, generated[0]["id"])
                result = load_workbook(result_path, data_only=False)
                try:
                    self.assertEqual(result.active.title, "原数据")
                    self.assertEqual(result.active["A1"].value, "保留标题")
                    self.assertTrue(result.active["A1"].font.bold)
                    self.assertEqual(result.active["B2"].value, "hello")
                    self.assertEqual(result.active["C3"].value, "=1+1")
                    self.assertEqual(result["其他表"]["D4"].value, "仍需保留")
                finally:
                    result.close()

                original = load_workbook(source_path, data_only=False)
                try:
                    self.assertEqual(original.active["B2"].value, "旧值")
                finally:
                    original.close()

    def test_existing_workbook_load_failures_are_not_swallowed(self):
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(CHAT_ATTACHMENTS_ROOT=Path(tmp)):
                root = Path(tmp) / "7"
                root.mkdir(parents=True)
                corrupt = root / "corrupt.xlsx"
                corrupt.write_bytes(b"not an xlsx")

                with self.assertRaises(zipfile.BadZipFile):
                    maybe_generate_excel_artifact(
                        user_id=7,
                        request_text="在上传的 xlsx 的 A1 写入 hello 并导出 Excel",
                        model_reply=None,
                        source_attachments=[{
                            "id": corrupt.name,
                            "name": "corrupt.xlsx",
                        }],
                    )

    def test_existing_workbook_rejects_zip_package_resource_limits(self):
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(CHAT_ATTACHMENTS_ROOT=Path(tmp)):
                root = Path(tmp) / "7"
                root.mkdir(parents=True)
                source_path = root / "source.xlsx"
                source = Workbook()
                source.active["A1"] = "value"
                source.save(source_path)
                source.close()

                limits = (
                    ("MAX_XLSX_SOURCE_BYTES", 1),
                    ("MAX_XLSX_ZIP_ENTRIES", 1),
                    ("MAX_XLSX_EXPANDED_BYTES", 1),
                    ("MAX_XLSX_SINGLE_ENTRY_BYTES", 1),
                    ("MAX_XLSX_WORKSHEET_XML_BYTES", 1),
                )
                for constant, limit in limits:
                    with self.subTest(constant=constant):
                        with patch.object(xiaoce_artifacts, constant, limit):
                            with self.assertRaises(XlsxResourceLimitError):
                                self.generate_from_existing(7, source_path)

    def test_existing_workbook_rejects_oversized_dimensions_and_cell_counts(self):
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(CHAT_ATTACHMENTS_ROOT=Path(tmp)):
                root = Path(tmp) / "7"
                root.mkdir(parents=True)
                source_path = root / "source.xlsx"
                source = Workbook()
                source.active["A1"] = "one"
                source.active["B1"] = "two"
                source.active["C1"] = "three"
                second = source.create_sheet("Second")
                second["A1"] = "four"
                second["A2"] = "five"
                source.save(source_path)
                source.close()

                cases = (
                    ("MAX_XLSX_DIMENSION_CELLS", 2),
                    ("MAX_XLSX_WORKSHEET_CELLS", 2),
                    ("MAX_XLSX_TOTAL_CELLS", 4),
                    ("MAX_XLSX_WORKSHEETS", 1),
                    ("MAX_XLSX_WORKSHEET_ROWS", 1),
                )
                for constant, limit in cases:
                    with self.subTest(constant=constant):
                        with patch.object(xiaoce_artifacts, constant, limit):
                            with patch("openpyxl.load_workbook") as load_workbook_mock:
                                with self.assertRaises(XlsxResourceLimitError):
                                    self.generate_from_existing(7, source_path)
                                load_workbook_mock.assert_not_called()

    def test_nonstandard_relationship_target_cannot_bypass_worksheet_limits(self):
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(CHAT_ATTACHMENTS_ROOT=Path(tmp)):
                root = Path(tmp) / "7"
                root.mkdir(parents=True)
                source_path = root / "source.xlsx"
                source = Workbook()
                source.active["A1"] = "value"
                source.save(source_path)
                source.close()
                self.relocate_first_worksheet(
                    source_path,
                    target="evil.xml",
                    package_path="xl/evil.xml",
                    dimension="A1:XFD1048576",
                )

                with patch("openpyxl.load_workbook") as load_workbook_mock:
                    with self.assertRaises(XlsxResourceLimitError):
                        self.generate_from_existing(7, source_path)
                    load_workbook_mock.assert_not_called()

    def test_invalid_external_and_traversing_worksheet_targets_are_rejected(self):
        cases = (
            ("../evil.xml", "evil.xml", None),
            ("https://example.test/evil.xml", "xl/evil.xml", "External"),
            ("missing.xml", "xl/evil.xml", None),
        )
        for target, package_path, target_mode in cases:
            with self.subTest(target=target):
                with tempfile.TemporaryDirectory() as tmp:
                    with override_settings(CHAT_ATTACHMENTS_ROOT=Path(tmp)):
                        root = Path(tmp) / "7"
                        root.mkdir(parents=True)
                        source_path = root / "source.xlsx"
                        source = Workbook()
                        source.active["A1"] = "value"
                        source.save(source_path)
                        source.close()
                        self.relocate_first_worksheet(
                            source_path,
                            target=target,
                            package_path=package_path,
                            target_mode=target_mode,
                        )

                        with patch("openpyxl.load_workbook") as load_workbook_mock:
                            with self.assertRaises(XlsxResourceLimitError):
                                self.generate_from_existing(7, source_path)
                            load_workbook_mock.assert_not_called()

    def test_duplicate_normalized_worksheet_targets_are_rejected(self):
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(CHAT_ATTACHMENTS_ROOT=Path(tmp)):
                root = Path(tmp) / "7"
                root.mkdir(parents=True)
                source_path = root / "source.xlsx"
                source = Workbook()
                source.active["A1"] = "one"
                source.create_sheet("Second")["A1"] = "two"
                source.save(source_path)
                source.close()
                self.duplicate_worksheet_relationship_target(source_path)

                with patch("openpyxl.load_workbook") as load_workbook_mock:
                    with self.assertRaises(XlsxResourceLimitError):
                        self.generate_from_existing(7, source_path)
                    load_workbook_mock.assert_not_called()
