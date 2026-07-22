"""CSV / Excel 导入解析（智能表格）。"""
from __future__ import annotations

import csv
import io
import re
import zipfile
from datetime import date, datetime
from decimal import Decimal
from xml.etree import ElementTree

XLS_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
XLSX_MAGIC = b"PK"


def cell_to_text(cell) -> str:
    if cell is None:
        return ""
    if isinstance(cell, bool):
        return "是" if cell else "否"
    if isinstance(cell, datetime):
        return cell.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(cell, date):
        return cell.isoformat()
    if isinstance(cell, Decimal):
        text = format(cell, "f")
        return text.rstrip("0").rstrip(".") if "." in text else text
    if isinstance(cell, float):
        if cell == int(cell):
            return str(int(cell))
        return str(cell).strip()
    return str(cell).strip()


def rows_from_csv_text(raw: str) -> list[dict]:
    reader = csv.DictReader(io.StringIO(str(raw)))
    return [dict(item or {}) for item in reader]


def detect_spreadsheet_kind(content: bytes, filename: str = "") -> str:
    name = (filename or "").lower()
    if content[:8] == XLS_MAGIC:
        return "xls"
    if content[:2] == XLSX_MAGIC:
        return "xlsx"
    if name.endswith((".xlsx", ".xlsm")):
        return "xlsx"
    if name.endswith(".xls") and not name.endswith((".xlsx", ".xlsm")):
        return "xls"
    if name.endswith(".csv") or name.endswith(".txt"):
        return "csv"
    sample = content[:8192]
    if sample.startswith(b"\xef\xbb\xbf") or b"," in sample or b"\t" in sample:
        try:
            sample.decode("utf-8-sig")
            return "csv"
        except UnicodeDecodeError:
            pass
    return "unknown"


def _dict_rows_from_matrix(matrix: list[list[object]]) -> list[dict]:
    if not matrix:
        return []
    headers = [cell_to_text(h) for h in matrix[0]]
    out: list[dict] = []
    for row in matrix[1:]:
        if not row:
            continue
        item: dict[str, str] = {}
        empty = True
        for idx, header in enumerate(headers):
            if not header:
                continue
            value = cell_to_text(row[idx] if idx < len(row) else None)
            if value:
                empty = False
            item[header] = value
        if not empty:
            out.append(item)
    return out


def rows_from_xlsx_openpyxl(content: bytes, *, read_only: bool) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=read_only, data_only=True)
    try:
        ws = wb.active
        matrix: list[list[object]] = []
        for row in ws.iter_rows(values_only=True):
            if row is None:
                continue
            matrix.append(list(row))
        return _dict_rows_from_matrix(matrix)
    finally:
        wb.close()


def _xlsx_column_index(cell_ref: str) -> int:
    letters = "".join(char for char in cell_ref if char.isalpha()).upper()
    index = 0
    for char in letters:
        index = index * 26 + (ord(char) - ord("A") + 1)
    return max(index - 1, 0)


def _xlsx_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    try:
        xml = archive.read("xl/sharedStrings.xml")
    except KeyError:
        return []
    root = ElementTree.fromstring(xml)
    ns = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    return ["".join(node.text or "" for node in item.findall(".//s:t", ns)) for item in root.findall("s:si", ns)]


def _xlsx_sheet_paths(archive: zipfile.ZipFile) -> list[tuple[str, str]]:
    workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    rels = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    ns = {
        "s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "pr": "http://schemas.openxmlformats.org/package/2006/relationships",
    }
    rel_targets = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels.findall("pr:Relationship", ns)
        if "Id" in rel.attrib and "Target" in rel.attrib
    }
    sheets: list[tuple[str, str]] = []
    for sheet in workbook.findall("s:sheets/s:sheet", ns):
        name = sheet.attrib.get("name", f"Sheet{len(sheets) + 1}")
        rel_id = sheet.attrib.get(f"{{{ns['r']}}}id")
        target = rel_targets.get(rel_id or "")
        if not target:
            continue
        path = target.lstrip("/")
        if not path.startswith("xl/"):
            path = f"xl/{path}"
        sheets.append((name, path))
    return sheets


def _xlsx_cell_value(cell: ElementTree.Element, shared_strings: list[str], ns: dict[str, str]) -> object:
    cell_type = cell.attrib.get("t")
    value_node = cell.find("s:v", ns)
    if cell_type == "inlineStr":
        return "".join(node.text or "" for node in cell.findall(".//s:t", ns))
    if value_node is None or value_node.text is None:
        return None
    raw = value_node.text
    if cell_type == "s":
        try:
            return shared_strings[int(raw)]
        except (ValueError, IndexError):
            return raw
    if cell_type == "b":
        return raw == "1"
    try:
        if "." in raw or "e" in raw.lower():
            return float(raw)
        return int(raw)
    except ValueError:
        return raw


def rows_from_xlsx_zipxml(content: bytes) -> list[dict]:
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        shared_strings = _xlsx_shared_strings(archive)
        ns = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        sheet_paths = _xlsx_sheet_paths(archive)
        if not sheet_paths:
            return []
        _, sheet_path = sheet_paths[0]
        root = ElementTree.fromstring(archive.read(sheet_path))
        matrix: list[list[object]] = []
        for row in root.findall(".//s:sheetData/s:row", ns):
            values: list[object] = []
            for cell in row.findall("s:c", ns):
                column_index = _xlsx_column_index(cell.attrib.get("r", ""))
                while len(values) <= column_index:
                    values.append(None)
                values[column_index] = _xlsx_cell_value(cell, shared_strings, ns)
            matrix.append(values)
        return _dict_rows_from_matrix(matrix)


def rows_from_xlsx(content: bytes) -> list[dict]:
    errors: list[str] = []
    for read_only in (True, False):
        try:
            rows = rows_from_xlsx_openpyxl(content, read_only=read_only)
            if rows:
                return rows
        except Exception as exc:
            errors.append(f"openpyxl({'read_only' if read_only else 'normal'}): {exc}")
    try:
        rows = rows_from_xlsx_zipxml(content)
        if rows:
            return rows
    except Exception as exc:
        errors.append(f"zipxml: {exc}")
    detail = errors[-1] if errors else "empty workbook"
    raise ValueError(detail)


def rows_from_xls(content: bytes) -> list[dict]:
    try:
        import xlrd
    except ImportError as exc:
        raise ValueError("旧版 .xls 需要 xlrd 库，请联系管理员安装或另存为 .xlsx") from exc

    book = xlrd.open_workbook(file_contents=content)
    sheet = book.sheet_by_index(0)
    if sheet.nrows == 0:
        return []
    matrix: list[list[object]] = []
    for r in range(sheet.nrows):
        row_vals: list[object] = []
        for c in range(sheet.ncols):
            ctype = sheet.cell_type(r, c)
            val = sheet.cell_value(r, c)
            if ctype == xlrd.XL_CELL_DATE:
                try:
                    dt = xlrd.xldate_as_datetime(val, book.datemode)
                    row_vals.append(dt.date() if dt.hour == 0 and dt.minute == 0 and dt.second == 0 else dt)
                except Exception:
                    row_vals.append(val)
            elif ctype == xlrd.XL_CELL_BOOLEAN:
                row_vals.append(bool(val))
            else:
                row_vals.append(val)
        matrix.append(row_vals)
    return _dict_rows_from_matrix(matrix)


def parse_spreadsheet_upload(content: bytes, filename: str = "") -> tuple[list[dict], str]:
    kind = detect_spreadsheet_kind(content, filename)
    if kind == "xlsx":
        return rows_from_xlsx(content), "xlsx"
    if kind == "xls":
        return rows_from_xls(content), "xls"
    if kind == "csv":
        raw = content.decode("utf-8-sig", errors="replace")
        return rows_from_csv_text(raw), "csv"
    # 兜底：先 xlsx 再 csv
    if content[:2] == XLSX_MAGIC:
        return rows_from_xlsx(content), "xlsx"
    raw = content.decode("utf-8-sig", errors="replace")
    if raw.strip():
        return rows_from_csv_text(raw), "csv"
    raise ValueError("无法识别文件格式，请上传 .xlsx 或 .csv")


def normalize_import_text(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


_DATE_RE = re.compile(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}")
_CHECKBOX_VALUES = {"1", "0", "true", "false", "是", "否", "yes", "no", "y", "n", "已勾选", "未勾选"}


def sheet_name_from_filename(filename: str) -> str:
    base = (filename or "").rsplit("/", 1)[-1].rsplit("\\", 1)[-1]
    for ext in (".xlsx", ".xlsm", ".xls", ".csv", ".txt"):
        if base.lower().endswith(ext):
            base = base[: -len(ext)]
            break
    base = base.strip() or "导入表格"
    return base[:120]


def collect_headers(dict_rows: list[dict]) -> list[str]:
    if not dict_rows:
        return []
    ordered: list[str] = []
    seen: set[str] = set()
    for row in dict_rows:
        for title in row.keys():
            key = str(title or "").strip()
            if not key or key in seen:
                continue
            seen.add(key)
            ordered.append(key)
    return ordered


def _is_number(text: str) -> bool:
    if not text:
        return False
    try:
        float(text.replace(",", "").replace("%", "").strip())
        return True
    except ValueError:
        return False


def _is_date(text: str) -> bool:
    if not text:
        return False
    return bool(_DATE_RE.match(text.strip()))


def _looks_like_short_label(text: str) -> bool:
    """单选选项应是短标签，不含长段落 / 换行。"""
    t = text.strip()
    if not t or len(t) > 24:
        return False
    if "\n" in t or "\r" in t:
        return False
    # 排除明显的说明性长句
    if "）" in t and len(t) > 12:
        return False
    return True


def infer_column_spec(values: list[str]) -> tuple[str, list[str]]:
    """导入默认全部按文本建列；类型可在字段配置里再改。"""
    return "text", []
