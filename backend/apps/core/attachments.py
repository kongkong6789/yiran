"""对话附件：保存、文本抽取与图片 vision 准备。"""
from __future__ import annotations

import base64
import io
import json
import mimetypes
import uuid
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from django.conf import settings

MAX_ATTACH_BYTES = 20 * 1024 * 1024
MAX_ATTACH_FILES = 5
TEXT_EXTENSIONS = {
    ".md", ".markdown", ".txt", ".json", ".csv", ".py", ".log",
    ".yaml", ".yml", ".xml", ".html", ".htm", ".tsv",
}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"}
# 协作/对话可下载的二进制附件；其中 Excel 会进一步抽取文本。
DOC_EXTENSIONS = {
    ".pdf", ".doc", ".docx", ".xls", ".xlsx", ".xlsm", ".ppt", ".pptx",
    ".zip", ".rar", ".7z", ".tar", ".gz",
    ".mp3", ".wav", ".mp4", ".mov", ".avi",
    ".apk", ".ipa",
}
IMAGE_MIME = {
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".gif": "image/gif",
    ".webp": "image/webp",
    ".bmp": "image/bmp",
}
MAX_TEXT_INJECT = 12_000
SPREADSHEET_EXTENSIONS = {".xls", ".xlsx", ".xlsm"}
MAX_SPREADSHEET_SHEETS = 6
MAX_SPREADSHEET_ROWS = 200
MAX_SPREADSHEET_COLUMNS = 40
MAX_CELL_CHARS = 500


def attachments_root(user_id: int) -> Path:
    base = Path(getattr(settings, "CHAT_ATTACHMENTS_ROOT", settings.BASE_DIR / "chat_attachments"))
    return base / str(user_id)


def _decode_text(data: bytes) -> str:
    for enc in ("utf-8", "utf-8-sig", "gbk", "latin-1"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _cell_to_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        text = format(value, "f")
        return text.rstrip("0").rstrip(".") if "." in text else text
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\t", " ").strip()[:MAX_CELL_CHARS]


def _sheet_text(
    title: str,
    rows,
    *,
    total_rows: int | None = None,
    total_columns: int | None = None,
) -> str:
    lines = [f"## 工作表: {title}"]
    emitted = 0
    for raw_row in rows:
        if emitted >= MAX_SPREADSHEET_ROWS:
            break
        row = list(raw_row or ())
        truncated_columns = (
            len(row) > MAX_SPREADSHEET_COLUMNS
            or bool(total_columns and total_columns > MAX_SPREADSHEET_COLUMNS)
        )
        values = [_cell_to_text(value) for value in row[:MAX_SPREADSHEET_COLUMNS]]
        while values and not values[-1]:
            values.pop()
        if not any(values):
            continue
        line = "\t".join(values)
        if truncated_columns:
            line += "\t……（其余列已省略）"
        lines.append(line)
        emitted += 1
    if total_rows is not None and total_rows > MAX_SPREADSHEET_ROWS:
        lines.append(
            f"……（工作表共 {total_rows} 行，"
            f"仅扫描前 {MAX_SPREADSHEET_ROWS} 行以保证安全）"
        )
    elif emitted >= MAX_SPREADSHEET_ROWS:
        lines.append(f"……（仅展示前 {MAX_SPREADSHEET_ROWS} 行非空内容）")
    if emitted == 0:
        lines.append("（工作表为空）")
    return "\n".join(lines)


def _extract_xlsx_text(data: bytes) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        blocks = [
            _sheet_text(
                sheet.title,
                sheet.iter_rows(
                    min_row=1,
                    max_row=min(sheet.max_row, MAX_SPREADSHEET_ROWS),
                    max_col=min(sheet.max_column, MAX_SPREADSHEET_COLUMNS),
                    values_only=True,
                ),
                total_rows=sheet.max_row,
                total_columns=sheet.max_column,
            )
            for sheet in workbook.worksheets[:MAX_SPREADSHEET_SHEETS]
        ]
        if len(workbook.worksheets) > MAX_SPREADSHEET_SHEETS:
            blocks.append(
                f"……（文件共 {len(workbook.worksheets)} 个工作表，仅展示前 {MAX_SPREADSHEET_SHEETS} 个）"
            )
        return "\n\n".join(blocks)[:MAX_TEXT_INJECT]
    finally:
        workbook.close()


def _extract_xls_text(data: bytes) -> str:
    import xlrd

    workbook = xlrd.open_workbook(file_contents=data, on_demand=True)
    try:
        blocks: list[str] = []
        for sheet in workbook.sheets()[:MAX_SPREADSHEET_SHEETS]:
            rows = []
            for row_index in range(min(sheet.nrows, MAX_SPREADSHEET_ROWS)):
                values = []
                for column_index in range(min(sheet.ncols, MAX_SPREADSHEET_COLUMNS)):
                    cell = sheet.cell(row_index, column_index)
                    value = cell.value
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            value = xlrd.xldate_as_datetime(value, workbook.datemode)
                        except (TypeError, ValueError):
                            pass
                    elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                        value = bool(value)
                    values.append(value)
                rows.append(values)
            blocks.append(_sheet_text(sheet.name, rows, total_rows=sheet.nrows))
        if workbook.nsheets > MAX_SPREADSHEET_SHEETS:
            blocks.append(
                f"……（文件共 {workbook.nsheets} 个工作表，仅展示前 {MAX_SPREADSHEET_SHEETS} 个）"
            )
        return "\n\n".join(blocks)[:MAX_TEXT_INJECT]
    finally:
        workbook.release_resources()


def _extract_text(name: str, data: bytes) -> str:
    lower = name.lower()
    ext = Path(lower).suffix
    if ext in SPREADSHEET_EXTENSIONS:
        try:
            return _extract_xls_text(data) if ext == ".xls" else _extract_xlsx_text(data)
        except Exception:
            return "（Excel 文件解析失败，请确认文件未损坏、未加密，并重新上传。）"
    if ext not in TEXT_EXTENSIONS:
        return ""
    text = _decode_text(data)
    if ext == ".json":
        try:
            obj = json.loads(text)
            text = json.dumps(obj, ensure_ascii=False, indent=2)
        except json.JSONDecodeError:
            pass
    return text[:MAX_TEXT_INJECT]


def _is_image(name: str, mime: str = "") -> bool:
    ext = Path((name or "").lower()).suffix
    if ext in IMAGE_EXTENSIONS:
        return True
    return (mime or "").lower().startswith("image/")


def _image_mime(name: str, mime: str = "") -> str:
    if mime and mime.startswith("image/"):
        return mime
    ext = Path((name or "").lower()).suffix
    return IMAGE_MIME.get(ext) or mimetypes.guess_type(name)[0] or "image/jpeg"


def process_uploaded_files(files, user_id: int) -> list[dict]:
    """保存上传文件并返回元数据、文本或图片 base64。"""
    if not files:
        return []
    if len(files) > MAX_ATTACH_FILES:
        raise ValueError(f"最多上传 {MAX_ATTACH_FILES} 个附件")

    root = attachments_root(user_id)
    root.mkdir(parents=True, exist_ok=True)
    results: list[dict] = []

    for upload in files:
        name = (getattr(upload, "name", "") or "file").replace("\\", "/").split("/")[-1]
        mime = getattr(upload, "content_type", "") or ""
        data = upload.read()
        if len(data) > MAX_ATTACH_BYTES:
            raise ValueError(f"文件 {name} 超过 {MAX_ATTACH_BYTES // (1024 * 1024)}MB 上限")

        ext = Path(name.lower()).suffix
        is_image = _is_image(name, mime)
        is_text = ext in TEXT_EXTENSIONS
        is_doc = ext in DOC_EXTENSIONS
        if not is_image and not is_text and not is_doc:
            raise ValueError(
                f"暂不支持文件类型: {name}，"
                "请上传图片、文本(json/md/csv 等)或常见附件(pdf/office/zip 等)"
            )

        stored = f"{uuid.uuid4().hex}_{name}"
        path = root / stored
        path.write_bytes(data)

        item = {
            "id": stored,
            "name": name,
            "size": len(data),
            "mime": mime or (_image_mime(name) if is_image else (mimetypes.guess_type(name)[0] or "application/octet-stream")),
            "text": "",
            "has_text": False,
            "is_image": is_image,
            "is_file": not is_image,
            "stored_path": str(path),
            "url": f"/api/agent/attachments/{stored}",
        }
        if is_image:
            img_mime = _image_mime(name, mime)
            item["mime"] = img_mime
            item["image_base64"] = base64.b64encode(data).decode("ascii")
            item["data_url"] = f"data:{img_mime};base64,{item['image_base64']}"
        elif is_text or ext in SPREADSHEET_EXTENSIONS:
            text = _extract_text(name, data)
            item["text"] = text
            item["has_text"] = bool(text)

        results.append(item)
    return results


def load_stored_attachments(items: list[dict], user_id: int) -> list[dict]:
    """从可信的用户附件目录恢复模型所需文本或图片数据。"""
    loaded: list[dict] = []
    for raw in (items or [])[:MAX_ATTACH_FILES]:
        if not isinstance(raw, dict):
            continue
        item = dict(raw)
        stored_id = str(item.get("id") or "")
        path = resolve_attachment_path(user_id, stored_id)
        if path is None:
            item.update({
                "text": "（附件文件不可用或已删除，无法读取内容。）",
                "has_text": True,
                "image_base64": "",
                "data_url": "",
            })
            loaded.append(item)
            continue

        if path.stat().st_size > MAX_ATTACH_BYTES:
            item.update({
                "text": "（附件超过读取上限，无法解析内容。）",
                "has_text": True,
                "image_base64": "",
                "data_url": "",
            })
            loaded.append(item)
            continue

        data = path.read_bytes()
        name = str(item.get("name") or path.name)
        mime = str(item.get("mime") or mimetypes.guess_type(name)[0] or "application/octet-stream")
        is_image = _is_image(name, mime)
        item.update({
            "name": name,
            "size": len(data),
            "mime": mime,
            "is_image": is_image,
            "is_file": not is_image,
            "stored_path": str(path),
            "text": "",
            "has_text": False,
        })
        if is_image:
            image_mime = _image_mime(name, mime)
            encoded = base64.b64encode(data).decode("ascii")
            item.update({
                "mime": image_mime,
                "image_base64": encoded,
                "data_url": f"data:{image_mime};base64,{encoded}",
            })
        else:
            text = _extract_text(name, data)
            item.update({"text": text, "has_text": bool(text)})
        loaded.append(item)
    return loaded


def format_attachment_context(attachments: list[dict]) -> str:
    if not attachments:
        return ""
    parts = ["【用户上传附件】"]
    for item in attachments:
        parts.append(f"\n### 文件: {item.get('name', 'file')} ({item.get('size', 0)} bytes)")
        if item.get("is_image"):
            parts.append("(图片已随消息发送给视觉模型,请直接根据图像内容回答)")
            continue
        text = (item.get("text") or "").strip()
        if text:
            parts.append(text)
        else:
            parts.append("(该文件类型暂不支持自动解析,已记录文件名与大小)")
    return "\n".join(parts)


def vision_image_parts(attachments: list[dict]) -> list[dict]:
    """组装 OpenAI 兼容 vision content parts。"""
    parts: list[dict] = []
    for item in attachments:
        if not item.get("is_image"):
            continue
        data_url = item.get("data_url")
        if not data_url and item.get("image_base64"):
            mime = item.get("mime") or "image/jpeg"
            data_url = f"data:{mime};base64,{item['image_base64']}"
        if not data_url:
            continue
        parts.append({
            "type": "image_url",
            "image_url": {"url": data_url},
        })
    return parts


def attachment_public_meta(items: list[dict]) -> list[dict]:
    return [
        {
            "id": item.get("id"),
            "name": item.get("name"),
            "size": item.get("size"),
            "mime": item.get("mime"),
            "has_text": item.get("has_text"),
            "is_image": bool(item.get("is_image")),
            "is_file": bool(item.get("is_file") or not item.get("is_image")),
            "url": item.get("url") or "",
        }
        for item in items
    ]


def resolve_attachment_path(user_id: int, stored_id: str) -> Path | None:
    safe = (stored_id or "").replace("\\", "/").split("/")[-1]
    if not safe or ".." in safe:
        return None
    path = attachments_root(user_id) / safe
    if path.is_file():
        return path
    return None


def resolve_attachment_path_any(stored_id: str) -> Path | None:
    """管理员跨用户查找附件。"""
    safe = (stored_id or "").replace("\\", "/").split("/")[-1]
    if not safe or ".." in safe:
        return None
    root = Path(getattr(settings, "CHAT_ATTACHMENTS_ROOT", settings.BASE_DIR / "chat_attachments"))
    if not root.exists():
        return None
    for child in root.iterdir():
        if not child.is_dir():
            continue
        path = child / safe
        if path.is_file():
            return path
    return None
