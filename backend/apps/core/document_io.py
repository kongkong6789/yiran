"""Safe, structured text extraction for chat attachments.

The chat layer must never execute an uploaded document.  This module therefore
uses bounded, read-only parsers and returns a common result shape that makes a
successful extract distinguishable from an unsupported, encrypted, scanned, or
damaged file.
"""
from __future__ import annotations

import csv
import io
import json
import posixpath
import re
import zipfile
from dataclasses import dataclass, field
from html import unescape
from html.parser import HTMLParser
from pathlib import Path
from typing import Any
from xml.etree import ElementTree


MAX_EXTRACT_CHARS = 12_000
MAX_TEXT_SOURCE_BYTES = 4 * 1024 * 1024
MAX_CSV_ROWS = 200
MAX_CSV_COLUMNS = 40
MAX_CELL_CHARS = 500
MAX_RTF_TOKENS = 500_000
MAX_RTF_NESTING_DEPTH = 256

MAX_ZIP_ENTRIES = 512
MAX_ZIP_LIST_ENTRIES = 200
MAX_ZIP_SINGLE_ENTRY_BYTES = 32 * 1024 * 1024
MAX_ZIP_EXPANDED_BYTES = 96 * 1024 * 1024
MAX_ZIP_COMPRESSION_RATIO = 250
MAX_XML_ENTRY_BYTES = 8 * 1024 * 1024
MAX_OOXML_XML_BYTES_TOTAL = 32 * 1024 * 1024
MAX_XML_ELEMENTS_PER_ENTRY = 100_000
MAX_XML_ELEMENTS_TOTAL = 250_000

MAX_SPREADSHEET_SHEETS = 6
MAX_SPREADSHEET_ROWS = 200
MAX_SPREADSHEET_COLUMNS = 40
MAX_XLSX_PACKAGE_WORKSHEETS = 64
MAX_XLSX_WORKSHEET_ROWS = 250_000
MAX_XLSX_WORKSHEET_CELLS = 250_000
MAX_XLSX_TOTAL_CELLS = 500_000
MAX_XLSX_DIMENSION_CELLS = 5_000_000
MAX_PDF_PAGES = 80
MAX_PDF_STREAM_OUTPUT_BYTES = 16 * 1024 * 1024
MAX_PPTX_SLIDES = 100

TEXT_EXTENSIONS = {
    ".md",
    ".markdown",
    ".txt",
    ".json",
    ".jsonl",
    ".csv",
    ".tsv",
    ".log",
    ".yaml",
    ".yml",
    ".xml",
    ".html",
    ".htm",
    ".rtf",
    ".py",
    ".pyi",
    ".js",
    ".jsx",
    ".mjs",
    ".cjs",
    ".ts",
    ".tsx",
    ".java",
    ".go",
    ".rs",
    ".c",
    ".cc",
    ".cpp",
    ".cxx",
    ".h",
    ".hh",
    ".hpp",
    ".cs",
    ".php",
    ".rb",
    ".swift",
    ".kt",
    ".kts",
    ".scala",
    ".sql",
    ".sh",
    ".bash",
    ".zsh",
    ".fish",
    ".ps1",
    ".bat",
    ".cmd",
    ".css",
    ".scss",
    ".sass",
    ".less",
    ".vue",
    ".svelte",
    ".toml",
    ".ini",
    ".cfg",
    ".conf",
    ".properties",
}
TEXT_FILENAMES = {
    "dockerfile",
    "makefile",
    "gnumakefile",
    "jenkinsfile",
    "procfile",
    ".env",
    ".gitignore",
}
SPREADSHEET_EXTENSIONS = {".xls", ".xlsx", ".xlsm"}
READABLE_DOCUMENT_EXTENSIONS = {
    ".docx",
    ".pdf",
    ".pptx",
    ".zip",
    *SPREADSHEET_EXTENSIONS,
}
DOWNLOAD_ONLY_EXTENSIONS = {
    ".doc",
    ".ppt",
    ".rar",
    ".7z",
    ".tar",
    ".gz",
    ".mp3",
    ".wav",
    ".mp4",
    ".mov",
    ".avi",
    ".apk",
    ".ipa",
}
SUPPORTED_NON_IMAGE_EXTENSIONS = (
    TEXT_EXTENSIONS | READABLE_DOCUMENT_EXTENSIONS | DOWNLOAD_ONLY_EXTENSIONS
)


@dataclass(frozen=True)
class ExtractionResult:
    text: str = ""
    status: str = "success"
    error: str = ""
    truncated: bool = False
    metadata: dict[str, Any] = field(default_factory=dict)

    @property
    def has_text(self) -> bool:
        return bool(self.text.strip())

    def attachment_fields(self) -> dict[str, Any]:
        return {
            "text": self.text,
            "has_text": self.has_text,
            "extraction_status": self.status,
            "extraction_error": self.error,
            "truncated": self.truncated,
            "metadata": dict(self.metadata),
        }


class DocumentSafetyError(ValueError):
    """The file is malformed or exceeds a parser resource boundary."""


def _error_result(
    status: str,
    message: str,
    *,
    metadata: dict[str, Any] | None = None,
) -> ExtractionResult:
    return ExtractionResult(
        text=f"（{message}）",
        status=status,
        error=message,
        metadata=metadata or {},
    )


def _limit_text(
    text: str,
    *,
    truncated: bool = False,
    metadata: dict[str, Any] | None = None,
) -> ExtractionResult:
    normalized = str(text or "").replace("\x00", "").strip()
    was_truncated = truncated or len(normalized) > MAX_EXTRACT_CHARS
    if len(normalized) > MAX_EXTRACT_CHARS:
        normalized = normalized[:MAX_EXTRACT_CHARS].rstrip()
    if was_truncated and normalized:
        marker = "\n……（内容已按安全上限截断）"
        normalized = normalized[: max(0, MAX_EXTRACT_CHARS - len(marker))].rstrip() + marker
    if not normalized:
        return _error_result("empty", "文件中没有可读取的文本", metadata=metadata)
    return ExtractionResult(
        text=normalized,
        status="truncated" if was_truncated else "success",
        truncated=was_truncated,
        metadata=metadata or {},
    )


def _decode_text(data: bytes) -> tuple[str, str, bool]:
    source = data[:MAX_TEXT_SOURCE_BYTES]
    source_truncated = len(data) > len(source)
    for encoding in ("utf-8-sig", "gb18030", "utf-16", "latin-1"):
        try:
            return source.decode(encoding), encoding, source_truncated
        except (UnicodeDecodeError, UnicodeError):
            continue
    return source.decode("utf-8", errors="replace"), "utf-8-replace", source_truncated


class _VisibleHTMLParser(HTMLParser):
    _SKIPPED = {"script", "style", "noscript", "template"}
    _BREAKS = {
        "article",
        "br",
        "div",
        "footer",
        "h1",
        "h2",
        "h3",
        "h4",
        "h5",
        "h6",
        "header",
        "li",
        "p",
        "section",
        "tr",
    }

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.parts: list[str] = []
        self.skip_depth = 0

    def handle_starttag(self, tag: str, attrs) -> None:
        lowered = tag.casefold()
        if lowered in self._SKIPPED:
            self.skip_depth += 1
        elif self.skip_depth == 0 and lowered in self._BREAKS:
            self.parts.append("\n")
        elif self.skip_depth == 0 and lowered in {"td", "th"}:
            self.parts.append(" | ")

    def handle_endtag(self, tag: str) -> None:
        lowered = tag.casefold()
        if lowered in self._SKIPPED and self.skip_depth:
            self.skip_depth -= 1
        elif self.skip_depth == 0 and lowered in self._BREAKS:
            self.parts.append("\n")

    def handle_data(self, data: str) -> None:
        if self.skip_depth == 0 and data.strip():
            self.parts.append(data)

    def visible_text(self) -> str:
        value = unescape(" ".join(self.parts))
        value = re.sub(r"[ \t]+", " ", value)
        value = re.sub(r"\n[ \t]+", "\n", value)
        return re.sub(r"\n{3,}", "\n\n", value).strip()


def _extract_html(data: bytes) -> ExtractionResult:
    decoded, encoding, source_truncated = _decode_text(data)
    parser = _VisibleHTMLParser()
    try:
        parser.feed(decoded)
        parser.close()
    except Exception as exc:
        return _error_result(
            "error",
            f"HTML 解析失败：{exc}",
            metadata={"parser": "html", "encoding": encoding},
        )
    return _limit_text(
        parser.visible_text(),
        truncated=source_truncated,
        metadata={"parser": "html", "encoding": encoding, "active_content_removed": True},
    )


def _extract_json(data: bytes) -> ExtractionResult:
    decoded, encoding, source_truncated = _decode_text(data)
    metadata = {"parser": "json", "encoding": encoding}
    if source_truncated:
        return _limit_text(
            decoded,
            truncated=True,
            metadata={**metadata, "validated": False, "reason": "source_too_large"},
        )
    try:
        value = json.loads(decoded)
    except json.JSONDecodeError as exc:
        return _error_result(
            "error",
            f"JSON 解析失败：第 {exc.lineno} 行第 {exc.colno} 列格式错误",
            metadata=metadata,
        )
    rendered = json.dumps(value, ensure_ascii=False, indent=2)
    return _limit_text(rendered, metadata={**metadata, "validated": True})


def _cell_text(value: Any) -> str:
    return str(value if value is not None else "").replace("\x00", "").replace("\t", " ").strip()[
        :MAX_CELL_CHARS
    ]


def _extract_csv(data: bytes, *, delimiter: str | None = None) -> ExtractionResult:
    decoded, encoding, source_truncated = _decode_text(data)
    metadata: dict[str, Any] = {"parser": "tsv" if delimiter == "\t" else "csv", "encoding": encoding}
    if delimiter is None:
        try:
            delimiter = csv.Sniffer().sniff(decoded[:4096], delimiters=",;\t|").delimiter
        except csv.Error:
            delimiter = ","
    metadata["delimiter"] = delimiter
    lines: list[str] = []
    truncated = source_truncated
    max_columns = 0
    try:
        reader = csv.reader(io.StringIO(decoded), delimiter=delimiter)
        for index, row in enumerate(reader):
            if index >= MAX_CSV_ROWS:
                truncated = True
                break
            max_columns = max(max_columns, len(row))
            if len(row) > MAX_CSV_COLUMNS:
                truncated = True
            cells = [_cell_text(value) for value in row[:MAX_CSV_COLUMNS]]
            while cells and not cells[-1]:
                cells.pop()
            if cells:
                suffix = "\t……（其余列已省略）" if len(row) > MAX_CSV_COLUMNS else ""
                lines.append("\t".join(cells) + suffix)
    except (csv.Error, UnicodeError) as exc:
        return _error_result("error", f"表格文本解析失败：{exc}", metadata=metadata)
    metadata.update({"rows": len(lines), "max_columns_seen": max_columns})
    return _limit_text("\n".join(lines), truncated=truncated, metadata=metadata)


def _safe_xml_root(data: bytes, *, label: str) -> tuple[ElementTree.Element, int]:
    if len(data) > MAX_XML_ENTRY_BYTES:
        raise DocumentSafetyError(f"{label} XML 超过 {MAX_XML_ENTRY_BYTES // (1024 * 1024)}MB 安全上限")
    lowered = data[:4096].lower()
    if b"<!doctype" in lowered or b"<!entity" in lowered:
        raise DocumentSafetyError(f"{label} XML 包含禁止的 DTD 或实体声明")
    try:
        try:
            from defusedxml import ElementTree as SafeElementTree

            root = SafeElementTree.fromstring(data)
        except ImportError:
            root = ElementTree.fromstring(data)
    except Exception as exc:
        raise DocumentSafetyError(f"{label} XML 解析失败：{exc}") from exc
    count = 0
    for count, _ in enumerate(root.iter(), start=1):
        if count > MAX_XML_ELEMENTS_PER_ENTRY:
            raise DocumentSafetyError(f"{label} XML 元素数量超过安全上限")
    return root, count


def _safe_xml_element_count(data: bytes, *, label: str) -> int:
    """Validate XML with a clearing iterparse for package-wide preflight."""
    if len(data) > MAX_XML_ENTRY_BYTES:
        raise DocumentSafetyError(
            f"{label} XML 超过 {MAX_XML_ENTRY_BYTES // (1024 * 1024)}MB 安全上限"
        )
    lowered = data[:4096].lower()
    if b"<!doctype" in lowered or b"<!entity" in lowered:
        raise DocumentSafetyError(f"{label} XML 包含禁止的 DTD 或实体声明")
    try:
        try:
            from defusedxml import ElementTree as SafeElementTree

            iterator = SafeElementTree.iterparse(
                io.BytesIO(data),
                events=("start", "end"),
            )
        except ImportError:
            iterator = ElementTree.iterparse(
                io.BytesIO(data),
                events=("start", "end"),
            )
        count = 0
        for event, element in iterator:
            if event == "start":
                count += 1
                if count > MAX_XML_ELEMENTS_PER_ENTRY:
                    raise DocumentSafetyError(
                        f"{label} XML 元素数量超过安全上限"
                    )
            else:
                element.clear()
        return count
    except DocumentSafetyError:
        raise
    except Exception as exc:
        raise DocumentSafetyError(f"{label} XML 解析失败：{exc}") from exc


def _local_name(tag: str) -> str:
    return str(tag or "").rsplit("}", 1)[-1]


def _extract_xml(data: bytes) -> ExtractionResult:
    try:
        root, element_count = _safe_xml_root(data, label="XML")
    except DocumentSafetyError as exc:
        return _error_result("error", str(exc), metadata={"parser": "xml"})

    lines: list[str] = []
    truncated = False

    def visit(node: ElementTree.Element, path: str, depth: int) -> None:
        nonlocal truncated
        if truncated:
            return
        if depth > 64:
            truncated = True
            return
        name = _local_name(node.tag)
        current = f"{path}/{name}" if path else name
        attributes = ", ".join(f"{_local_name(key)}={value}" for key, value in node.attrib.items())
        value = re.sub(r"\s+", " ", node.text or "").strip()
        if attributes:
            lines.append(f"{current} [{attributes}]")
        if value:
            lines.append(f"{current}: {value}")
        if sum(len(line) for line in lines) > MAX_EXTRACT_CHARS:
            truncated = True
            return
        for child in list(node):
            visit(child, current, depth + 1)

    visit(root, "", 0)
    return _limit_text(
        "\n".join(lines),
        truncated=truncated,
        metadata={
            "parser": "xml",
            "root": _local_name(root.tag),
            "element_count": element_count,
        },
    )


# RTF destinations whose contents are metadata or active/display instructions,
# not user-visible document text.
_RTF_DESTINATIONS = frozenset(
    """
    aftncn aftnsep aftnsepc annotation atnauthor atndate atnicn atnid atnparent
    atnref atntime atrfend atrfstart author background bkmkend bkmkstart blipuid
    buptim category colorschememapping colortbl comment company creatim datafield
    datastore defchp defpap do doccomm docvar dptxbxtext falt fchars ffdeftext
    ffentrymcr ffexitmcr ffformat ffhelptext ffl ffname ffstat field filetbl
    fileinfo formfield fontemb fontfile fonttbl footer footerf footerl footerr
    footnote formprot generator gridtbl header headerf headerl headerr hl htmltag
    info keycode keywords latentstyles lchars levelnumbers leveltext lfolevel
    linkval list listlevel listname listoverride listoverridetable listpicture
    liststylename listtable manager mhtmltag mmath mmathpr nextfile nonesttables
    objalias objclass objdata object objname objsect objtime oldcprops oldpprops
    oldsprops oldtprops operator panose password passwordhash pgp pgptbl picprop
    pict pn pnseclvl pntext printim private propname protend protstart protusertbl
    pxe result revtbl rsidtbl rxe shp shpbottom shpbxcolumn shpbxignore shpbxmargin
    shpbxpage shptop shpinst shplid shprslt shptxt shpycolumn shpyignore shpymargin
    shpypage staticval stylesheet subject sv template title txe ud upr userprops
    wgrffmtfilter windowcaption writereservation writereservhash xe xform
    xmlattrname xmlattrvalue xmlclose xmlname xmlnstbl xmlopen
    """.split()
)
_RTF_SPECIAL = {
    "par": "\n",
    "line": "\n",
    "tab": "\t",
    "emdash": "—",
    "endash": "–",
    "bullet": "•",
    "lquote": "‘",
    "rquote": "’",
    "ldblquote": "“",
    "rdblquote": "”",
}
_RTF_TOKEN = re.compile(
    r"\\([a-zA-Z]{1,32})(-?\d{1,10})?[ ]?"
    r"|\\'([0-9a-fA-F]{2})"
    r"|\\([^a-zA-Z])"
    r"|([{}])"
    r"|[\r\n]+"
    r"|(.)",
    re.DOTALL,
)


def _extract_rtf(data: bytes) -> ExtractionResult:
    source = data[:MAX_TEXT_SOURCE_BYTES]
    source_truncated = len(source) < len(data)
    raw = source.decode("latin-1", errors="replace")
    stack: list[tuple[int, bool]] = []
    ucskip = 1
    curskip = 0
    ignorable = False
    out: list[str] = []
    output_chars = 0
    output_truncated = False

    def append_output(value: str) -> None:
        nonlocal output_chars, output_truncated
        if not value or output_truncated:
            return
        remaining = MAX_EXTRACT_CHARS - output_chars
        if remaining <= 0:
            output_truncated = True
            return
        if len(value) > remaining:
            out.append(value[:remaining])
            output_chars += remaining
            output_truncated = True
            return
        out.append(value)
        output_chars += len(value)

    try:
        for token_count, match in enumerate(_RTF_TOKEN.finditer(raw), start=1):
            if token_count > MAX_RTF_TOKENS:
                raise DocumentSafetyError("RTF token 数量超过安全上限")
            word, arg, hex_value, symbol, brace, char = match.groups()
            if brace:
                if brace == "{":
                    if len(stack) >= MAX_RTF_NESTING_DEPTH:
                        raise DocumentSafetyError("RTF 嵌套深度超过安全上限")
                    stack.append((ucskip, ignorable))
                elif stack:
                    ucskip, ignorable = stack.pop()
                else:
                    raise DocumentSafetyError("RTF 分组括号不匹配")
                continue
            if symbol:
                if symbol == "*":
                    ignorable = True
                elif symbol in "{}\\" and not ignorable:
                    append_output(symbol)
                continue
            if word:
                lowered = word.casefold()
                if lowered in _RTF_DESTINATIONS:
                    ignorable = True
                elif not ignorable and lowered in _RTF_SPECIAL:
                    append_output(_RTF_SPECIAL[lowered])
                elif lowered == "uc" and arg is not None:
                    ucskip = max(0, int(arg))
                elif lowered == "u" and arg is not None and not ignorable:
                    codepoint = int(arg)
                    if codepoint < 0:
                        codepoint += 0x10000
                    append_output(chr(codepoint))
                    curskip = ucskip
                if output_truncated:
                    break
                continue
            if ignorable:
                continue
            if hex_value:
                if curskip:
                    curskip -= 1
                else:
                    append_output(
                        bytes.fromhex(hex_value).decode("cp1252", errors="replace")
                    )
            elif char:
                if curskip:
                    curskip -= 1
                else:
                    append_output(char)
            if output_truncated:
                break
        if stack and not output_truncated:
            raise DocumentSafetyError("RTF 分组括号未闭合")
    except (DocumentSafetyError, ValueError, OverflowError) as exc:
        return _error_result("error", f"RTF 解析失败：{exc}", metadata={"parser": "rtf"})
    text = re.sub(r"[ \t]+\n", "\n", "".join(out))
    text = re.sub(r"\n{3,}", "\n\n", text)
    return _limit_text(
        text,
        truncated=source_truncated or output_truncated,
        metadata={
            "parser": "rtf",
            "output_chars": output_chars,
        },
    )


@dataclass(frozen=True)
class _ZipInspection:
    entries: dict[str, zipfile.ZipInfo]
    entry_count: int
    expanded_bytes: int


def _normalized_zip_name(value: str) -> str:
    raw = str(value or "").replace("\\", "/")
    if not raw or "\x00" in raw or raw.startswith("/") or re.match(r"^[A-Za-z]:", raw):
        raise DocumentSafetyError("ZIP 条目路径无效")
    parts = raw.split("/")
    if any(part == ".." for part in parts):
        raise DocumentSafetyError("ZIP 条目存在路径穿越")
    normalized = posixpath.normpath(raw).lstrip("./")
    if not normalized or normalized in {".", ".."} or normalized.startswith("../"):
        raise DocumentSafetyError("ZIP 条目路径无效")
    return normalized.rstrip("/")


def _inspect_zip(archive: zipfile.ZipFile) -> _ZipInspection:
    infos = archive.infolist()
    if len(infos) > MAX_ZIP_ENTRIES:
        raise DocumentSafetyError(f"ZIP 条目超过 {MAX_ZIP_ENTRIES} 个安全上限")
    entries: dict[str, zipfile.ZipInfo] = {}
    seen: set[str] = set()
    expanded = 0
    for info in infos:
        normalized = _normalized_zip_name(info.filename)
        dedupe = normalized.casefold()
        if dedupe in seen:
            raise DocumentSafetyError("ZIP 包含重复条目")
        seen.add(dedupe)
        if info.flag_bits & 0x1:
            raise DocumentSafetyError("ZIP 包含加密条目，无法安全读取")
        if info.file_size > MAX_ZIP_SINGLE_ENTRY_BYTES:
            raise DocumentSafetyError("ZIP 单个条目解压后过大")
        expanded += info.file_size
        if expanded > MAX_ZIP_EXPANDED_BYTES:
            raise DocumentSafetyError("ZIP 解压后总体积超过安全上限")
        if (
            info.file_size > 1024 * 1024
            and info.compress_size > 0
            and info.file_size / info.compress_size > MAX_ZIP_COMPRESSION_RATIO
        ):
            raise DocumentSafetyError("ZIP 压缩比异常，疑似压缩炸弹")
        if not info.is_dir():
            entries[normalized] = info
    return _ZipInspection(entries=entries, entry_count=len(infos), expanded_bytes=expanded)


def _read_zip_entry(
    archive: zipfile.ZipFile,
    info: zipfile.ZipInfo,
    *,
    limit: int,
) -> bytes:
    output = bytearray()
    try:
        with archive.open(info, "r") as source:
            while True:
                chunk = source.read(min(64 * 1024, limit - len(output) + 1))
                if not chunk:
                    break
                output.extend(chunk)
                if len(output) > limit:
                    raise DocumentSafetyError(
                        f"ZIP 条目 {info.filename} 实际解压内容超过安全上限"
                    )
        return bytes(output)
    except DocumentSafetyError:
        raise
    except (
        zipfile.BadZipFile,
        zipfile.LargeZipFile,
        NotImplementedError,
        RuntimeError,
        OSError,
        EOFError,
        ValueError,
    ) as exc:
        raise DocumentSafetyError(
            f"ZIP 条目 {info.filename} 损坏或使用了不支持的压缩方式：{exc}"
        ) from exc


def _open_safe_zip(data: bytes) -> tuple[zipfile.ZipFile, _ZipInspection]:
    try:
        archive = zipfile.ZipFile(io.BytesIO(data))
        inspection = _inspect_zip(archive)
        return archive, inspection
    except (
        zipfile.BadZipFile,
        zipfile.LargeZipFile,
        NotImplementedError,
        OSError,
        RuntimeError,
        ValueError,
    ) as exc:
        if isinstance(exc, DocumentSafetyError):
            raise
        raise DocumentSafetyError(f"ZIP/OOXML 文件损坏：{exc}") from exc


def _validated_ooxml_xml(
    archive: zipfile.ZipFile,
    inspection: _ZipInspection,
    *,
    keep: set[str] | None = None,
) -> dict[str, bytes]:
    kept: dict[str, bytes] = {}
    total_elements = 0
    total_xml_bytes = 0
    for name, info in inspection.entries.items():
        if not name.casefold().endswith((".xml", ".rels")):
            continue
        xml = _read_zip_entry(archive, info, limit=MAX_XML_ENTRY_BYTES)
        total_xml_bytes += len(xml)
        if total_xml_bytes > MAX_ZIP_EXPANDED_BYTES:
            raise DocumentSafetyError("OOXML XML 实际解压总量超过安全上限")
        count = _safe_xml_element_count(xml, label=name)
        total_elements += count
        if total_elements > MAX_XML_ELEMENTS_TOTAL:
            raise DocumentSafetyError("OOXML XML 元素总数超过安全上限")
        if total_xml_bytes > MAX_OOXML_XML_BYTES_TOTAL:
            raise DocumentSafetyError("OOXML XML 实际解压总量超过安全上限")
        if keep and name in keep:
            kept[name] = xml
    return kept


_WORKBOOK_NAMESPACES = {
    "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "http://purl.oclc.org/ooxml/spreadsheetml/main",
}
_OFFICE_RELATIONSHIP_NAMESPACES = {
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "http://purl.oclc.org/ooxml/officeDocument/relationships",
}
_PACKAGE_RELATIONSHIP_NAMESPACES = {
    "http://schemas.openxmlformats.org/package/2006/relationships",
    "http://purl.oclc.org/ooxml/package/relationships",
}
_XLSX_WORKSHEET_RELATIONSHIP_TYPES = {
    f"{namespace}/worksheet"
    for namespace in _OFFICE_RELATIONSHIP_NAMESPACES
}
_XLSX_CHARTSHEET_RELATIONSHIP_TYPES = {
    f"{namespace}/chartsheet"
    for namespace in _OFFICE_RELATIONSHIP_NAMESPACES
}


def _xml_namespace(tag: str) -> str:
    value = str(tag or "")
    return value[1:].split("}", 1)[0] if value.startswith("{") and "}" in value else ""


def _safe_ooxml_target(base_path: str, target: str) -> str:
    raw = str(target or "").replace("\\", "/")
    if not raw or "\x00" in raw or re.match(r"^[A-Za-z]:", raw):
        raise DocumentSafetyError("XLSX 工作表关系目标无效")
    if raw.startswith("/"):
        normalized = posixpath.normpath(raw.lstrip("/"))
    else:
        normalized = posixpath.normpath(posixpath.join(posixpath.dirname(base_path), raw))
    if (
        not normalized
        or normalized in {".", ".."}
        or normalized.startswith("../")
        or not normalized.startswith("xl/")
    ):
        raise DocumentSafetyError("XLSX 工作表关系目标存在路径越界")
    return normalized


def _xlsx_coordinate(value: str) -> tuple[int, int] | None:
    match = re.fullmatch(r"\$?([A-Za-z]{1,3})\$?([1-9]\d{0,6})", str(value or ""))
    if not match:
        return None
    column = 0
    for char in match.group(1).upper():
        column = column * 26 + ord(char) - ord("A") + 1
    row = int(match.group(2))
    if column > 16_384 or row > 1_048_576:
        return None
    return row, column


def _xlsx_dimension_area(reference: str) -> int:
    endpoints = str(reference or "").split(":")
    if len(endpoints) not in {1, 2}:
        raise DocumentSafetyError("XLSX 工作表维度格式无效")
    start = _xlsx_coordinate(endpoints[0])
    end = _xlsx_coordinate(endpoints[-1])
    if start is None or end is None:
        raise DocumentSafetyError("XLSX 工作表维度超出 Excel 范围")
    return (abs(end[0] - start[0]) + 1) * (abs(end[1] - start[1]) + 1)


def _xlsx_worksheet_targets(
    workbook_xml: bytes,
    relationships_xml: bytes,
    inspection: _ZipInspection,
) -> list[str]:
    workbook, _ = _safe_xml_root(workbook_xml, label="xl/workbook.xml")
    relationships, _ = _safe_xml_root(
        relationships_xml,
        label="xl/_rels/workbook.xml.rels",
    )
    if (
        _local_name(workbook.tag) != "workbook"
        or _xml_namespace(workbook.tag) not in _WORKBOOK_NAMESPACES
    ):
        raise DocumentSafetyError("XLSX workbook.xml 根节点无法识别")
    if (
        _local_name(relationships.tag) != "Relationships"
        or _xml_namespace(relationships.tag) not in _PACKAGE_RELATIONSHIP_NAMESPACES
    ):
        raise DocumentSafetyError("XLSX workbook 关系文件根节点无法识别")

    relation_map: dict[str, tuple[str, str, str]] = {}
    for relation in relationships:
        if _local_name(relation.tag) != "Relationship":
            continue
        relation_id = str(relation.attrib.get("Id") or "")
        if not relation_id or relation_id in relation_map:
            raise DocumentSafetyError("XLSX workbook 关系 ID 缺失或重复")
        relation_map[relation_id] = (
            str(relation.attrib.get("Type") or ""),
            str(relation.attrib.get("Target") or ""),
            str(relation.attrib.get("TargetMode") or ""),
        )

    targets: list[str] = []
    seen_ids: set[str] = set()
    seen_targets: set[str] = set()
    for sheet in workbook.iter():
        if _local_name(sheet.tag) != "sheet":
            continue
        relationship_ids = [
            sheet.attrib.get(f"{{{namespace}}}id")
            for namespace in _OFFICE_RELATIONSHIP_NAMESPACES
            if sheet.attrib.get(f"{{{namespace}}}id")
        ]
        if len(relationship_ids) != 1:
            raise DocumentSafetyError("XLSX 工作表关系 ID 无法识别")
        relationship_id = str(relationship_ids[0])
        if relationship_id in seen_ids:
            raise DocumentSafetyError("XLSX 工作表关系 ID 重复")
        seen_ids.add(relationship_id)
        relation = relation_map.get(relationship_id)
        if relation is None:
            raise DocumentSafetyError("XLSX 工作表关系不存在")
        relation_type, raw_target, target_mode = relation
        if target_mode.casefold() not in {"", "internal"}:
            raise DocumentSafetyError("XLSX 工作表关系 TargetMode 只能为 Internal")
        if relation_type in _XLSX_CHARTSHEET_RELATIONSHIP_TYPES:
            # Chart sheets are not cell-bearing and are not exposed through
            # ``openpyxl.workbook.worksheets``. Still require a bounded,
            # internal, existing target instead of silently accepting it.
            target = _safe_ooxml_target("xl/workbook.xml", raw_target)
            if target not in inspection.entries:
                raise DocumentSafetyError("XLSX 图表工作表关系指向不存在的 ZIP 条目")
            continue
        if relation_type not in _XLSX_WORKSHEET_RELATIONSHIP_TYPES:
            raise DocumentSafetyError("XLSX 工作表使用了不受支持的关系类型")
        target = _safe_ooxml_target("xl/workbook.xml", raw_target)
        if target not in inspection.entries:
            raise DocumentSafetyError("XLSX 工作表关系指向不存在的 ZIP 条目")
        dedupe = target.casefold()
        if dedupe in seen_targets:
            raise DocumentSafetyError("XLSX 多个工作表关系指向同一条目")
        seen_targets.add(dedupe)
        targets.append(target)
        if len(targets) > MAX_XLSX_PACKAGE_WORKSHEETS:
            raise DocumentSafetyError("XLSX 工作表数量超过安全上限")
    if not targets:
        raise DocumentSafetyError("XLSX 未包含可识别的工作表")
    return targets


def _inspect_xlsx_worksheet(xml: bytes, *, path: str) -> int:
    root, _ = _safe_xml_root(xml, label=path)
    if (
        _local_name(root.tag) != "worksheet"
        or _xml_namespace(root.tag) not in _WORKBOOK_NAMESPACES
    ):
        raise DocumentSafetyError(f"XLSX 工作表 {path} 根节点无法识别")
    rows = 0
    cells = 0
    for element in root.iter():
        local_name = _local_name(element.tag)
        if local_name == "dimension":
            reference = str(element.attrib.get("ref") or "")
            if reference and _xlsx_dimension_area(reference) > MAX_XLSX_DIMENSION_CELLS:
                raise DocumentSafetyError(f"XLSX 工作表 {path} 声明维度过大")
        elif local_name == "row":
            rows += 1
            if rows > MAX_XLSX_WORKSHEET_ROWS:
                raise DocumentSafetyError(f"XLSX 工作表 {path} 物理行数过多")
        elif local_name == "c":
            cells += 1
            if cells > MAX_XLSX_WORKSHEET_CELLS:
                raise DocumentSafetyError(f"XLSX 工作表 {path} 单元格数量过多")
    return cells


def _validate_xlsx_package(
    archive: zipfile.ZipFile,
    inspection: _ZipInspection,
) -> list[str]:
    workbook_path = "xl/workbook.xml"
    relationships_path = "xl/_rels/workbook.xml.rels"
    if workbook_path not in inspection.entries or relationships_path not in inspection.entries:
        raise DocumentSafetyError("XLSX 缺少 workbook.xml 或其关系文件")
    core = _validated_ooxml_xml(
        archive,
        inspection,
        keep={workbook_path, relationships_path},
    )
    targets = _xlsx_worksheet_targets(
        core[workbook_path],
        core[relationships_path],
        inspection,
    )
    total_cells = 0
    for target in targets:
        worksheet_xml = _read_zip_entry(
            archive,
            inspection.entries[target],
            limit=MAX_XML_ENTRY_BYTES,
        )
        total_cells += _inspect_xlsx_worksheet(worksheet_xml, path=target)
        if total_cells > MAX_XLSX_TOTAL_CELLS:
            raise DocumentSafetyError("XLSX 全部工作表单元格总数过多")
    return targets


def _extract_docx(data: bytes) -> ExtractionResult:
    try:
        archive, inspection = _open_safe_zip(data)
        with archive:
            path = "word/document.xml"
            if path not in inspection.entries:
                raise DocumentSafetyError("DOCX 缺少 word/document.xml")
            xml = _validated_ooxml_xml(archive, inspection, keep={path})[path]
        root, element_count = _safe_xml_root(xml, label="DOCX document")
    except DocumentSafetyError as exc:
        return _error_result("error", str(exc), metadata={"parser": "docx"})

    body = next((node for node in root.iter() if _local_name(node.tag) == "body"), None)
    if body is None:
        return _error_result("error", "DOCX 缺少文档正文", metadata={"parser": "docx"})
    blocks: list[str] = []
    paragraphs = 0
    tables = 0
    truncated = False
    for child in list(body):
        tag = _local_name(child.tag)
        if tag == "p":
            text = "".join(node.text or "" for node in child.iter() if _local_name(node.tag) == "t").strip()
            if text:
                blocks.append(text)
                paragraphs += 1
        elif tag == "tbl":
            rows: list[str] = []
            for row in (node for node in child.iter() if _local_name(node.tag) == "tr"):
                cells: list[str] = []
                for cell in (node for node in list(row) if _local_name(node.tag) == "tc"):
                    cell_text = " ".join(
                        filter(
                            None,
                            (
                                "".join(
                                    text.text or ""
                                    for text in paragraph.iter()
                                    if _local_name(text.tag) == "t"
                                ).strip()
                                for paragraph in cell.iter()
                                if _local_name(paragraph.tag) == "p"
                            ),
                        )
                    )
                    cells.append(cell_text)
                if any(cells):
                    rows.append(" | ".join(cells))
            if rows:
                blocks.append("\n".join(rows))
                tables += 1
        if sum(len(block) for block in blocks) > MAX_EXTRACT_CHARS:
            truncated = True
            break
    return _limit_text(
        "\n\n".join(blocks),
        truncated=truncated,
        metadata={
            "parser": "docx",
            "paragraph_count": paragraphs,
            "table_count": tables,
            "xml_element_count": element_count,
        },
    )


def _slide_number(name: str) -> int:
    match = re.search(r"/slide(\d+)\.xml$", name, flags=re.IGNORECASE)
    return int(match.group(1)) if match else 0


def _extract_pptx(data: bytes) -> ExtractionResult:
    try:
        archive, inspection = _open_safe_zip(data)
        with archive:
            slide_paths = sorted(
                (
                    name
                    for name in inspection.entries
                    if re.fullmatch(r"ppt/slides/slide\d+\.xml", name, flags=re.IGNORECASE)
                ),
                key=_slide_number,
            )
            if not slide_paths:
                raise DocumentSafetyError("PPTX 中没有可识别的幻灯片")
            selected = slide_paths[:MAX_PPTX_SLIDES]
            xml_by_path = _validated_ooxml_xml(archive, inspection, keep=set(selected))
        blocks: list[str] = []
        truncated = len(slide_paths) > len(selected)
        for index, path in enumerate(selected, start=1):
            root, _ = _safe_xml_root(xml_by_path[path], label=f"PPTX slide {index}")
            texts = [
                re.sub(r"\s+", " ", node.text or "").strip()
                for node in root.iter()
                if _local_name(node.tag) == "t" and (node.text or "").strip()
            ]
            if texts:
                blocks.append(f"## 幻灯片 {index}\n" + "\n".join(texts))
            if sum(len(block) for block in blocks) > MAX_EXTRACT_CHARS:
                truncated = True
                break
    except DocumentSafetyError as exc:
        return _error_result("error", str(exc), metadata={"parser": "pptx"})
    return _limit_text(
        "\n\n".join(blocks),
        truncated=truncated,
        metadata={
            "parser": "pptx",
            "slide_count": len(slide_paths),
            "slides_read": min(len(selected), len(blocks)),
        },
    )


def _sheet_cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\x00", "").replace("\t", " ").strip()[:MAX_CELL_CHARS]


def _sheet_text(
    title: str,
    rows,
    *,
    total_rows: int | None = None,
    total_columns: int | None = None,
) -> tuple[str, bool, int]:
    lines = [f"## 工作表: {title}"]
    emitted = 0
    truncated = False
    for raw_row in rows:
        if emitted >= MAX_SPREADSHEET_ROWS:
            truncated = True
            break
        row = list(raw_row or ())
        values = [_sheet_cell_text(value) for value in row[:MAX_SPREADSHEET_COLUMNS]]
        while values and not values[-1]:
            values.pop()
        if not any(values):
            continue
        more_columns = len(row) > MAX_SPREADSHEET_COLUMNS or bool(
            total_columns and total_columns > MAX_SPREADSHEET_COLUMNS
        )
        if more_columns:
            truncated = True
        lines.append(
            "\t".join(values) + ("\t……（其余列已省略）" if more_columns else "")
        )
        emitted += 1
    if total_rows is not None and total_rows > MAX_SPREADSHEET_ROWS:
        truncated = True
        lines.append(
            f"……（工作表共 {total_rows} 行，"
            f"仅扫描前 {MAX_SPREADSHEET_ROWS} 行以保证安全）"
        )
    if emitted == 0:
        lines.append("（工作表为空）")
    return "\n".join(lines), truncated, emitted


def _extract_xlsx(data: bytes) -> ExtractionResult:
    try:
        archive, inspection = _open_safe_zip(data)
        with archive:
            worksheet_targets = _validate_xlsx_package(archive, inspection)
    except DocumentSafetyError as exc:
        return _error_result("error", str(exc), metadata={"parser": "xlsx"})

    try:
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        try:
            actual_targets = {
                str(getattr(sheet, "_worksheet_path", "") or "").lstrip("/")
                for sheet in workbook.worksheets
            }
            if (
                not all(actual_targets)
                or actual_targets != set(worksheet_targets)
                or len(actual_targets) != len(workbook.worksheets)
            ):
                raise DocumentSafetyError(
                    "Excel 解析器读取的工作表集合与安全预检结果不一致"
                )
            blocks: list[str] = []
            row_counts: dict[str, int] = {}
            truncated = len(workbook.worksheets) > MAX_SPREADSHEET_SHEETS
            for sheet in workbook.worksheets[:MAX_SPREADSHEET_SHEETS]:
                block, sheet_truncated, emitted = _sheet_text(
                    sheet.title,
                    sheet.iter_rows(
                        min_row=1,
                        max_row=min(sheet.max_row, MAX_SPREADSHEET_ROWS),
                        max_col=min(sheet.max_column, MAX_SPREADSHEET_COLUMNS),
                        values_only=True,
                    ),
                    total_rows=sheet.max_row,
                    total_columns=sheet.max_column,
                )
                blocks.append(block)
                row_counts[sheet.title] = emitted
                truncated = truncated or sheet_truncated
        finally:
            workbook.close()
    except DocumentSafetyError as exc:
        return _error_result(
            "error",
            str(exc),
            metadata={"parser": "xlsx"},
        )
    except Exception as exc:
        return _error_result(
            "error",
            "Excel 文件解析失败，请确认文件未损坏或加密",
            metadata={"parser": "xlsx", "detail": str(exc)[:200]},
        )
    return _limit_text(
        "\n\n".join(blocks),
        truncated=truncated,
        metadata={
            "parser": "xlsx",
            "sheet_count": len(row_counts),
            "rows_read": row_counts,
            "validated_worksheet_count": len(worksheet_targets),
        },
    )


def _extract_xls(data: bytes) -> ExtractionResult:
    try:
        import xlrd

        workbook = xlrd.open_workbook(file_contents=data, on_demand=True)
        try:
            blocks: list[str] = []
            row_counts: dict[str, int] = {}
            truncated = workbook.nsheets > MAX_SPREADSHEET_SHEETS
            for sheet in workbook.sheets()[:MAX_SPREADSHEET_SHEETS]:
                rows = [
                    [
                        sheet.cell(row_index, column_index).value
                        for column_index in range(min(sheet.ncols, MAX_SPREADSHEET_COLUMNS))
                    ]
                    for row_index in range(min(sheet.nrows, MAX_SPREADSHEET_ROWS))
                ]
                block, sheet_truncated, emitted = _sheet_text(
                    sheet.name,
                    rows,
                    total_rows=sheet.nrows,
                    total_columns=sheet.ncols,
                )
                blocks.append(block)
                row_counts[sheet.name] = emitted
                truncated = truncated or sheet_truncated
        finally:
            workbook.release_resources()
    except Exception as exc:
        return _error_result(
            "error",
            "旧版 Excel 文件解析失败，请确认文件未损坏或加密",
            metadata={"parser": "xls", "detail": str(exc)[:200]},
        )
    return _limit_text(
        "\n\n".join(blocks),
        truncated=truncated,
        metadata={
            "parser": "xls",
            "sheet_count": len(row_counts),
            "rows_read": row_counts,
        },
    )


def _extract_pdf(data: bytes) -> ExtractionResult:
    try:
        from pypdf import PdfReader
        from pypdf.errors import PdfReadError
        from pypdf import filters as pdf_filters
    except ImportError:
        return _error_result(
            "unsupported",
            "PDF 解析组件尚未安装，当前只能下载该文件",
            metadata={"parser": "pdf"},
        )

    try:
        # pypdf exposes decompression guards as module constants.  Only lower
        # them: concurrent readers then share the same conservative ceiling.
        for setting_name in (
            "ZLIB_MAX_OUTPUT_LENGTH",
            "LZW_MAX_OUTPUT_LENGTH",
            "RUN_LENGTH_MAX_OUTPUT_LENGTH",
            "JBIG2_MAX_OUTPUT_LENGTH",
            "MAX_ARRAY_BASED_STREAM_OUTPUT_LENGTH",
            "MAX_DECLARED_STREAM_LENGTH",
        ):
            current = getattr(pdf_filters, setting_name, MAX_PDF_STREAM_OUTPUT_BYTES)
            setattr(
                pdf_filters,
                setting_name,
                min(int(current), MAX_PDF_STREAM_OUTPUT_BYTES),
            )
        reader = PdfReader(io.BytesIO(data), strict=False)
        if reader.is_encrypted:
            try:
                unlocked = reader.decrypt("")
            except Exception:
                unlocked = 0
            if not unlocked:
                return _error_result(
                    "encrypted",
                    "PDF 已加密，无法读取内容",
                    metadata={"parser": "pdf"},
                )
        page_count = len(reader.pages)
        selected_pages = min(page_count, MAX_PDF_PAGES)
        blocks: list[str] = []
        truncated = page_count > selected_pages
        for index in range(selected_pages):
            text = reader.pages[index].extract_text() or ""
            text = text.replace("\x00", "").strip()
            if text:
                blocks.append(f"## 第 {index + 1} 页\n{text}")
            if sum(len(block) for block in blocks) > MAX_EXTRACT_CHARS:
                truncated = True
                break
    except (PdfReadError, OSError, ValueError, TypeError, KeyError) as exc:
        return _error_result(
            "error",
            f"PDF 文件解析失败：{str(exc)[:200]}",
            metadata={"parser": "pdf"},
        )
    except Exception as exc:
        return _error_result(
            "error",
            f"PDF 内容无法安全读取：{str(exc)[:200]}",
            metadata={"parser": "pdf"},
        )
    metadata = {
        "parser": "pdf",
        "page_count": page_count,
        "pages_read": min(selected_pages, len(blocks)),
    }
    if not blocks:
        return _error_result(
            "scanned",
            "PDF 未检测到可提取文字，可能是扫描件，需要 OCR 后才能读取",
            metadata=metadata,
        )
    return _limit_text("\n\n".join(blocks), truncated=truncated, metadata=metadata)


def _extract_zip_listing(data: bytes) -> ExtractionResult:
    try:
        archive, inspection = _open_safe_zip(data)
        with archive:
            entries = sorted(inspection.entries.items(), key=lambda pair: pair[0].casefold())
            selected = entries[:MAX_ZIP_LIST_ENTRIES]
            lines = [
                f"- {name} ({info.file_size} bytes，压缩后 {info.compress_size} bytes)"
                for name, info in selected
            ]
    except DocumentSafetyError as exc:
        return _error_result("error", str(exc), metadata={"parser": "zip_listing"})
    return _limit_text(
        "ZIP 文件目录（仅列出，不解压内容）：\n" + "\n".join(lines),
        truncated=len(entries) > len(selected),
        metadata={
            "parser": "zip_listing",
            "entry_count": inspection.entry_count,
            "file_count": len(entries),
            "expanded_bytes": inspection.expanded_bytes,
            "contents_extracted": False,
        },
    )


def _extract_plain_text(data: bytes, *, parser: str) -> ExtractionResult:
    decoded, encoding, source_truncated = _decode_text(data)
    return _limit_text(
        decoded,
        truncated=source_truncated,
        metadata={"parser": parser, "encoding": encoding},
    )


def _unsupported_message(extension: str) -> str:
    if extension in {".doc", ".ppt"}:
        return "旧版 Office 格式当前不做内容解析，请转换为 DOCX 或 PPTX 后上传"
    if extension in {".rar", ".7z", ".tar", ".gz"}:
        return "该压缩格式当前仅支持下载；如需查看目录，请上传 ZIP"
    if extension in {".mp3", ".wav", ".mp4", ".mov", ".avi"}:
        return "音视频附件当前没有配置转写能力，只能下载"
    if extension in {".apk", ".ipa"}:
        return "应用安装包不做内容解析，只能下载"
    return "该文件类型当前不支持自动解析，只能下载"


def extract_document(name: str, data: bytes) -> ExtractionResult:
    """Extract a supported document into bounded, model-safe text."""
    filename = Path(str(name or "file").replace("\\", "/")).name
    extension = Path(filename.casefold()).suffix
    basename = filename.casefold()
    try:
        if extension in {".html", ".htm"}:
            return _extract_html(data)
        if extension == ".json":
            return _extract_json(data)
        if extension == ".csv":
            return _extract_csv(data)
        if extension == ".tsv":
            return _extract_csv(data, delimiter="\t")
        if extension == ".xml":
            return _extract_xml(data)
        if extension == ".rtf":
            return _extract_rtf(data)
        if extension == ".docx":
            return _extract_docx(data)
        if extension == ".pptx":
            return _extract_pptx(data)
        if extension in {".xlsx", ".xlsm"}:
            return _extract_xlsx(data)
        if extension == ".xls":
            return _extract_xls(data)
        if extension == ".pdf":
            return _extract_pdf(data)
        if extension == ".zip":
            return _extract_zip_listing(data)
        if extension in DOWNLOAD_ONLY_EXTENSIONS:
            message = _unsupported_message(extension)
            return _error_result(
                "unsupported",
                message,
                metadata={"parser": "none", "extension": extension},
            )
        if extension in TEXT_EXTENSIONS or basename in TEXT_FILENAMES:
            return _extract_plain_text(data, parser=extension.lstrip(".") or basename)
    except DocumentSafetyError as exc:
        return _error_result(
            "error",
            str(exc),
            metadata={"parser": extension.lstrip(".") or "unknown"},
        )
    return _error_result(
        "unsupported",
        _unsupported_message(extension),
        metadata={"parser": "none", "extension": extension},
    )


def is_supported_non_image_name(name: str) -> bool:
    filename = Path(str(name or "").replace("\\", "/")).name.casefold()
    return Path(filename).suffix in SUPPORTED_NON_IMAGE_EXTENSIONS or filename in TEXT_FILENAMES


__all__ = [
    "DOWNLOAD_ONLY_EXTENSIONS",
    "DocumentSafetyError",
    "ExtractionResult",
    "MAX_EXTRACT_CHARS",
    "READABLE_DOCUMENT_EXTENSIONS",
    "SPREADSHEET_EXTENSIONS",
    "SUPPORTED_NON_IMAGE_EXTENSIONS",
    "TEXT_EXTENSIONS",
    "TEXT_FILENAMES",
    "extract_document",
    "is_supported_non_image_name",
]
