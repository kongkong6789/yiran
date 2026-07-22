from pathlib import Path
from datetime import datetime, timedelta
import mimetypes
import json
import io
import re
from urllib.parse import quote

from django.contrib.auth import get_user_model
from django.http import FileResponse, HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from django.db import transaction
from django.db.models import Count, Min, Q
from django.db.models.functions import TruncDate
from rest_framework import status
from rest_framework.decorators import api_view, parser_classes, permission_classes
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response

from .agent_chat import run_chat
from .attachments import (
    attachment_public_meta,
    process_uploaded_files,
    resolve_attachment_path,
    resolve_attachment_path_any,
)
from .models import AuditLog, ChatMessage, ChatSession, TaskFollowUp, TaskResultRecord, WorkAutomation, WorkAutomationRun, WorkTask, WorkTaskArtifact, WorkTodo
from .organizations import current_organization, organization_user_ids
from .automation_scheduler import configure_next_run

User = get_user_model()


def _work_automation_payload(row: WorkAutomation) -> dict:
    return {
        "id": row.id,
        "name": row.name,
        "triggerType": row.trigger_type,
        "triggerRule": row.trigger_rule,
        "action": row.action,
        "channel": row.notification_channel,
        "recipientContactIds": row.recipient_contact_ids,
        "enabled": row.enabled,
        "nextRunAt": row.next_run_at.isoformat() if row.next_run_at else None,
        "lastRunAt": row.last_run_at.isoformat() if row.last_run_at else None,
        "lastRunStatus": row.last_run_status,
        "lastError": row.last_error,
        "runCount": row.run_count,
        "lastTestedAt": row.last_tested_at.isoformat() if row.last_tested_at else None,
        "lastTestStatus": row.last_test_status,
        "createdAt": row.created_at.isoformat(),
        "updatedAt": row.updated_at.isoformat(),
    }


def _validate_work_automation(user, data: dict, current: WorkAutomation | None = None):
    name = str(data.get("name", current.name if current else "")).strip()[:128]
    action = str(data.get("action", current.action if current else "")).strip()
    trigger_type = str(data.get("triggerType", current.trigger_type if current else WorkAutomation.TriggerType.SCHEDULE))
    trigger_rule = str(data.get("triggerRule", current.trigger_rule if current else "")).strip()[:255]
    channel = str(data.get("channel", current.notification_channel if current else WorkAutomation.NotificationChannel.NONE))
    enabled = bool(data.get("enabled", current.enabled if current else False))
    raw_recipients = data.get("recipientContactIds", current.recipient_contact_ids if current else [])
    recipient_ids = list(dict.fromkeys(int(value) for value in raw_recipients if str(value).isdigit()))
    if not name or not action or not trigger_rule:
        raise ValueError("自动化名称、触发规则和执行动作必填。")
    if trigger_type not in WorkAutomation.TriggerType.values:
        raise ValueError("触发方式无效。")
    if channel not in WorkAutomation.NotificationChannel.values:
        raise ValueError("通知方式无效。")
    if enabled and trigger_type == WorkAutomation.TriggerType.DATA and trigger_rule != "待办状态变化时":
        raise ValueError("当前仅“待办状态变化时”已接入真实事件源；该规则可先保存为停用，接入对应数据源后再启用。")
    if channel == WorkAutomation.NotificationChannel.WECOM:
        if not recipient_ids:
            raise ValueError("请选择至少一位企业微信接收人。")
        from apps.wecom.access import resolve_accessible_config
        from apps.wecom.models import WeComContact
        config = resolve_accessible_config(user)
        available = set(WeComContact.objects.filter(
            config=config, available=True, id__in=recipient_ids,
        ).values_list("id", flat=True)) if config else set()
        if available != set(recipient_ids):
            raise ValueError("部分企业微信接收人不存在、已停用或不属于当前企业。")
    return {
        "name": name,
        "trigger_type": trigger_type,
        "trigger_rule": trigger_rule,
        "action": action,
        "notification_channel": channel,
        "recipient_contact_ids": recipient_ids if channel == WorkAutomation.NotificationChannel.WECOM else [],
        "enabled": enabled,
    }


def _is_admin(user) -> bool:
    return bool(getattr(user, "is_authenticated", False) and (user.is_staff or user.is_superuser))


def _display_name(user) -> str:
    profile = getattr(user, "settings", None)
    return (getattr(profile, "display_name", "") or user.username).strip()


def _work_todo_payload(row: WorkTodo) -> dict:
    return {
        "id": str(row.public_id),
        "title": row.title,
        "description": row.description,
        "priority": row.priority,
        "priorityLabel": row.get_priority_display(),
        "status": row.status,
        "statusLabel": row.get_status_display(),
        "dueAt": row.due_at.isoformat() if row.due_at else None,
        "creator": {"id": row.creator_id, "name": _display_name(row.creator)},
        "assignee": {"id": row.assignee_id, "name": _display_name(row.assignee)},
        "syncStatus": row.sync_status,
        "syncErrorReason": row.sync_error_reason,
        "remindTypes": row.remind_types,
        "completedAt": row.completed_at.isoformat() if row.completed_at else None,
        "createdAt": row.created_at.isoformat(),
        "updatedAt": row.updated_at.isoformat(),
    }


@api_view(["GET"])
@permission_classes([AllowAny])
def health(request):
    return Response({"status": "ok", "service": "agent-saas-backend"})


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def audit_logs(request):
    """兼容用原始审计接口，仅供超级管理员访问。"""
    if not request.user.is_superuser:
        return Response({"ok": False, "detail": "仅超级管理员可查看日志。"}, status=403)
    limit = int(request.query_params.get("limit", 50))
    logs = AuditLog.objects.all()[:limit]
    data = [
        {
            "id": log.id,
            "trace_id": log.trace_id,
            "actor": log.actor,
            "intent": log.intent,
            "action": log.action,
            "payload": log.payload,
            "decision": log.decision,
            "checks": log.checks,
            "result": log.result,
            "created_at": log.created_at.isoformat(),
        }
        for log in logs
    ]
    return Response({"count": len(data), "results": data})


# ================= 超级管理员日志中心 =================

# 操作动作 -> 操作类型（用于图表分组与表格标签）
_OP_TYPE_RULES = [
    ("skill", "技能调用", ("skill", "agent", "council", "chat")),
    ("knowledge", "知识库管理", ("knowledge", "doc", "ontology", "table", "datalake", "memory")),
    ("app", "应用管理", ("wecom", "connector", "mcp", "app", "automation", "webhook")),
    ("user", "用户管理", ("user", "auth", "account", "login", "logout", "register", "password", "avatar")),
    ("role", "角色与权限", ("org", "organization", "team", "role", "perm", "member")),
    ("system", "系统设置", ("system", "config", "setting")),
]

_DECISION_STATUS = {
    "allow": ("success", "成功"),
    "block": ("failed", "失败"),
    "need_approval": ("pending", "待审批"),
    "dry_run": ("dryrun", "演练"),
}

# 顶部页签 -> 过滤条件
_LOG_CATEGORIES = {"operation", "login", "system", "security", "data_change"}


def _classify_op_type(action: str, intent: str = "") -> tuple[str, str]:
    text = f"{action or ''} {intent or ''}".lower()
    for key, label, needles in _OP_TYPE_RULES:
        if any(needle in text for needle in needles):
            return key, label
    return "other", "其他操作"


def _apply_log_category(qs, category: str):
    if category == "login":
        return qs.filter(Q(action__icontains="login") | Q(action__icontains="logout") | Q(action__icontains="auth"))
    if category == "system":
        return qs.filter(Q(action__icontains="system") | Q(action__icontains="config") | Q(action__icontains="setting"))
    if category == "security":
        # 安全相关：被拦截/待审批的动作，以及登录失败、成员移除、角色/权限调整等敏感操作
        return qs.filter(
            Q(decision__in=["block", "need_approval"])
            | Q(action__icontains="login_failed")
            | Q(action__icontains="remove_member")
            | Q(action__icontains="assign_users")
            | Q(action__icontains="role")
            | Q(action__icontains="perm")
            | Q(action__icontains="password")
        )
    if category == "data_change":
        return qs.filter(
            Q(action__icontains="create") | Q(action__icontains="update") | Q(action__icontains="delete")
            | Q(action__icontains="add") | Q(action__icontains="remove") | Q(action__icontains="config")
        )
    return qs


def _parse_day(value: str):
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except (TypeError, ValueError):
        return None


def _kpi(value: int, previous: int) -> dict:
    if previous > 0:
        delta = round((value - previous) / previous * 100, 1)
    else:
        delta = 100.0 if value > 0 else 0.0
    return {"value": value, "deltaPct": abs(delta), "trend": "up" if delta > 0 else "down" if delta < 0 else "flat"}


def _resource_from_payload(payload: dict) -> str:
    if not isinstance(payload, dict):
        return ""
    for key in ("name", "title", "resource_name", "organization_name", "team_name", "target"):
        if payload.get(key):
            return str(payload[key])[:80]
    for key in ("organization_id", "team_id", "user_id", "id"):
        if payload.get(key) not in (None, ""):
            return f"#{payload[key]}"
    return ""


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def audit_overview(request):
    """超级管理员日志中心：概览指标 + 趋势 + 类型分布 + TOP 用户 + 明细分页。"""
    if not request.user.is_superuser:
        return Response({"ok": False, "detail": "仅超级管理员可查看日志。"}, status=403)

    User = get_user_model()
    params = request.query_params
    category = params.get("category") or "operation"
    if category not in _LOG_CATEGORIES:
        category = "operation"

    today = timezone.localdate()
    end_day = _parse_day(params.get("end") or "") or today
    start_day = _parse_day(params.get("start") or "") or (end_day - timedelta(days=29))
    if start_day > end_day:
        start_day, end_day = end_day, start_day
    span_days = (end_day - start_day).days + 1

    tz = timezone.get_current_timezone()
    start_dt = timezone.make_aware(datetime.combine(start_day, datetime.min.time()), tz)
    end_dt = timezone.make_aware(datetime.combine(end_day, datetime.max.time()), tz)
    prev_start_dt = start_dt - timedelta(days=span_days)

    # 概览/图表基线：页签 + 日期范围
    base = _apply_log_category(AuditLog.objects.all(), category)
    ranged = base.filter(created_at__gte=start_dt, created_at__lte=end_dt)
    prev = base.filter(created_at__gte=prev_start_dt, created_at__lt=start_dt)

    total_ops = ranged.count()
    error_ops = ranged.filter(decision="block").count()
    sensitive_ops = ranged.filter(decision="need_approval").count()
    active_users = ranged.values("actor").distinct().count()

    prev_total = prev.count()
    prev_error = prev.filter(decision="block").count()
    prev_sensitive = prev.filter(decision="need_approval").count()
    prev_active = prev.values("actor").distinct().count()

    total_users = User.objects.count()
    new_users = User.objects.filter(date_joined__gte=start_dt, date_joined__lte=end_dt).count()
    prev_new_users = User.objects.filter(date_joined__gte=prev_start_dt, date_joined__lt=start_dt).count()

    kpis = {
        "totalOps": _kpi(total_ops, prev_total),
        "totalUsers": {"value": total_users, **{k: v for k, v in _kpi(new_users, prev_new_users).items() if k != "value"}},
        "errorOps": _kpi(error_ops, prev_error),
        "sensitiveOps": _kpi(sensitive_ops, prev_sensitive),
        "activeUsers": _kpi(active_users, prev_active),
    }

    # 趋势：按天补零
    trend_map = {
        row["day"].isoformat(): row["c"]
        for row in ranged.annotate(day=TruncDate("created_at")).values("day").annotate(c=Count("id"))
        if row["day"]
    }
    trend = []
    cursor = start_day
    while cursor <= end_day:
        iso = cursor.isoformat()
        trend.append({"date": iso, "count": trend_map.get(iso, 0)})
        cursor += timedelta(days=1)

    # 类型分布
    dist_counter: dict[str, dict] = {}
    for item in ranged.values("action", "intent").annotate(c=Count("id")):
        key, label = _classify_op_type(item["action"], item["intent"])
        bucket = dist_counter.setdefault(key, {"key": key, "label": label, "count": 0})
        bucket["count"] += item["c"]
    distribution = sorted(dist_counter.values(), key=lambda x: x["count"], reverse=True)
    for entry in distribution:
        entry["pct"] = round(entry["count"] / total_ops * 100, 1) if total_ops else 0.0

    # TOP 操作用户
    top_rows = list(ranged.values("actor").annotate(c=Count("id")).order_by("-c")[:5])
    actor_names = {r["actor"] for r in top_rows}

    # 明细：叠加类型/状态/用户/关键字过滤 + 分页
    rows_qs = ranged
    op_type = params.get("type") or ""
    if op_type and op_type != "all":
        needles = next((r[2] for r in _OP_TYPE_RULES if r[0] == op_type), None)
        if needles:
            cond = Q()
            for needle in needles:
                cond |= Q(action__icontains=needle) | Q(intent__icontains=needle)
            rows_qs = rows_qs.filter(cond)
        elif op_type == "other":
            known = [n for r in _OP_TYPE_RULES for n in r[2]]
            cond = Q()
            for needle in known:
                cond |= Q(action__icontains=needle) | Q(intent__icontains=needle)
            rows_qs = rows_qs.exclude(cond)
    status_filter = params.get("status") or ""
    decision_by_status = {v[0]: k for k, v in _DECISION_STATUS.items()}
    if status_filter and status_filter != "all" and status_filter in decision_by_status:
        rows_qs = rows_qs.filter(decision=decision_by_status[status_filter])
    user_filter = params.get("user") or ""
    if user_filter and user_filter != "all":
        rows_qs = rows_qs.filter(actor=user_filter)
    keyword = (params.get("q") or "").strip()
    if keyword:
        rows_qs = rows_qs.filter(
            Q(actor__icontains=keyword) | Q(intent__icontains=keyword)
            | Q(action__icontains=keyword) | Q(trace_id__icontains=keyword)
        )

    try:
        page = max(1, int(params.get("page", 1)))
    except ValueError:
        page = 1
    try:
        page_size = min(100, max(5, int(params.get("pageSize", 10))))
    except ValueError:
        page_size = 10
    total_rows = rows_qs.count()
    offset = (page - 1) * page_size
    page_rows = list(rows_qs.order_by("-created_at")[offset:offset + page_size])

    # 用户展示信息（display_name / 角色 / 头像首字母）
    needed_usernames = actor_names | {r.actor for r in page_rows}
    users_by_name = {}
    for u in User.objects.filter(username__in=needed_usernames).select_related("settings"):
        display = (getattr(getattr(u, "settings", None), "display_name", "") or u.username).strip()
        if u.is_superuser:
            role_label = "超级管理员"
        elif u.is_staff:
            role_label = "管理员"
        else:
            role_label = "成员"
        users_by_name[u.username] = {
            "name": display,
            "roleLabel": role_label,
            "avatarUrl": getattr(getattr(u, "settings", None), "avatar_url", "") or "",
        }

    def _actor_info(actor: str) -> dict:
        if actor in users_by_name:
            return users_by_name[actor]
        if actor in ("agent", "system", "cron", ""):
            return {"name": actor or "系统", "roleLabel": "系统", "avatarUrl": ""}
        return {"name": actor, "roleLabel": "成员", "avatarUrl": ""}

    top_users = [
        {**_actor_info(r["actor"]), "actor": r["actor"], "count": r["c"]}
        for r in top_rows
    ]

    def _row_payload(log: AuditLog) -> dict:
        type_key, type_label = _classify_op_type(log.action, log.intent)
        status_key, status_label = _DECISION_STATUS.get(log.decision, ("dryrun", log.decision or "-"))
        payload = log.payload if isinstance(log.payload, dict) else {}
        return {
            "id": log.id,
            "time": timezone.localtime(log.created_at).strftime("%Y-%m-%d %H:%M:%S"),
            "user": _actor_info(log.actor),
            "operationType": {"key": type_key, "label": type_label},
            "content": log.intent or log.action or "-",
            "detail": log.action or "",
            "resourceType": type_label,
            "resourceName": _resource_from_payload(payload) or "-",
            "ip": str(payload.get("ip") or payload.get("ip_address") or "-"),
            "status": {"key": status_key, "label": status_label},
            "traceId": log.trace_id,
            "decision": log.decision,
            "payload": log.payload if isinstance(log.payload, dict) else {},
            "checks": log.checks if isinstance(log.checks, list) else [],
            "result": log.result if isinstance(log.result, dict) else {},
        }

    rows = [_row_payload(log) for log in page_rows]

    # 过滤下拉的可选项
    all_actors = list(base.values_list("actor", flat=True).distinct()[:200])
    user_options = [
        {"value": actor, "label": _actor_info(actor)["name"]}
        for actor in sorted(set(all_actors))
        if actor
    ]

    return Response({
        "ok": True,
        "range": {"start": start_day.isoformat(), "end": end_day.isoformat()},
        "kpis": kpis,
        "trend": trend,
        "distribution": distribution,
        "topUsers": top_users,
        "rows": rows,
        "pagination": {"page": page, "pageSize": page_size, "total": total_rows},
        "filters": {
            "operationTypes": [{"value": r[0], "label": r[1]} for r in _OP_TYPE_RULES] + [{"value": "other", "label": "其他操作"}],
            "statuses": [{"value": v[0], "label": v[1]} for v in _DECISION_STATUS.values()],
            "users": user_options,
        },
    })


def _task_result_payload(row: TaskResultRecord) -> dict:
    return {
        "id": row.id, "traceId": row.trace_id, "sopId": row.sop_id, "status": row.status,
        "title": row.title, "snapshot": row.snapshot, "resolvedAttentionIds": row.resolved_attention_ids,
        "followUps": [{"id": item.id, "title": item.title, "description": item.description, "status": item.status, "dueAt": item.due_at.isoformat() if item.due_at else None, "createdAt": item.created_at.isoformat()} for item in row.follow_ups.all()],
        "createdAt": row.created_at.isoformat(), "updatedAt": row.updated_at.isoformat(),
    }


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def task_results(request):
    trace_id = str(request.data.get("traceId") or "").strip()[:64]
    snapshot = request.data.get("snapshot")
    if not trace_id or not isinstance(snapshot, dict):
        return Response({"ok": False, "detail": "traceId 和 snapshot 必填。"}, status=400)
    row, _ = TaskResultRecord.objects.update_or_create(
        user=request.user, trace_id=trace_id,
        defaults={"sop_id": str(request.data.get("sopId") or "")[:128], "status": str(request.data.get("status") or "success")[:32], "title": str(request.data.get("title") or "任务结果")[:255], "snapshot": snapshot},
    )
    return Response({"ok": True, "result": _task_result_payload(row)}, status=201)


def _owned_task_result(request, trace_id: str):
    return get_object_or_404(TaskResultRecord.objects.prefetch_related("follow_ups"), user=request.user, trace_id=trace_id)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def task_result_detail(request, trace_id: str):
    return Response({"ok": True, "result": _task_result_payload(_owned_task_result(request, trace_id))})


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def task_result_export(request, trace_id: str):
    row = _owned_task_result(request, trace_id)
    payload = json.dumps(_task_result_payload(row), ensure_ascii=False, indent=2)
    response = HttpResponse(payload, content_type="application/json; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="task-result-{row.trace_id}.json"'
    return response


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def task_result_follow_up(request, trace_id: str):
    row = _owned_task_result(request, trace_id)
    title = str(request.data.get("title") or "").strip()[:255]
    if not title:
        return Response({"ok": False, "detail": "跟进任务标题必填。"}, status=400)
    follow = TaskFollowUp.objects.create(result=row, creator=request.user, title=title, description=str(request.data.get("description") or "")[:4000])
    return Response({"ok": True, "followUp": {"id": follow.id, "title": follow.title, "status": follow.status}}, status=201)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def task_attention_resolve(request, trace_id: str, item_id: str):
    row = _owned_task_result(request, trace_id)
    resolved = list(row.resolved_attention_ids or [])
    if item_id not in resolved:
        resolved.append(item_id)
        row.resolved_attention_ids = resolved
        row.save(update_fields=["resolved_attention_ids", "updated_at"])
    return Response({"ok": True, "resolvedAttentionIds": resolved})


def _work_task_payload(row: WorkTask) -> dict:
    sender_settings = getattr(row.sender, "settings", None)
    assignee_members = []
    for user in row.assignees.all():
        profile = getattr(user, "settings", None)
        assignee_members.append({
            "id": user.id,
            "name": (getattr(profile, "display_name", "") or user.get_full_name() or user.username).strip(),
            "avatarUrl": getattr(profile, "avatar_url", "") or "",
        })
    return {
        "id": row.id,
        "traceId": row.trace_id,
        "title": row.title,
        "sopId": row.sop_id,
        "senderId": row.sender_id,
        "sender": row.sender.get_full_name() or row.sender.username,
        "senderAvatar": getattr(sender_settings, "avatar_url", "") or "",
        "assigneeIds": list(row.assignees.values_list("id", flat=True)),
        "assignees": row.assignee_names,
        "assigneeMembers": assignee_members,
        "assigneeWeComUserIds": row.assignee_wecom_userids,
        "agentName": row.agent_name,
        "deadline": row.deadline.isoformat() if row.deadline else None,
        "priority": row.priority,
        "priorityLabel": row.get_priority_display(),
        "status": row.status,
        "statusLabel": row.get_status_display(),
        "progress": row.progress,
        "notificationMode": row.notification_mode,
        "notificationTarget": row.notification_target,
        "notificationStatus": row.notification_status,
        "notificationRecordId": row.notification_record_id,
        "timeline": row.timeline,
        "artifacts": [_work_task_artifact_payload(item) for item in row.artifacts.all()],
        "createdAt": row.created_at.isoformat(),
        "updatedAt": row.updated_at.isoformat(),
    }


def _work_task_artifact_payload(row: WorkTaskArtifact) -> dict:
    base = f"/tasks/{row.task.trace_id}/artifacts/{row.id}"
    return {
        "id": row.id,
        "name": row.name,
        "filename": row.filename,
        "type": "document" if row.kind == WorkTaskArtifact.Kind.MARKDOWN else "data" if row.kind == WorkTaskArtifact.Kind.JSON else "file",
        "format": row.get_kind_display(),
        "size": f"{max(1, round(row.size / 1024))} KB",
        "created_at": row.created_at.isoformat(),
        "preview_url": f"{base}/preview/" if row.kind != WorkTaskArtifact.Kind.XLSX else "",
        "download_url": f"{base}/download/",
    }


def _safe_filename(value: str, extension: str) -> str:
    stem = re.sub(r'[\\/:*?"<>|\r\n]+', "_", value).strip(" ._")[:80] or "任务产物"
    return f"{stem}.{extension}"


def _flatten(prefix: str, value, rows: list[tuple[str, str]]) -> None:
    if isinstance(value, dict):
        for key, item in value.items():
            _flatten(f"{prefix}.{key}" if prefix else str(key), item, rows)
    elif isinstance(value, list):
        rows.append((prefix, json.dumps(value, ensure_ascii=False)))
    else:
        rows.append((prefix, "" if value is None else str(value)))


def _markdown_table(rows: list[tuple[str, str]]) -> list[str]:
    lines = ["| 项目 | 内容 |", "| --- | --- |"]
    for key, value in rows:
        lines.append(f"| {key} | {value or '—'} |")
    return lines


def _humanize_scalar(value) -> str:
    if value is None:
        return "—"
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (int, float)):
        return f"{value:,}" if isinstance(value, int) else str(value)
    return str(value)


def _humanize_mapping(value: dict, depth: int = 0) -> list[str]:
    lines: list[str] = []
    for key, item in value.items():
        label = str(key)
        if isinstance(item, dict):
            lines.append(f"### {label}")
            lines.extend(_humanize_mapping(item, depth + 1))
        elif isinstance(item, list):
            lines.append(f"### {label}")
            if not item:
                lines.append("- 无")
            elif all(not isinstance(entry, (dict, list)) for entry in item):
                for entry in item:
                    lines.append(f"- {_humanize_scalar(entry)}")
            else:
                for index, entry in enumerate(item, start=1):
                    lines.append(f"#### 第 {index} 项")
                    if isinstance(entry, dict):
                        lines.extend(_markdown_table([(str(k), _humanize_scalar(v) if not isinstance(v, (dict, list)) else json.dumps(v, ensure_ascii=False)) for k, v in entry.items()]))
                        lines.append("")
                    else:
                        lines.append(f"- {_humanize_scalar(entry)}")
        else:
            lines.append(f"- **{label}**：{_humanize_scalar(item)}")
    return lines


def _generate_work_task_artifacts(row: WorkTask, parameters: dict, result_data: dict) -> list[WorkTaskArtifact]:
    from openpyxl import Workbook

    summary = [
        f"# {row.title}",
        "",
        "## 任务概览",
        "",
        *_markdown_table([
            ("任务编号", row.trace_id),
            ("SOP", row.sop_id or "未匹配"),
            ("执行智能体", row.agent_name or "未设置"),
            ("优先级", row.get_priority_display()),
            ("截止时间", row.deadline.isoformat() if row.deadline else "未设置"),
            ("负责人", "、".join(row.assignee_names) or "未设置"),
        ]),
        "",
        "## 执行参数",
        "",
    ]
    if parameters:
        summary.extend(_humanize_mapping(parameters))
    else:
        summary.append("- 无额外参数")
    summary.extend(["", "## 执行结果", ""])
    if result_data:
        summary.extend(_humanize_mapping(result_data))
    else:
        summary.append("- 暂无结果数据")
    summary.extend([
        "",
        "> 如需查看完整原始数据，请下载 JSON 或 Excel 附件。",
    ])
    markdown = "\n".join(summary).encode("utf-8")
    json_bytes = json.dumps({
        "traceId": row.trace_id, "task": row.title, "sopId": row.sop_id,
        "priority": row.priority, "parameters": parameters, "result": result_data,
    }, ensure_ascii=False, indent=2).encode("utf-8")

    workbook = Workbook()
    info = workbook.active
    info.title = "任务信息"
    for values in [
        ("字段", "内容"), ("任务", row.title), ("Trace ID", row.trace_id),
        ("SOP", row.sop_id), ("执行智能体", row.agent_name),
        ("优先级", row.get_priority_display()), ("负责人", "、".join(row.assignee_names)),
        ("截止时间", row.deadline.isoformat() if row.deadline else ""),
    ]:
        info.append(values)
    params_sheet = workbook.create_sheet("执行参数")
    params_sheet.append(("字段", "值"))
    flattened: list[tuple[str, str]] = []
    _flatten("", parameters, flattened)
    for values in flattened:
        params_sheet.append(values)
    result_sheet = workbook.create_sheet("执行结果")
    result_sheet.append(("字段", "值"))
    flattened = []
    _flatten("", result_data, flattened)
    for values in flattened:
        result_sheet.append(values)
    stream = io.BytesIO()
    workbook.save(stream)
    artifacts = [
        (WorkTaskArtifact.Kind.MARKDOWN, "任务执行报告", _safe_filename(row.title, "md"), "text/markdown; charset=utf-8", markdown),
        (WorkTaskArtifact.Kind.JSON, "任务原始数据", _safe_filename(row.title, "json"), "application/json; charset=utf-8", json_bytes),
        (WorkTaskArtifact.Kind.XLSX, "任务执行数据", _safe_filename(row.title, "xlsx"), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", stream.getvalue()),
    ]
    result = []
    for kind, name, filename, content_type, content in artifacts:
        artifact, _ = WorkTaskArtifact.objects.update_or_create(
            task=row, kind=kind,
            defaults={"name": name, "filename": filename, "content_type": content_type, "content": content, "size": len(content)},
        )
        result.append(artifact)
    return result


def _iso(dt) -> str | None:
    return dt.isoformat() if dt else None


def _work_automation_payload(row: WorkAutomation) -> dict:
    return {
        "id": row.id,
        "name": row.name,
        "triggerType": row.trigger_type,
        "triggerRule": row.trigger_rule,
        "action": row.action,
        "channel": row.notification_channel,
        "recipientContactIds": list(row.recipient_contact_ids or []),
        "enabled": bool(row.enabled),
        "nextRunAt": _iso(row.next_run_at),
        "lastRunAt": _iso(row.last_run_at),
        "lastRunStatus": row.last_run_status or "",
        "lastError": row.last_error or "",
        "runCount": row.run_count,
        "lastTestedAt": _iso(row.last_tested_at),
        "lastTestStatus": row.last_test_status or "",
        "createdAt": _iso(row.created_at) or "",
        "updatedAt": _iso(row.updated_at) or "",
    }


@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def work_automations(request):
    organization = current_organization(request.user)
    if organization is None:
        return Response({"ok": False, "detail": "未加入组织，无法管理自动化。"}, status=400)

    base_qs = WorkAutomation.objects.filter(organization=organization, creator=request.user)

    if request.method == "GET":
        rows = list(base_qs.order_by("-updated_at", "-id")[:200])
        start = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
        today_runs = WorkAutomationRun.objects.filter(
            organization=organization,
            creator=request.user,
            started_at__gte=start,
        ).count()
        enabled_rows = [r for r in rows if r.enabled]
        next_candidates = sorted(
            [r.next_run_at for r in enabled_rows if r.next_run_at],
            key=lambda x: x,
        )
        stats = {
            "saved": len(rows),
            "enabled": len(enabled_rows),
            "nextRunAt": _iso(next_candidates[0]) if next_candidates else None,
            "todayRuns": today_runs,
        }
        return Response({
            "ok": True,
            "count": len(rows),
            "stats": stats,
            "results": [_work_automation_payload(row) for row in rows],
        })

    data = request.data or {}
    name = str(data.get("name") or "").strip()[:128]
    trigger_type = str(data.get("triggerType") or WorkAutomation.TriggerType.SCHEDULE).strip()
    trigger_rule = str(data.get("triggerRule") or "").strip()[:255]
    action = str(data.get("action") or "").strip()
    channel = str(data.get("channel") or WorkAutomation.NotificationChannel.NONE).strip()
    enabled = bool(data.get("enabled", False))
    recipient_ids = data.get("recipientContactIds") or []
    if not isinstance(recipient_ids, list):
        recipient_ids = []
    recipient_ids = [int(x) for x in recipient_ids if str(x).isdigit() or isinstance(x, int)]

    if not name:
        return Response({"ok": False, "detail": "name 必填。"}, status=400)
    if not action:
        return Response({"ok": False, "detail": "action 必填。"}, status=400)
    if trigger_type not in WorkAutomation.TriggerType.values:
        return Response({"ok": False, "detail": "triggerType 无效。"}, status=400)
    if channel not in WorkAutomation.NotificationChannel.values:
        return Response({"ok": False, "detail": "channel 无效。"}, status=400)
    if not trigger_rule and trigger_type != WorkAutomation.TriggerType.MANUAL:
        return Response({"ok": False, "detail": "triggerRule 必填。"}, status=400)

    row = WorkAutomation(
        organization=organization,
        creator=request.user,
        name=name,
        trigger_type=trigger_type,
        trigger_rule=trigger_rule or "手动",
        action=action,
        notification_channel=channel,
        recipient_contact_ids=recipient_ids,
        enabled=enabled,
    )
    try:
        configure_next_run(row)
    except ValueError as exc:
        return Response({"ok": False, "detail": str(exc)}, status=400)
    row.save()
    return Response({"ok": True, "automation": _work_automation_payload(row)}, status=201)


@api_view(["GET", "PATCH", "DELETE"])
@permission_classes([IsAuthenticated])
def work_automation_detail(request, automation_id: int):
    organization = current_organization(request.user)
    if organization is None:
        return Response({"ok": False, "detail": "未加入组织，无法管理自动化。"}, status=400)

    row = get_object_or_404(
        WorkAutomation,
        id=automation_id,
        organization=organization,
        creator=request.user,
    )

    if request.method == "GET":
        return Response({"ok": True, "automation": _work_automation_payload(row)})

    if request.method == "DELETE":
        row.delete()
        return Response({"ok": True})

    data = request.data or {}
    if "name" in data:
        name = str(data.get("name") or "").strip()[:128]
        if not name:
            return Response({"ok": False, "detail": "name 不能为空。"}, status=400)
        row.name = name
    if "triggerType" in data:
        trigger_type = str(data.get("triggerType") or "").strip()
        if trigger_type not in WorkAutomation.TriggerType.values:
            return Response({"ok": False, "detail": "triggerType 无效。"}, status=400)
        row.trigger_type = trigger_type
    if "triggerRule" in data:
        row.trigger_rule = str(data.get("triggerRule") or "").strip()[:255]
    if "action" in data:
        action = str(data.get("action") or "").strip()
        if not action:
            return Response({"ok": False, "detail": "action 不能为空。"}, status=400)
        row.action = action
    if "channel" in data:
        channel = str(data.get("channel") or "").strip()
        if channel not in WorkAutomation.NotificationChannel.values:
            return Response({"ok": False, "detail": "channel 无效。"}, status=400)
        row.notification_channel = channel
    if "recipientContactIds" in data:
        recipient_ids = data.get("recipientContactIds") or []
        if not isinstance(recipient_ids, list):
            recipient_ids = []
        row.recipient_contact_ids = [
            int(x) for x in recipient_ids if str(x).isdigit() or isinstance(x, int)
        ]
    if "enabled" in data:
        row.enabled = bool(data.get("enabled"))

    try:
        configure_next_run(row)
    except ValueError as exc:
        return Response({"ok": False, "detail": str(exc)}, status=400)
    row.save()
    return Response({"ok": True, "automation": _work_automation_payload(row)})


@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def work_tasks(request):
    if request.method == "GET":
        view = str(request.query_params.get("view") or "sent")
        rows = WorkTask.objects.select_related("sender", "sender__settings").prefetch_related("assignees__settings", "artifacts")
        rows = rows.filter(assignees=request.user) if view == "received" else rows.filter(sender=request.user)
        rows = rows.distinct()[:100]
        return Response({"ok": True, "count": len(rows), "results": [_work_task_payload(row) for row in rows]})

    trace_id = str(request.data.get("traceId") or "").strip()[:64]
    title = str(request.data.get("title") or "").strip()[:500]
    priority = str(request.data.get("priority") or WorkTask.Priority.NORMAL)
    if not trace_id or not title:
        return Response({"ok": False, "detail": "traceId 和 title 必填。"}, status=400)
    if priority not in WorkTask.Priority.values:
        return Response({"ok": False, "detail": "优先级无效。"}, status=400)
    deadline_raw = str(request.data.get("deadline") or "").strip()
    deadline = parse_datetime(deadline_raw) if deadline_raw else None
    if deadline and timezone.is_naive(deadline):
        deadline = timezone.make_aware(deadline)
    from apps.wecom.access import resolve_accessible_config
    from apps.wecom.models import WeComContact
    contact_ids = [int(value) for value in request.data.get("recipientContactIds", []) if str(value).isdigit()]
    raw_assignee_contact_ids = request.data.get("assigneeWeComContactIds")
    assignee_contact_ids = (
        [int(value) for value in raw_assignee_contact_ids if str(value).isdigit()]
        if isinstance(raw_assignee_contact_ids, list)
        else contact_ids
    )
    config = resolve_accessible_config(request.user)
    contacts = list(WeComContact.objects.filter(
        config=config, available=True, id__in=assignee_contact_ids,
    )) if config and assignee_contact_ids else []
    if assignee_contact_ids and len(contacts) != len(set(assignee_contact_ids)):
        return Response({"ok": False, "detail": "部分企业微信负责人不存在、已停用或不属于当前企业。"}, status=400)
    contact_by_id = {item.id: item for item in contacts}
    if assignee_contact_ids:
        assignee_userids = [contact_by_id[item].wecom_userid for item in assignee_contact_ids]
    else:
        legacy_values = request.data.get("assigneeWeComUserIds") or request.data.get("recipientUserIds") or []
        requested_userids = list(dict.fromkeys(str(value) for value in legacy_values if str(value)))
        allowed_userids = set(WeComContact.objects.filter(
            config=config, available=True, wecom_userid__in=requested_userids,
        ).values_list("wecom_userid", flat=True)) if config else set()
        if requested_userids and len(allowed_userids) != len(requested_userids):
            return Response({"ok": False, "detail": "部分企业微信负责人不存在、已停用或不属于当前企业。"}, status=400)
        assignee_userids = requested_userids
    assignee_names = [str(value)[:128] for value in request.data.get("assigneeNames", []) if str(value)]
    timeline = request.data.get("timeline") if isinstance(request.data.get("timeline"), list) else []
    row, _ = WorkTask.objects.update_or_create(
        sender=request.user,
        trace_id=trace_id,
        defaults={
            "title": title,
            "sop_id": str(request.data.get("sopId") or "")[:128],
            "agent_name": str(request.data.get("agentName") or "")[:128],
            "priority": priority,
            "deadline": deadline,
            "status": WorkTask.Status.RUNNING,
            "progress": min(max(int(request.data.get("progress") or 75), 0), 100),
            "assignee_wecom_userids": assignee_userids,
            "assignee_names": assignee_names,
            "notification_mode": str(request.data.get("notificationMode") or "")[:16],
            "notification_target": str(request.data.get("notificationTarget") or "")[:500],
            "timeline": timeline,
        },
    )
    from apps.wecom.models import UserWeComBinding
    organization = current_organization(request.user)
    bindings = UserWeComBinding.objects.filter(
        wecom_userid__in=assignee_userids,
        status=UserWeComBinding.Status.MATCHED,
    )
    if organization is not None:
        bindings = bindings.filter(
            platform_user_id__in=organization_user_ids(organization),
            wecom_config__organization=organization,
        )
    else:
        bindings = bindings.none()
    platform_ids = bindings.values_list("platform_user_id", flat=True)
    row.assignees.set(platform_ids)
    parameters = request.data.get("parameters") if isinstance(request.data.get("parameters"), dict) else {}
    result_data = request.data.get("resultData") if isinstance(request.data.get("resultData"), dict) else {}
    if request.data.get("generateArtifacts", True):
        _generate_work_task_artifacts(row, parameters, result_data)
    return Response({"ok": True, "task": _work_task_payload(row)}, status=201)


@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def work_automations(request):
    organization = current_organization(request.user)
    if organization is None:
        return Response({"ok": False, "detail": "当前账号尚未加入企业。"}, status=400)
    rows = WorkAutomation.objects.filter(organization=organization, creator=request.user)
    if request.method == "GET":
        local_now = timezone.localtime()
        day_start = timezone.make_aware(datetime.combine(local_now.date(), datetime.min.time()))
        stats = {
            "saved": rows.count(),
            "enabled": rows.filter(enabled=True).count(),
            "nextRunAt": rows.filter(enabled=True, next_run_at__isnull=False).aggregate(value=Min("next_run_at"))["value"],
            "todayRuns": WorkAutomationRun.objects.filter(
                organization=organization, creator=request.user, started_at__gte=day_start,
            ).count(),
        }
        stats["nextRunAt"] = stats["nextRunAt"].isoformat() if stats["nextRunAt"] else None
        return Response({"ok": True, "count": stats["saved"], "stats": stats, "results": [_work_automation_payload(row) for row in rows]})
    try:
        values = _validate_work_automation(request.user, request.data)
    except ValueError as exc:
        return Response({"ok": False, "detail": str(exc)}, status=400)
    from .automation_scheduler import configure_next_run
    row = WorkAutomation(organization=organization, creator=request.user, **values)
    try:
        configure_next_run(row)
    except ValueError as exc:
        return Response({"ok": False, "detail": str(exc)}, status=400)
    row.save()
    return Response({"ok": True, "automation": _work_automation_payload(row)}, status=201)


@api_view(["PATCH", "DELETE"])
@permission_classes([IsAuthenticated])
def work_automation_detail(request, automation_id: int):
    organization = current_organization(request.user)
    row = get_object_or_404(
        WorkAutomation, id=automation_id, organization=organization, creator=request.user,
    )
    if request.method == "DELETE":
        row.delete()
        return Response({"ok": True})
    try:
        values = _validate_work_automation(request.user, request.data, row)
    except ValueError as exc:
        return Response({"ok": False, "detail": str(exc)}, status=400)
    for key, value in values.items():
        setattr(row, key, value)
    from .automation_scheduler import configure_next_run
    try:
        configure_next_run(row)
    except ValueError as exc:
        return Response({"ok": False, "detail": str(exc)}, status=400)
    row.save()
    return Response({"ok": True, "automation": _work_automation_payload(row)})


@api_view(["PATCH"])
@permission_classes([IsAuthenticated])
def work_task_detail(request, trace_id: str):
    row = get_object_or_404(WorkTask, sender=request.user, trace_id=trace_id)
    status_value = str(request.data.get("status") or row.status)
    if status_value not in WorkTask.Status.values:
        return Response({"ok": False, "detail": "任务状态无效。"}, status=400)
    row.status = status_value
    if "progress" in request.data:
        row.progress = min(max(int(request.data["progress"]), 0), 100)
    if "sopId" in request.data:
        row.sop_id = str(request.data["sopId"] or "")[:128]
    if isinstance(request.data.get("timeline"), list):
        row.timeline = request.data["timeline"]
    if "notificationStatus" in request.data:
        row.notification_status = str(request.data["notificationStatus"])[:32]
    if request.data.get("notificationRecordId"):
        row.notification_record_id = int(request.data["notificationRecordId"])
    row.save()
    parameters = request.data.get("parameters")
    result_data = request.data.get("resultData")
    if isinstance(parameters, dict) or isinstance(result_data, dict):
        _generate_work_task_artifacts(
            row,
            parameters if isinstance(parameters, dict) else {},
            result_data if isinstance(result_data, dict) else {},
        )
    return Response({"ok": True, "task": _work_task_payload(row)})


def _visible_work_task(request, trace_id: str):
    return get_object_or_404(
        WorkTask.objects.select_related("sender", "sender__settings").prefetch_related("assignees__settings", "artifacts").filter(
            Q(sender=request.user) | Q(assignees=request.user),
        ).distinct(),
        trace_id=trace_id,
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def work_task_artifact_preview(request, trace_id: str, artifact_id: int):
    task = _visible_work_task(request, trace_id)
    artifact = get_object_or_404(task.artifacts, id=artifact_id)
    if artifact.kind == WorkTaskArtifact.Kind.XLSX:
        return Response({"ok": False, "detail": "Excel 文件请下载后查看。"}, status=400)
    return HttpResponse(bytes(artifact.content), content_type=artifact.content_type)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def work_task_artifact_download(request, trace_id: str, artifact_id: int):
    task = _visible_work_task(request, trace_id)
    artifact = get_object_or_404(task.artifacts, id=artifact_id)
    response = HttpResponse(bytes(artifact.content), content_type=artifact.content_type)
    suffix = Path(artifact.filename).suffix if re.fullmatch(r"\.[A-Za-z0-9]+", Path(artifact.filename).suffix) else ""
    ascii_stem = re.sub(r"[^A-Za-z0-9_-]+", "-", Path(artifact.filename).stem).strip("-")
    ascii_name = f"{ascii_stem or f'artifact-{artifact.id}'}{suffix}"
    response["Content-Disposition"] = (
        f'attachment; filename="{ascii_name}"; filename*=UTF-8\'\'{quote(artifact.filename)}'
    )
    response["Content-Length"] = artifact.size
    return response


def _attachment_meta(items: list[dict]) -> list[dict]:
    return attachment_public_meta(items)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def agent_models(request):
    """返回网关当前可用的对话 / 生图模型(来自 /v1/models)。"""
    from apps.council import images as image_svc

    return Response(image_svc.list_gateway_models())


@api_view(["POST"])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser, JSONParser])
def agent_chat(request):
    """对话 Agent:按登录用户隔离会话与 MCP 配置。"""
    message = str(request.data.get("message") or "").strip()

    try:
        attachments = process_uploaded_files(
            request.FILES.getlist("files"),
            request.user.id,
        )
    except ValueError as exc:
        return Response({"ok": False, "error": str(exc)}, status=400)

    if not message and not attachments:
        return Response({"ok": False, "error": "消息或附件不能为空"}, status=400)

    conversation_id = request.data.get("conversation_id")
    if conversation_id:
        session = get_object_or_404(ChatSession, id=conversation_id, user=request.user)
    else:
        title_src = message or (attachments[0]["name"] if attachments else "新对话")
        session = ChatSession.objects.create(
            user=request.user,
            title=title_src[:40] or "新对话",
        )

    history = list(
        session.messages.order_by("-created_at", "-id")
        .values("role", "content")[:30]
    )
    history.reverse()

    user_meta = {"attachments": _attachment_meta(attachments)} if attachments else {}
    display = message
    if attachments:
        names = "、".join(a["name"] for a in attachments)
        display = f"{message}\n\n[附件: {names}]".strip() if message else f"[附件: {names}]"
    ChatMessage.objects.create(
        session=session,
        role="user",
        content=display,
        meta=user_meta,
    )

    skill_ids = request.data.get("skill_ids") or []
    if isinstance(skill_ids, str):
        skill_ids = [skill_ids]

    model = str(request.data.get("model") or "").strip() or None
    knowledge_mode = str(request.data.get("knowledge_mode") or "auto").strip().lower()
    if knowledge_mode not in {"auto", "none", "selected"}:
        knowledge_mode = "auto"
    raw_knowledge_ids = request.data.get("knowledge_base_ids") or []
    if isinstance(raw_knowledge_ids, str):
        raw_knowledge_ids = request.data.getlist("knowledge_base_ids") or [raw_knowledge_ids]
    knowledge_base_ids = []
    for raw_id in raw_knowledge_ids:
        try:
            knowledge_base_ids.append(int(raw_id))
        except (TypeError, ValueError):
            continue

    try:
        result = run_chat(
            message,
            history,
            user=request.user,
            skill_ids=skill_ids,
            attachments=attachments,
            model=model,
            knowledge_mode=knowledge_mode,
            knowledge_base_ids=knowledge_base_ids,
            session_key=str(session.id),
        )
    except Exception as exc:
        import logging
        logging.getLogger(__name__).exception("agent_chat failed")
        return Response({"ok": False, "error": str(exc)}, status=500)

    if result.get("ok") and result.get("reply"):
        ChatMessage.objects.create(
            session=session,
            role="assistant",
            content=result["reply"],
            meta={
                "llm": result.get("llm"),
                "llm_model": result.get("llm_model") or "",
                "knowledge_hit": result.get("knowledge_hit"),
                "mcp": result.get("mcp") or {},
                "refs": result.get("refs") or {},
                "skills": result.get("skills") or [],
                "attachments": result.get("attachments") or [],
                "nas_files": result.get("nas_files") or [],
            },
        )
        if session.title == "新对话":
            session.title = (message or attachments[0]["name"])[:40]
        session.save(update_fields=["title", "updated_at"])

    result["conversation_id"] = str(session.id)
    result["conversation_title"] = session.title
    result.setdefault("attachments", _attachment_meta(attachments))
    code = 200 if result.get("ok") else 400
    return Response(result, status=code)


@api_view(["GET"])
@permission_classes([AllowAny])
def agent_attachment(request, stored_id: str):
    """读取用户上传的附件(图片预览)。

    支持 Header Token 或 ?token= 查询参数(便于 <img src> 拉取)。
    """
    from rest_framework.authtoken.models import Token

    user = request.user if getattr(request.user, "is_authenticated", False) else None
    if user is None or not user.is_authenticated:
        raw = (request.query_params.get("token") or "").strip()
        if raw:
            row = Token.objects.filter(key=raw).select_related("user").first()
            user = row.user if row else None
    if user is None or not getattr(user, "is_authenticated", False):
        return Response({"ok": False, "error": "未登录"}, status=401)

    path = resolve_attachment_path(user.id, stored_id)
    if not path and _is_admin(user):
        path = resolve_attachment_path_any(stored_id)
    if not path:
        return Response({"ok": False, "error": "附件不存在"}, status=404)
    mime = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
    original = path.name.split("_", 1)[-1] if "_" in path.name else path.name
    return FileResponse(path.open("rb"), as_attachment=False, filename=original, content_type=mime)


def _session_payload(session: ChatSession, include_messages: bool = False) -> dict:
    owner = session.user
    payload = {
        "id": str(session.id),
        "title": session.title,
        "created_at": session.created_at.isoformat(),
        "updated_at": session.updated_at.isoformat(),
        "user_id": owner.id if owner else None,
        "username": owner.username if owner else "",
    }
    if include_messages:
        payload["messages"] = [
            {
                "id": row.id,
                "role": row.role,
                "content": row.content,
                "meta": row.meta,
                "created_at": row.created_at.isoformat(),
            }
            for row in session.messages.all()
        ]
    return payload


def _sessions_for_user(user, *, all_users: bool = False):
    qs = ChatSession.objects.select_related("user")
    if all_users and _is_admin(user):
        return qs.all()
    return qs.filter(user=user)


@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def chat_sessions(request):
    if request.method == "POST":
        title = str(request.data.get("title") or "新对话").strip()[:120]
        session = ChatSession.objects.create(
            user=request.user,
            title=title or "新对话",
        )
        return Response(_session_payload(session, include_messages=True), status=201)

    limit = min(max(int(request.query_params.get("limit", 50)), 1), 200)
    username = str(request.query_params.get("username") or "").strip()
    qs = _sessions_for_user(request.user, all_users=True)
    if username and _is_admin(request.user):
        qs = qs.filter(user__username=username)
    sessions = list(qs[:limit])
    return Response({
        "count": len(sessions),
        "is_admin": _is_admin(request.user),
        "results": [_session_payload(item) for item in sessions],
    })


@api_view(["GET", "DELETE"])
@permission_classes([IsAuthenticated])
def chat_session_detail(request, session_id):
    if _is_admin(request.user):
        session = get_object_or_404(ChatSession.objects.select_related("user"), id=session_id)
    else:
        session = get_object_or_404(
            ChatSession.objects.select_related("user"),
            id=session_id,
            user=request.user,
        )
    if request.method == "DELETE":
        # 管理员可删任意会话；普通用户只能删自己的
        session.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    return Response(_session_payload(session, include_messages=True))
